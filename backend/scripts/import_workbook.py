"""WO v4.22 (Phase 2D-3) — multi-source ICB operational ETL into icb_mes.

One-shot loader (Q-Ph2D-03): TRUNCATE icb_mes + reload from the REAL Production-Server
workbooks, then re-seed the Materials master data from the mockup. Writes ONLY icb_mes;
reference-reads icb_costings but never mutates the catalogue / faje / the Cost Calculator.

Sources:
  * ENTERPRISE PLANNING - 2026.xlsx / JOBS (PLANNED)   -> production_jobs + planning_slots (v4.21).
  * 01 - MRP 2026.xlsx / Material Requirement Planning  -> demand_lines (v4.22 re-point, §0.1).
  * 02 - Live Daily Count 2026.xlsx (6 category sheets) -> live_daily_count (v4.22, §0.2).
  * Book1 TRUCK REGISTER 2026.xlsx / JOBS & CHASSIS     -> chassis_register (v4.22, §0.3).
  * frontend mockup JSON -> mes_materials / stock_positions / suppliers (§0.6-A, v4.21).

MRP layout (resolved at build-time): left inventory side cols 0-8; then 5 day-blocks at
offsets 10/20/30/40/50 (ITEM CODE, MATERIAL, REQ QTY, JOB 1..JOB 6). Row 1 = day dates
(offset+1); ROW 3 = the REAL 5-digit job numbers per JOB column; row 5 = "JOB n" labels;
data row 6+. Each non-zero (item × job-col) cell -> a demand_line.

Run from repo root:
    python -m backend.scripts.import_workbook [--planning P] [--mrp P] [--ldc P] [--chassis P] [--today YYYY-MM-DD]
"""
import argparse
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parents[1]
_REPO = Path(__file__).resolve().parents[2]
_DATA = _REPO / "frontend" / "src" / "data"
_LATEST = _REPO / "latest documents"
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from sqlalchemy import text as sa_text  # noqa: E402

from app.database import (  # noqa: E402
    Branch, Customer, SapItemCode, SessionLocal,
)
from app.models.mes import (  # noqa: E402
    ChassisRegister, DemandLine, LiveDailyCount, MesMaterial, PlanningSlot,
    ProductionJob, StockPosition, Supplier,
)

DEFAULT_PLANNING = (_REPO.parent / "Burt Costing Model" / "ICB business process"
                    / "ENTERPRISE PLANNING - 2026.xlsx")
DEFAULT_MRP = _LATEST / "01 - MRP 2026.xlsx"
DEFAULT_LDC = _LATEST / "02 - Live Daily Count 2026.xlsx"
DEFAULT_CHASSIS = _LATEST / "Book1 TRUCK REGISTER 2026.xlsx"

# All icb_mes tables — one-shot TRUNCATE target (Q-Ph2D-03).
_MES_TABLES = [
    "production_jobs", "work_orders", "tasks", "sign_offs", "photos", "rework_tickets",
    "planning_slots", "planning_acks", "stock_counts", "discrepancies", "po_suggestions",
    "demand_lines", "mes_materials", "stock_positions", "suppliers",
    "live_daily_count", "chassis_register",
]

# JOBS sheet — PLANNED block column indices (0-based; header row 2, data row 3+).
_J_JOB, _J_CUST, _J_DESC, _J_CHASSIS_RX, _J_VIN = 10, 11, 12, 13, 14
_J_VACUUM, _J_ASSY_COMP, _J_INVOICED, _J_LEFT, _J_PRICE = 18, 22, 23, 24, 25

# MRP sheet — Material Requirement Planning layout.
_MRP_BLOCK_OFFSETS = [10, 20, 30, 40, 50]   # 5 day-blocks
_MRP_JOBS_PER_BLOCK = 6                       # JOB 1..JOB 6 at offset+3 .. offset+8
_MRP_DATE_ROW = 1                             # DAY n + date (date cell at offset+1)
_MRP_JOBNUM_ROW = 3                           # the REAL 5-digit job numbers per JOB column
_MRP_DATA_START = 6

# Live Daily Count — header-name -> field (case-insensitive, stripped). First-wins
# (the sheets carry a duplicate "ORDERED" + a SAP-comparison block we ignore).
_LDC_SHEETS = ["ALU MATERIAL", "STEEL MATERIAL", "TIMBER MATERIAL", "EPS MATERIAL", "PU MATERIAL", "COILS"]
_LDC_FIELDS = {
    "sap code": "sap_code", "item description": "description", "uom": "uom",
    "on hand": "on_hand", "rejected stock": "rejected_stock", "max stock": "max_stock",
    "top up": "top_up", "ordered": "ordered", "price": "price", "pricing": "price",
    "variance qty": "variance_qty", "variance value": "variance_value",
}

# Chassis JOBS & CHASSIS — 17 hoist columns (single header row 1; cols A-Q).
_CHASSIS_HOIST = [
    ("job_number", 0), ("customer_name", 1), ("telephone", 2), ("contact_person", 3),
    ("vehicle_id_no", 4), ("model", 5), ("make", 6), ("description", 7), ("submit_status", 8),
    ("date_received_1", 9), ("vcl_1", 10), ("date_left_1", 11), ("dcl_1", 12),
    ("date_received_2", 13), ("vcl_2", 14), ("date_left_2", 15), ("dcl_2", 16),
]
_CHASSIS_DATE_FIELDS = {"date_received_1", "date_left_1", "date_received_2", "date_left_2"}
_CHASSIS_MAXLEN = {
    "job_number": 32, "customer_name": 255, "telephone": 64, "contact_person": 255,
    "vehicle_id_no": 64, "model": 64, "make": 64, "description": 255, "submit_status": 64,
    "vcl_1": 64, "dcl_1": 64, "vcl_2": 64, "dcl_2": 64,
}


# ── value helpers ──────────────────────────────────────────────────────────────
def _is_dt(v):
    return isinstance(v, datetime)


def _aware(v):
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    return None


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _iso_week(d: date) -> str:
    y, w, _ = d.isocalendar()
    return f"{y}-W{w:02d}"


def _num(v):
    try:
        return float(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.upper() != "#N/A" else None


def _job_str(v):
    if isinstance(v, (int, float)):
        return str(int(v))
    return _str(v)


def _ddmmyyyy(v):
    """Parse a chassis date cell (DD/MM/YYYY string or a datetime) -> date, else None."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _str(v)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _json_safe(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def _date_iso(s):
    return date.fromisoformat(str(s)[:10]) if s else None


def _dt_iso(s):
    if not s:
        return None
    s = str(s).replace("Z", "+00:00")
    try:
        d = datetime.fromisoformat(s)
    except ValueError:
        d = datetime.fromisoformat(s[:10] + "T00:00:00+00:00")
    return d if d.tzinfo else d.replace(tzinfo=timezone.utc)


# ── DB helpers ───────────────────────────────────────────────────────────────
def _truncate(db):
    db.execute(sa_text(
        "TRUNCATE " + ", ".join(f"icb_mes.{t}" for t in _MES_TABLES)
        + " RESTART IDENTITY CASCADE"))


def _reset_sequences(db):
    for t in ("production_jobs", "planning_slots", "demand_lines", "mes_materials",
              "stock_positions", "suppliers", "live_daily_count", "chassis_register"):
        db.execute(sa_text(
            f"SELECT setval('icb_mes.{t}_id_seq', "
            f"(SELECT COALESCE(MAX(id), 1) FROM icb_mes.{t}))"))


def _ref_maps(db):
    customers = {}
    for cid, name in db.query(Customer.id, Customer.name).all():
        if name:
            customers.setdefault(name.strip().lower(), cid)
    sap = set()
    for (code,) in db.query(SapItemCode.item_code).all():
        if code:
            sap.add(str(code).strip().lower())
    return customers, sap


# ── 1. production_jobs from JOBS PLANNED (unchanged from v4.21) ───────────────
def _load_jobs(db, ws, today, cutoff, customers_map, report):
    jhb = db.query(Branch).filter_by(code="JHB").first()
    branch_id = jhb.id if jhb else None
    jobs = []
    seen = set()
    skipped = matched_cust = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r is None or len(r) <= _J_LEFT:
            continue
        job_number = _job_str(r[_J_JOB])
        if not job_number or job_number.upper() == "JOB" or job_number in seen:
            continue
        seen.add(job_number)
        inv, left = r[_J_INVOICED], r[_J_LEFT]
        invoiced, left_dt = _is_dt(inv), _is_dt(left)
        if invoiced and left_dt and left.date() < cutoff:
            skipped += 1
            continue
        vacuum = r[_J_VACUUM] if _is_dt(r[_J_VACUUM]) else None
        if left_dt:
            status = "completed"
        elif vacuum is not None:
            status = "in_production"
        else:
            status = "planning"
        cust = _str(r[_J_CUST])
        if cust and cust.strip().lower() in customers_map:
            matched_cust += 1
        vin = _str(r[_J_VIN])
        pj = ProductionJob(
            calculation_record_id=None, source="workbook", branch_id=branch_id,
            job_number=job_number, status=status,
            customer_name=cust, description=_str(r[_J_DESC]), selling_zar=_num(r[_J_PRICE]),
            chassis_received_at=_aware(r[_J_CHASSIS_RX]),
            chassis_data_json=(json.dumps({"vin": vin}) if vin else None),
            planned_start_date=_aware(vacuum),
            completed_at=(_aware(left) if left_dt else _aware(r[_J_ASSY_COMP])),
        )
        db.add(pj)
        jobs.append((job_number, pj, vacuum, status))
    db.flush()
    report["production_jobs"] = len(jobs)
    report["jobs_skipped_30d"] = skipped
    report["customer_name_match"] = f"{matched_cust}/{len(jobs)} match icb_costings.customers"
    return jobs


# ── 2. planning_slots (MVP derivation, unchanged from v4.21) ──────────────────
def _load_slots(db, jobs, report):
    by_week = defaultdict(list)
    for job_number, pj, vacuum, status in jobs:
        if status == "completed" or vacuum is None:
            continue
        by_week[_monday(vacuum.date())].append(pj)
    n = 0
    for monday, pjs in by_week.items():
        for pos, pj in enumerate(sorted(pjs, key=lambda p: p.job_number), start=1):
            db.add(PlanningSlot(
                production_job_id=pj.id, week=monday,
                bay=f"Bay-{pos}", lane="vacuum", slot_position=pos, status="scheduled"))
            n += 1
    report["planning_slots"] = n
    report["planning_unscheduled_pool"] = sum(1 for *_, s in jobs if s == "planning")


# ── 3. demand_lines from MRP Material Requirement Planning (v4.22 re-point) ───
def _load_demand_mrp(db, ws, jobs, sap_set, report):
    pj_by_job = {jn: pj for (jn, pj, vac, st) in jobs}
    head = list(ws.iter_rows(min_row=1, max_row=_MRP_JOBNUM_ROW, values_only=True))
    date_row = head[_MRP_DATE_ROW - 1] if len(head) >= _MRP_DATE_ROW else ()
    jobnum_row = head[_MRP_JOBNUM_ROW - 1] if len(head) >= _MRP_JOBNUM_ROW else ()

    job_at, col_to_off = {}, {}
    for off in _MRP_BLOCK_OFFSETS:
        for k in range(_MRP_JOBS_PER_BLOCK):
            c = off + 3 + k
            col_to_off[c] = off
            jn = _job_str(jobnum_row[c]) if c < len(jobnum_row) else None
            if jn and jn.upper() != "ENTER JOB NUMBER":
                job_at[c] = jn
    block_date = {}
    for off in _MRP_BLOCK_OFFSETS:
        dc = off + 1
        dv = date_row[dc] if dc < len(date_row) else None
        block_date[off] = dv.date() if isinstance(dv, datetime) else None

    n = 0
    codes_seen, codes_hit, jobs_seen, matched = set(), set(), set(), set()
    for row in ws.iter_rows(min_row=_MRP_DATA_START, values_only=True):
        sap = _str(row[0]) if row else None
        if not sap:
            continue
        codes_seen.add(sap.lower())
        if sap.lower() in sap_set:
            codes_hit.add(sap.lower())
        for c, jn in job_at.items():
            qty = _num(row[c]) if c < len(row) else None
            if not qty or qty <= 0:
                continue
            d = block_date.get(col_to_off[c])
            pj = pj_by_job.get(jn)
            jobs_seen.add(jn)
            if pj:
                matched.add(jn)
            db.add(DemandLine(
                sap_code=sap, qty=qty, job_ref=jn,
                production_job_id=(pj.id if pj else None),
                need_by=d, week_bucket=(_iso_week(_monday(d)) if d else None)))
            n += 1
    report["demand_lines"] = n
    report["demand_active_job_columns"] = len(job_at)
    report["demand_jobs"] = f"{len(jobs_seen)} distinct MRP jobs; {len(matched)} matched to production_jobs"
    hit = (100 * len(codes_hit) / len(codes_seen)) if codes_seen else 0
    report["sap_enrichment_hit_rate"] = (
        f"{len(codes_hit)}/{len(codes_seen)} distinct ITEM CODEs match "
        f"icb_costings.sap_item_codes ({hit:.0f}%)")


# ── 4. live_daily_count from the 6 category sheets (by header name) ───────────
def _load_live_daily_count(db, wb, report):
    per_cat, total = {}, 0
    for sheet in _LDC_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        category = sheet.replace(" MATERIAL", "").strip()
        header_idx, colmap = None, {}
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
            c0 = _str(row[0]) if row else None
            if c0 and c0.lower() == "sap code":
                header_idx = ridx
                for ci, cell in enumerate(row):
                    name = _str(cell)
                    field = _LDC_FIELDS.get(name.lower()) if name else None
                    if field and field not in colmap:     # first-wins (dup ORDERED)
                        colmap[field] = ci
                break
        if header_idx is None or "sap_code" not in colmap:
            per_cat[category] = 0
            continue

        def cell(row, field):
            ci = colmap.get(field)
            return row[ci] if ci is not None and ci < len(row) else None

        n = 0
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            sap = _str(cell(row, "sap_code"))
            if not sap:
                continue
            db.add(LiveDailyCount(
                sap_code=sap[:64], description=_str(cell(row, "description")),
                uom=_str(cell(row, "uom")), category=category,
                on_hand=_num(cell(row, "on_hand")), rejected_stock=_num(cell(row, "rejected_stock")),
                max_stock=_num(cell(row, "max_stock")), top_up=_num(cell(row, "top_up")),
                ordered=_num(cell(row, "ordered")), price=_num(cell(row, "price")),
                variance_qty=_num(cell(row, "variance_qty")),
                variance_value=_num(cell(row, "variance_value"))))
            n += 1
        per_cat[category] = n
        total += n
    report["live_daily_count"] = total
    report["live_daily_count_by_category"] = per_cat


# ── 5. chassis_register from JOBS & CHASSIS (17 hoist cols + raw_row_json) ────
def _load_chassis(db, ws, report):
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    keys = [( _str(h).strip() if _str(h) else f"col{ci}") for ci, h in enumerate(header)]
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        vals = {}
        for field, ci in _CHASSIS_HOIST:
            c = row[ci] if ci < len(row) else None
            if field in _CHASSIS_DATE_FIELDS:
                vals[field] = _ddmmyyyy(c)
            else:
                s = _str(c)
                vals[field] = s[:_CHASSIS_MAXLEN[field]] if s and field in _CHASSIS_MAXLEN else s
        raw = {}
        for ci, c in enumerate(row):
            if c is None:
                continue
            key = keys[ci] if ci < len(keys) and keys[ci] else f"col{ci}"
            raw[key] = _json_safe(c)
        db.add(ChassisRegister(raw_row_json=raw, **vals))
        n += 1
    report["chassis_register"] = n


# ── 6. Materials/Buying/Stores re-seed from the mockup (§0.6-A, unchanged) ────
def _reseed_materials(db, report):
    mats = json.loads((_DATA / "icb_materials_data.json").read_text(encoding="utf-8"))
    for m in mats.get("materials", []):
        db.add(MesMaterial(
            sap_code=m.get("sap_code"), description=m.get("description"),
            supplier=m.get("supplier"), lead_days=m.get("lead_days"),
            last_price=m.get("last_price"), abc_class=m.get("abc_class"), dept=m.get("dept")))
    for s in mats.get("stock_positions", []):
        db.add(StockPosition(
            sap_code=s.get("sap_code"), sap_stock=s.get("sap_stock"),
            allocated=s.get("allocated"), free=s.get("free"),
            open_po_qty=s.get("open_po_qty"), open_po_eta=_date_iso(s.get("open_po_eta")),
            last_refreshed=_dt_iso(s.get("last_refreshed"))))
    for sup in mats.get("suppliers", []):
        db.add(Supplier(
            name=sup.get("name"), contact_person=sup.get("contact_person"),
            payment_terms=sup.get("payment_terms"), phone=sup.get("phone")))
    report["mes_materials"] = len(mats.get("materials", []))
    report["stock_positions"] = len(mats.get("stock_positions", []))
    report["suppliers"] = len(mats.get("suppliers", []))


def run(planning_wb: Path, mrp_wb: Path, ldc_wb: Path, chassis_wb: Path, today: date) -> dict:
    cutoff = today - timedelta(days=30)
    report = {"today": today.isoformat(), "active_cutoff": cutoff.isoformat(),
              "planning_wb": str(planning_wb), "mrp_wb": str(mrp_wb),
              "ldc_wb": str(ldc_wb), "chassis_wb": str(chassis_wb)}
    wb_plan = openpyxl.load_workbook(planning_wb, read_only=True, data_only=True)
    wb_mrp = openpyxl.load_workbook(mrp_wb, read_only=True, data_only=True)
    wb_ldc = openpyxl.load_workbook(ldc_wb, read_only=True, data_only=True)
    wb_chassis = openpyxl.load_workbook(chassis_wb, read_only=True, data_only=True)
    db = SessionLocal()
    try:
        _truncate(db)
        customers_map, sap_set = _ref_maps(db)
        jobs = _load_jobs(db, wb_plan["JOBS"], today, cutoff, customers_map, report)
        _load_slots(db, jobs, report)
        _load_demand_mrp(db, wb_mrp["Material Requirement Planning"], jobs, sap_set, report)
        _load_live_daily_count(db, wb_ldc, report)
        _load_chassis(db, wb_chassis["JOBS & CHASSIS"], report)
        _reseed_materials(db, report)
        db.commit()
        _reset_sequences(db)
        db.commit()
        return report
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        for w in (wb_plan, wb_mrp, wb_ldc, wb_chassis):
            w.close()


def main():
    ap = argparse.ArgumentParser(description="WO v4.22 — multi-source ICB operational ETL.")
    ap.add_argument("--planning", default=str(DEFAULT_PLANNING), help="ENTERPRISE PLANNING workbook")
    ap.add_argument("--mrp", default=str(DEFAULT_MRP), help="01 - MRP 2026.xlsx")
    ap.add_argument("--ldc", default=str(DEFAULT_LDC), help="02 - Live Daily Count 2026.xlsx")
    ap.add_argument("--chassis", default=str(DEFAULT_CHASSIS), help="Book1 TRUCK REGISTER 2026.xlsx")
    ap.add_argument("--today", default=None, help="YYYY-MM-DD override for the active-job cutoff")
    args = ap.parse_args()
    paths = {"planning": Path(args.planning), "mrp": Path(args.mrp),
             "ldc": Path(args.ldc), "chassis": Path(args.chassis)}
    for label, p in paths.items():
        if not p.exists():
            sys.exit(f"[import_workbook] {label} workbook not found: {p}")
    today = date.fromisoformat(args.today) if args.today else date.today()
    report = run(paths["planning"], paths["mrp"], paths["ldc"], paths["chassis"], today)
    print("\n[import_workbook] Complete. Load report:")
    for k, v in report.items():
        if not k.startswith("_"):
            print(f"  {k:<28} {v}")


if __name__ == "__main__":
    main()
