"""WO v4.23 — load the icb_sap (SAP-mock) schema from the Inventory + PRICE workbooks.

UPSERT + soft-delete (WO v4.27 §3.6, replacing the v4.23 one-shot TRUNCATE+RELOAD): items
absent from the import are flagged validFor='N', returning items UPSERTed back to 'Y' — nothing
is physically deleted, so the demand_lines->OITM FK (landed in 0011) can never be cascade-
violated (ADR 0013 forward pattern). Writes ONLY icb_sap (read-only from app code). Sources:
  * 04 - Inventory 2026.xlsx / Sheet1 — primary: OWHS (Whse row 2), OITM + OITW (rows 3+).
    Header row 1: Item No. | Item Description | Stock UoM | In Stock | Committed | Ordered |
    Available | Item Price | Total | Confirmed.
  * PRICE 2017 MARCH (08 April 2026).xlsx / Last P.P — enrichment: U_ItemGroup,
    U_Manufacturer, U_LastEvaluatedPrice, MinLevel (UPDATE OITM by ItemCode; PRICE-only
    items are reported as orphans).
  * ItmsGrpCod derived from the ItemCode prefix (CON=1, GRP=2, ... others=99).

Run from repo root:
    python -m backend.scripts.import_inventory_to_sap_mock [--inventory P] [--price P]
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parents[1]
_REPO = Path(__file__).resolve().parents[2]
_LATEST = _REPO / "latest documents"
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from sqlalchemy import text as sa_text  # noqa: E402
from sqlalchemy.dialects.postgresql import insert as pg_insert  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.models.sap import OITM, OITW, OWHS  # noqa: E402

DEFAULT_INVENTORY = _LATEST / "04 - Inventory 2026.xlsx"
DEFAULT_PRICE = (_REPO.parent / "Burt Costing Model" / "Nadie Costings"
                 / "PRICE 2017 MARCH (08 April 2026).xlsx")
WHS_CODE = "HEIDEL"
WHS_NAME = "Heidelberg (JHB main)"

# ItmsGrpCod from the 3-letter ItemCode prefix (WO §3.2; ADR 0013). others -> 99.
_GRP_MAP = {"CON": 1, "GRP": 2, "STE": 3, "TRA": 4, "FOA": 5, "BUY": 6, "MIS": 7, "CUT": 8,
            "TIM": 9, "BOL": 10, "ELE": 11, "DOR": 12, "PAI": 13, "STP": 14, "CHE": 15, "ICB": 16}

# Inventory Sheet1 column indices (header row 1; data row 3+).
_I_CODE, _I_NAME, _I_UOM, _I_INSTOCK, _I_COMMITTED, _I_ORDERED, _I_PRICE = 0, 1, 2, 3, 4, 5, 7


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.upper() != "#N/A" else None


def _num(v):
    try:
        return float(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _grp_cod(code):
    m = re.match(r"[A-Za-z]+", code or "")
    return _GRP_MAP.get((m.group()[:3].upper() if m else ""), 99)


def _upsert(db, model, rows, index_elements, update_cols):
    """Chunked INSERT ... ON CONFLICT (index_elements) DO UPDATE (Postgres UPSERT)."""
    for i in range(0, len(rows), 1000):
        chunk = rows[i:i + 1000]
        stmt = pg_insert(model).values(chunk)
        db.execute(stmt.on_conflict_do_update(
            index_elements=index_elements,
            set_={c: getattr(stmt.excluded, c) for c in update_cols},
        ))


def _soft_delete_absent_oitm(db, incoming_codes) -> int:
    """Flag OITM rows absent from `incoming_codes` as validFor='N' (NO physical delete → the
    demand_lines->OITM FK is never violated). Temp-table anti-join: fast at ~5.5k codes, and
    correct for an empty incoming set (unlike a literal `NOT IN ()`)."""
    db.execute(sa_text("DROP TABLE IF EXISTS _incoming_codes"))
    db.execute(sa_text("CREATE TEMP TABLE _incoming_codes (code varchar PRIMARY KEY) ON COMMIT DROP"))
    if incoming_codes:
        db.execute(sa_text("INSERT INTO _incoming_codes (code) VALUES (:c) ON CONFLICT DO NOTHING"),
                   [{"c": c} for c in incoming_codes])
    res = db.execute(sa_text(
        'UPDATE icb_sap."OITM" SET "validFor" = \'N\' '
        'WHERE "ItemCode" NOT IN (SELECT code FROM _incoming_codes) AND "validFor" <> \'N\''))
    db.execute(sa_text("DROP TABLE IF EXISTS _incoming_codes"))
    return res.rowcount or 0


def _load_inventory(db, inventory, report):
    wb = openpyxl.load_workbook(inventory, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    oitm, oitw = [], []
    seen = set()
    onhand_pos = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        code = _str(r[_I_CODE]) if r else None
        if not code or code.lower().startswith("whse"):
            continue
        code = code[:64]
        if code in seen:
            continue
        seen.add(code)
        price = _num(r[_I_PRICE]) if len(r) > _I_PRICE else None
        on_hand = _num(r[_I_INSTOCK]) or 0
        oitm.append({"ItemCode": code, "ItemName": _str(r[_I_NAME]), "InvntryUom": _str(r[_I_UOM]),
                     "ItmsGrpCod": _grp_cod(code), "U_LastPurchasePrice": price, "validFor": "Y"})
        oitw.append({"ItemCode": code, "WhsCode": WHS_CODE, "OnHand": on_hand,
                     "IsCommited": _num(r[_I_COMMITTED]) or 0, "OnOrder": _num(r[_I_ORDERED]) or 0,
                     "AvgPrice": price})
        if on_hand > 0:
            onhand_pos += 1
    wb.close()
    # WO v4.27 §3.6 — UPSERT + soft-delete (NOT TRUNCATE+RELOAD): soft-delete first (absent ->
    # validFor='N'), then UPSERT returning/new items back to 'Y'. No physical delete → the
    # demand_lines->OITM FK can never be cascade-violated (ADR 0013).
    soft_deleted = _soft_delete_absent_oitm(db, [o["ItemCode"] for o in oitm])
    _upsert(db, OITM, oitm, ["ItemCode"],
            ["ItemName", "InvntryUom", "ItmsGrpCod", "U_LastPurchasePrice", "validFor"])
    db.flush()
    _upsert(db, OITW, oitw, ["ItemCode", "WhsCode"],
            ["OnHand", "IsCommited", "OnOrder", "AvgPrice"])   # "Available" is GENERATED
    db.flush()
    report["OITM"] = len(oitm)
    report["OITW"] = len(oitw)
    report["OITW_onhand_positive"] = onhand_pos
    report["OITM_soft_deleted"] = soft_deleted


def _enrich_from_price(db, price, report):
    wb = openpyxl.load_workbook(price, read_only=True, data_only=True)
    ws = wb["Last P.P"] if "Last P.P" in wb.sheetnames else wb[wb.sheetnames[0]]
    existing = {row[0] for row in db.execute(sa_text('SELECT "ItemCode" FROM icb_sap."OITM"')).all()}
    # The 'Last P.P' sheet's "Item No." / "Item Description" headers are SWAPPED — the
    # ItemCode (GRP-MPS-A-0001) sits under "Item Description". So detect the header row +
    # the group/mfr/evalprice/minlevel columns by their (correct) header names, but detect
    # the CODE column by value (the column whose values best overlap OITM.ItemCode).
    header_idx, col, sample = None, {}, []
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        low = [(_str(c) or "").lower() for c in row]
        if header_idx is None and (any("item no" in x for x in low) or any("item desc" in x for x in low)):
            header_idx = ridx
            for ci, x in enumerate(low):
                if "item group" in x:
                    col.setdefault("group", ci)
                elif "manufacturer" in x:
                    col.setdefault("mfr", ci)
                elif "last evaluated" in x:
                    col.setdefault("evalprice", ci)
                elif "minimum inventory" in x or ("min" in x and "level" in x):
                    col.setdefault("minlevel", ci)
            continue
        if header_idx is not None and any(c is not None for c in row):
            sample.append(row)
            if len(sample) >= 10:
                break
    code_ci, best = None, 0
    for ci in range(min((max(len(r) for r in sample) if sample else 0), 8)):
        hits = sum(1 for r in sample if ci < len(r) and ((_str(r[ci]) or "")[:64] in existing))
        if hits > best:
            best, code_ci = hits, ci

    enriched = orphan = price_rows = 0
    if header_idx is not None and code_ci is not None:
        updates, seen = [], set()
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            code = _str(row[code_ci]) if code_ci < len(row) else None
            if not code or code in seen:
                continue
            seen.add(code)
            price_rows += 1
            if code[:64] not in existing:
                orphan += 1
                continue

            def g(k):
                ci = col.get(k)
                return row[ci] if ci is not None and ci < len(row) else None

            updates.append({"c": code[:64], "g": _str(g("group")), "m": _str(g("mfr")),
                            "e": _num(g("evalprice")), "ml": _num(g("minlevel"))})
        if updates:
            db.execute(sa_text(
                'UPDATE icb_sap."OITM" SET "U_ItemGroup"=:g, "U_Manufacturer"=:m, '
                '"U_LastEvaluatedPrice"=:e, "MinLevel"=:ml WHERE "ItemCode"=:c'), updates)
            enriched = len(updates)
    wb.close()
    report["price_code_column"] = code_ci
    report["price_rows"] = price_rows
    report["oitm_enriched"] = enriched
    report["orphan_price_codes"] = orphan


def run(inventory: Path, price: Path) -> dict:
    report = {"inventory": str(inventory), "price": str(price)}
    db = SessionLocal()
    try:
        # WO v4.27 §3.6 — UPSERT the warehouse (no TRUNCATE; the demand_lines->OITM FK is live).
        ow = pg_insert(OWHS).values(WhsCode=WHS_CODE, WhsName=WHS_NAME, Inactive="N")
        db.execute(ow.on_conflict_do_update(
            index_elements=["WhsCode"],
            set_={"WhsName": ow.excluded.WhsName, "Inactive": ow.excluded.Inactive}))
        db.flush()
        report["OWHS"] = 1
        _load_inventory(db, inventory, report)
        _enrich_from_price(db, price, report)
        # Orphan reconciliation: demand_lines.sap_code not present in OITM (for the FK / v4.23 report).
        orphans = db.execute(sa_text(
            'SELECT count(DISTINCT dl.sap_code) FROM icb_mes.demand_lines dl '
            'LEFT JOIN icb_sap."OITM" o ON o."ItemCode" = dl.sap_code WHERE o."ItemCode" IS NULL')).scalar()
        total = db.execute(sa_text("SELECT count(DISTINCT sap_code) FROM icb_mes.demand_lines")).scalar()
        report["demand_sapcode_orphans"] = f"{orphans}/{total} distinct demand sap_codes NOT in OITM"
        db.commit()
        return report
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main():
    ap = argparse.ArgumentParser(description="WO v4.23 — load icb_sap from Inventory + PRICE.")
    ap.add_argument("--inventory", default=str(DEFAULT_INVENTORY))
    ap.add_argument("--price", default=str(DEFAULT_PRICE))
    args = ap.parse_args()
    inv, price = Path(args.inventory), Path(args.price)
    if not inv.exists():
        sys.exit(f"[sap-mock] inventory not found: {inv}")
    if not price.exists():
        sys.exit(f"[sap-mock] price workbook not found: {price}")
    from scripts._environment_guard import confirm_if_shared_db
    confirm_if_shared_db("import_inventory_to_sap_mock",
                         destroys="UPSERT + soft-delete the icb_sap mock (OITM/OITW/OWHS) from the workbooks.")
    report = run(inv, price)
    print("\n[import_inventory_to_sap_mock] Complete. Load report:")
    for k, v in report.items():
        print(f"  {k:<26} {v}")


if __name__ == "__main__":
    main()
