"""WO v4.24 — extract job 32735's ground truth from COSTING MODULE 2026.xlsx into a
committed JSON fixture (so the replay test is deterministic + needs no 23 MB workbook in CI).

Provenance for the fixture: run from repo root with the workbook present:
    python -m backend.app.spikes.v4_24.extract_fixture

Captures, for job 32735 (the workbook's live job), the Vacuum Materials slice:
  * specs      — resolved per-panel specs (VACUUM ORDERS AF/AG block = '2026 COSTINGS'!D*)
  * cutlist    — VACUUM ORDERS computed cut-list per panel (A-F + AB/AC), the geometry targets
  * bom_vacuum — 2026 BOM 'Vacuum Materials' section (the replay target) + section/grand totals
  * tables     — the per-panel description lookup tables (H/L/N/P/R 2:66) the BOM feed VLOOKUPs

READ-ONLY against the workbook; writes only the fixture JSON. Not part of the runtime app.
"""
import json
from pathlib import Path

import openpyxl

_REPO = Path(__file__).resolve().parents[4]
WORKBOOK = _REPO / "latest documents" / "COSTING MODULE 2026.xlsx"
FIXTURE = Path(__file__).resolve().parents[3] / "tests" / "spikes" / "fixtures" / "job_32735_vacuum.json"


def _v(x):
    """Normalise a cell value: '- -'/blank -> None; trim strings; keep numbers."""
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        return None if s in ("", "- -", "#N/A") else s
    return x


def _rows(ws, lo, hi):
    out = {}
    for ridx, row in enumerate(ws.iter_rows(min_row=lo, max_row=hi, values_only=True), start=lo):
        out[ridx] = row
    return out


def extract() -> dict:
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    # ── VACUUM ORDERS: spec block (AF/AG = tuple idx 31/32) + cut-list (A-F=0..5, AB/AC=27/28) ──
    vo = wb["VACUUM ORDERS"]
    r = _rows(vo, 1, 25)

    def af(rn):   # AF column (1-based 32 -> tuple idx 31)
        return _v(r[rn][31]) if rn in r and len(r[rn]) > 31 else None

    def ag(rn):
        return _v(r[rn][32]) if rn in r and len(r[rn]) > 32 else None

    specs = {
        "job": 32735,
        "body_type": af(20),
        "length_mm": af(4), "width_mm": af(5), "height_mm": af(6),
        "roof":      {"thickness": af(7),  "material": ag(7)},
        "sides":     {"thickness": af(8),  "material": ag(8)},
        "floor":     {"thickness": af(9),  "material": ag(9)},
        "front":     {"thickness": af(10), "material": ag(10)},
        "rear":      {"thickness": af(11), "material": ag(11)},
        "partition": {"thickness": af(12), "material": ag(12)},
        "frame_dims": {"AF13": af(13), "AF14": af(14), "AF15": af(15), "AF16": af(16)},
        "af_block_raw": {f"AF{n}": af(n) for n in range(4, 22)},
        "ag_block_raw": {f"AG{n}": ag(n) for n in range(4, 22)},
    }

    cutlist = []
    for rn in range(5, 21):
        row = r.get(rn)
        if not row:
            continue
        name = _v(row[0])
        if not name or name.lower() in ("panel",):
            continue
        cutlist.append({
            "row": rn, "panel": name,
            "thickness": _v(row[1]), "material": _v(row[2]),
            "width": _v(row[3]), "length": _v(row[4]), "qty": _v(row[5]),
            "bom_description": _v(row[27]) if len(row) > 27 else None,
            "bom_qty": _v(row[28]) if len(row) > 28 else None,
        })

    # ── 2026 BOM: Vacuum Materials section (replay target) ──
    bom = wb["2026 BOM"]
    bom_vacuum, section_total, grand_total, in_sec = [], None, None, False
    for ridx, row in enumerate(bom.iter_rows(min_row=1, max_row=80, max_col=5, values_only=True), start=1):
        a = _v(row[0])
        if a == "TOTAL COST OF SALE":
            grand_total = _v(row[3])
        if a == "Vacuum Materials":
            in_sec, section_total = True, _v(row[3])
            continue
        if in_sec:
            qty = _v(row[1]) if len(row) > 1 else None
            # a section header = label in A, a total in D, but no per-line qty in B -> stop
            if a and qty is None and _v(row[3]) is not None and a != "Material Description":
                break
            if a and a != "Material Description" and qty is not None:
                bom_vacuum.append({
                    "description": a, "qty": qty,
                    "code": _v(row[2]) if len(row) > 2 else None,
                    "unit_price": _v(row[3]) if len(row) > 3 else None,
                    "line_total": _v(row[4]) if len(row) > 4 else None,
                })

    # ── local description lookup tables on VACUUM ORDERS (H/L/N/P/R = cols 8/12/14/16/18) ──
    tables = {}
    for label, kcol in (("H", 8), ("L", 12), ("N", 14), ("P", 16), ("R", 18)):
        pairs = []
        for rn in range(2, 67):
            row = r.get(rn)
            if not row or len(row) <= kcol:
                continue
            key, val = _v(row[kcol - 1]), _v(row[kcol]) if len(row) > kcol else None
            if key is not None or val is not None:
                pairs.append([key, val])
        tables[label] = pairs

    wb.close()
    return {"specs": specs, "cutlist": cutlist, "bom_vacuum": bom_vacuum,
            "section_total": section_total, "grand_total": grand_total, "lookup_tables": tables}


def main():
    if not WORKBOOK.exists():
        raise SystemExit(f"workbook not found: {WORKBOOK}")
    data = extract()
    s = data["specs"]
    print("== specs (resolved, job 32735) ==")
    print(f"  body_type={s['body_type']}  L×W×H={s['length_mm']}×{s['width_mm']}×{s['height_mm']}")
    for p in ("roof", "sides", "floor", "front", "rear", "partition"):
        print(f"  {p:<10} thk={s[p]['thickness']}  mat={s[p]['material']}")
    print(f"  frame_dims={s['frame_dims']}")
    print(f"\n== cutlist ({len(data['cutlist'])} panels) ==")
    for c in data["cutlist"]:
        print(f"  r{c['row']:<2} {c['panel']:<16} thk={c['thickness']} mat={c['material']} "
              f"W={c['width']} L={c['length']} qty={c['qty']} | bom='{c['bom_description']}' bomqty={c['bom_qty']}")
    print(f"\n== bom_vacuum ({len(data['bom_vacuum'])} lines)  section_total={data['section_total']} grand_total={data['grand_total']} ==")
    for b in data["bom_vacuum"]:
        print(f"  {str(b['description'])[:34]:<36} qty={b['qty']} code={b['code']} unit={b['unit_price']} tot={b['line_total']}")
    print(f"\n== lookup tables (non-empty rows) ==")
    for k, v in data["lookup_tables"].items():
        nonempty = [p for p in v if p[0] is not None]
        print(f"  {k}2:{k}66 -> {len(nonempty)} mappings; sample {nonempty[:3]}")

    FIXTURE.parent.mkdir(parents=True, exist_ok=True)
    FIXTURE.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    print(f"\n[fixture written] {FIXTURE}")


if __name__ == "__main__":
    main()
