import importlib.util
import io
import os
import sys
import tempfile
import time
import uuid
from typing import Optional

from fastapi import Request, APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from ..database import (
    get_db, AdminSetting, TrailerType, Material, MaterialCategory,
    BillOfMaterial, BOMSection,
)
from ..deps import get_current_user, require_admin, user_can, _is_dev_mode
from ..services import (
    _resolve_bom_section, archive_trailer_template_binding,
    restore_orphan_for_trailer, resolve_report_template,
)
from ..templates_config import templates

router = APIRouter()

# ── Excel file patterns ───────────────────────────────────────────────────────
EXCEL_FILE_PATTERNS = {
    "costing":  {"label": "Costing workbook",  "patterns": ["GRP", "COSTING"], "exts": [".xlsx", ".xlsm", ".xls"]},
    "formulas": {"label": "Formulas workbook", "patterns": ["FORMULA"],        "exts": [".xls", ".xlsx", ".xlsm"]},
}


def _scan_excel_folder(folder: str) -> dict:
    result = {}
    files = []
    if folder and os.path.isdir(folder):
        try:
            files = os.listdir(folder)
        except Exception:
            files = []
    for role, spec in EXCEL_FILE_PATTERNS.items():
        match = None
        for f in files:
            up = f.upper()
            if any(p in up for p in spec["patterns"]) and any(up.endswith(e.upper()) for e in spec["exts"]):
                match = f
                break
        result[role] = {
            "label":    spec["label"],
            "found":    bool(match),
            "filename": match or "",
            "path":     os.path.join(folder, match) if match else "",
        }
    return result


def _apply_excel_folder_setting():
    from ..database import SessionLocal, AdminSetting as _AS
    db = SessionLocal()
    try:
        s = db.query(_AS).filter_by(key="excel_folder").first()
        folder = (s.value if s else "").strip()
    finally:
        db.close()
    if not folder:
        return
    scan = _scan_excel_folder(folder)
    from .. import excel_importer as _xi
    if scan["costing"]["found"]:
        _xi.DEFAULT_GRP_PATH = scan["costing"]["path"]
        os.environ["EXCEL_PATH"] = scan["costing"]["path"]
    if scan["formulas"]["found"]:
        _xi.DEFAULT_FORMULAS_PATH = scan["formulas"]["path"]
        os.environ["FORMULAS_XLS_PATH"] = scan["formulas"]["path"]


# ── Per-import workbook upload state ─────────────────────────────────────────
_UPLOAD_ROOT = os.path.join(tempfile.gettempdir(), "burtcost_uploads")
os.makedirs(_UPLOAD_ROOT, exist_ok=True)
_UPLOADS: dict = {}
_UPLOAD_TTL_SECONDS = 60 * 60 * 2


def _purge_old_uploads():
    now = time.time()
    for uid, info in list(_UPLOADS.items()):
        if now - info.get("created_at", 0) > _UPLOAD_TTL_SECONDS:
            try:
                import shutil
                shutil.rmtree(info.get("dir", ""), ignore_errors=True)
            finally:
                _UPLOADS.pop(uid, None)


def _save_upload(file: UploadFile, dest_path: str) -> int:
    size = 0
    with open(dest_path, "wb") as out:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            out.write(chunk)
    return size


def _resolve_upload_paths(upload_id: Optional[str]) -> tuple:
    if not upload_id:
        return None, None
    info = _UPLOADS.get(upload_id)
    if not info:
        raise HTTPException(status_code=404,
            detail="Upload session not found or expired. Please re-upload the workbooks.")
    return info["costing"], info.get("formulas")


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/api/admin/browse-folders")
async def api_browse_folders(request: Request, path: str = "", db: Session = Depends(get_db)):
    if not _is_dev_mode():
        raise HTTPException(status_code=403, detail="Folder browsing only available in local mode.")
    require_admin(request, db)

    target = (path or "").strip()
    if not target:
        target = os.path.join(os.path.expanduser("~"), "Documents")
        if not os.path.isdir(target):
            target = os.path.expanduser("~")

    if sys.platform.startswith("win") and target in ("", "\\"):
        import string
        drives = []
        for letter in string.ascii_uppercase:
            d = f"{letter}:\\"
            if os.path.isdir(d):
                drives.append({"name": d, "path": d})
        return {"path": "", "parent": None, "is_root": True, "dirs": drives,
                "scan": _scan_excel_folder("")}

    target = os.path.abspath(target)
    if not os.path.isdir(target):
        raise HTTPException(status_code=404, detail=f"Not a folder: {target}")

    try:
        entries = sorted(os.listdir(target), key=lambda s: s.lower())
    except PermissionError:
        raise HTTPException(status_code=403, detail="Permission denied")
    dirs = []
    for name in entries:
        if name.startswith("."):
            continue
        full = os.path.join(target, name)
        try:
            if os.path.isdir(full):
                dirs.append({"name": name, "path": full})
        except OSError:
            pass

    parent = os.path.dirname(target)
    if parent == target:
        parent = "" if sys.platform.startswith("win") else None

    return {
        "path": target,
        "parent": parent,
        "is_root": False,
        "dirs": dirs,
        "scan": _scan_excel_folder(target),
    }


@router.get("/api/admin/excel-folder")
async def api_get_excel_folder(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    s = db.query(AdminSetting).filter_by(key="excel_folder").first()
    folder = (s.value if s else "").strip()
    return {"folder": folder, "files": _scan_excel_folder(folder)}


@router.post("/api/admin/excel-folder")
async def api_set_excel_folder(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    body = await request.json()
    folder = (body.get("folder") or "").strip()
    if folder and not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Folder does not exist: {folder}")

    s = db.query(AdminSetting).filter_by(key="excel_folder").first()
    if s:
        s.value = folder
    else:
        s = AdminSetting(key="excel_folder", value=folder)
        db.add(s)
    db.commit()

    _apply_excel_folder_setting()
    scan = _scan_excel_folder(folder)
    all_found = folder and all(v["found"] for v in scan.values())
    return {"ok": True, "folder": folder, "files": scan, "all_found": bool(all_found)}


@router.get("/api/import/sheets")
async def list_excel_sheets(path: str, request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="File not found")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return {"sheets": wb.sheetnames}


@router.post("/api/import/preview")
async def import_preview(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    body = await request.json()
    excel_path = body.get("excel_path", "").strip()
    sheet_name = body.get("sheet_name", "").strip()
    if not os.path.isfile(excel_path):
        raise HTTPException(status_code=404, detail="Excel file not found")
    spec = importlib.util.spec_from_file_location(
        "parse_grp_sheet",
        os.path.join(os.path.dirname(__file__), "..", "..", "tools", "parse_grp_sheet.py"),
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    wb_data, wb_formulas = mod.load_workbooks(excel_path)
    if sheet_name not in wb_data.sheetnames:
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found")
    result = mod.parse_sheet(sheet_name, wb_data, wb_formulas)
    return result


@router.post("/api/import/execute")
async def import_execute(request: Request, db: Session = Depends(get_db)):
    from datetime import datetime, timezone
    require_admin(request, db)
    body = await request.json()

    parsed = body.get("parsed")
    trailer_name_override = body.get("trailer_name_override", "").strip()
    replace_existing = body.get("replace_existing", False)

    if not parsed:
        raise HTTPException(status_code=400, detail="No parsed data provided")

    trailer_name  = trailer_name_override or parsed["trailer_name"]
    defaults      = parsed.get("trailer_defaults", {})
    sections_data = parsed.get("sections", [])

    existing_tt = db.query(TrailerType).filter_by(name=trailer_name).first()
    if existing_tt and not replace_existing:
        raise HTTPException(
            status_code=409,
            detail=f"Trailer type '{trailer_name}' already exists. "
                   "Set replace_existing=true to delete and reimport.",
        )
    if existing_tt and replace_existing:
        archive_trailer_template_binding(existing_tt, db)
        db.query(BillOfMaterial).filter_by(trailer_type_id=existing_tt.id).delete()
        db.delete(existing_tt)
        db.flush()

    tt = TrailerType(
        name=trailer_name,
        description=f"Imported from sheet: {parsed.get('source_sheet', trailer_name)}",
        is_active=True,
        default_length=defaults.get("length"),
        default_width=defaults.get("width"),
        default_height=defaults.get("height"),
    )
    db.add(tt)
    db.flush()

    restored_orphan = restore_orphan_for_trailer(tt, db)
    items_created = 0

    for sort_idx, sect in enumerate(sections_data):
        sect_name  = sect["name"]
        multiplier = float(sect.get("multiplier", 1.0))

        bom_sec = db.query(BOMSection).filter_by(name=sect_name).first()
        if not bom_sec:
            bom_sec = BOMSection(name=sect_name, sort_order=sort_idx, multiplier=multiplier)
            db.add(bom_sec)
        else:
            bom_sec.multiplier = multiplier
        db.flush()

        cat = db.query(MaterialCategory).filter_by(name=sect_name).first()
        if not cat:
            cat = MaterialCategory(name=sect_name)
            db.add(cat)
            db.flush()

        for item_sort, item in enumerate(sect.get("items", [])):
            mat_name = item["material_name"]
            price    = float(item.get("price_per_unit") or 0)
            uom      = item.get("unit_of_measure", "each")
            formula  = item.get("formula_expression") or "1"
            waste    = float(item.get("waste_percentage") or 0)
            notes    = item.get("notes", "")

            mat = (db.query(Material)
                   .filter_by(name=mat_name, category_id=cat.id, is_active=True)
                   .first())
            if not mat:
                mat = Material(
                    name=mat_name,
                    category_id=cat.id,
                    price_per_unit=price,
                    unit_of_measure=uom,
                    is_active=True,
                    last_updated=datetime.now(timezone.utc),
                )
                db.add(mat)
                db.flush()

            bom = BillOfMaterial(
                trailer_type_id=tt.id,
                material_id=mat.id,
                formula_expression=formula,
                waste_percentage=waste,
                notes=notes,
                bom_section=sect_name,
                bom_section_id=_resolve_bom_section(db, sect_name),
                sort_order=item_sort,
            )
            db.add(bom)
            items_created += 1

    db.commit()
    restored_info = None
    if restored_orphan:
        resolved = resolve_report_template(tt)
        restored_info = {
            "group_name":    tt.group.name if tt.group else None,
            "template_name": resolved.name if resolved else None,
        }
    return {
        "ok": True,
        "trailer_type_id":  tt.id,
        "trailer_name":     trailer_name,
        "sections_imported": len(sections_data),
        "items_imported":    items_created,
        "restored_template_binding": restored_info,
    }


@router.get("/admin/import", response_class=HTMLResponse)
async def admin_import_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login")
    if not user_can(user, "menu.import", db):
        raise HTTPException(status_code=403, detail="Permission denied: menu.import")
    return templates.TemplateResponse("admin_import.html", {
        "request": request, "user": user, "sheets": [],
    })


@router.get("/admin/import/grp", response_class=HTMLResponse)
async def admin_import_grp_page(request: Request, db: Session = Depends(get_db)):
    """Phase C/D UI — preview + commit a complex GRP sheet (RIGID DRY
    FREIGHT-style trailers with category gating + chained formulas)."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login")
    if not user_can(user, "menu.import", db):
        raise HTTPException(status_code=403, detail="Permission denied: menu.import")
    return templates.TemplateResponse("admin_import_grp.html", {
        "request": request, "user": user,
    })


@router.get("/api/import/sheets")
async def api_list_sheets(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    from app.excel_importer import list_sheets
    return list_sheets()


@router.post("/api/import/upload")
async def api_import_upload(
    request: Request,
    costing: UploadFile = File(...),
    formulas: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    require_admin(request, db)
    _purge_old_uploads()

    upload_id = uuid.uuid4().hex
    upload_dir = os.path.join(_UPLOAD_ROOT, upload_id)
    os.makedirs(upload_dir, exist_ok=True)

    costing_ext = os.path.splitext(costing.filename or "")[1].lower() or ".xlsx"
    costing_path = os.path.join(upload_dir, f"costing{costing_ext}")
    costing_size = _save_upload(costing, costing_path)

    formulas_path = None
    formulas_size = 0
    if formulas is not None and formulas.filename:
        formulas_ext = os.path.splitext(formulas.filename)[1].lower() or ".xls"
        formulas_path = os.path.join(upload_dir, f"formulas{formulas_ext}")
        formulas_size = _save_upload(formulas, formulas_path)

    try:
        from app.excel_importer import list_sheets
        sheets = list_sheets(grp_path=costing_path)
    except Exception as e:
        import shutil
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise HTTPException(status_code=400,
            detail=f"Could not read costing workbook: {type(e).__name__}: {e}")

    _UPLOADS[upload_id] = {
        "dir": upload_dir,
        "costing": costing_path,
        "formulas": formulas_path,
        "created_at": time.time(),
        "costing_name": costing.filename,
        "formulas_name": formulas.filename if formulas else None,
    }

    return {
        "upload_id": upload_id,
        "sheets": sheets,
        "costing_name": costing.filename,
        "costing_size": costing_size,
        "formulas_name": formulas.filename if formulas else None,
        "formulas_size": formulas_size,
    }


@router.post("/api/import/reuse")
async def api_import_reuse(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    body = await request.json()
    upload_id = (body.get("upload_id") or "").strip()
    info = _UPLOADS.get(upload_id)
    if not info:
        raise HTTPException(status_code=404,
            detail="Upload session not found or expired. Please pick the files again.")
    if not os.path.exists(info["costing"]):
        _UPLOADS.pop(upload_id, None)
        raise HTTPException(status_code=410,
            detail="Cached files are no longer available. Please pick the files again.")
    try:
        from app.excel_importer import list_sheets
        sheets = list_sheets(grp_path=info["costing"])
    except Exception as e:
        raise HTTPException(status_code=400,
            detail=f"Could not read cached workbook: {type(e).__name__}: {e}")
    return {
        "upload_id": upload_id,
        "sheets": sheets,
        "costing_name": info.get("costing_name") or os.path.basename(info["costing"]),
        "costing_size": os.path.getsize(info["costing"]),
        "formulas_name": info.get("formulas_name"),
        "formulas_size": (os.path.getsize(info["formulas"]) if info.get("formulas") and os.path.exists(info["formulas"]) else 0),
    }


@router.get("/admin/import/example-screenshot.png")
async def admin_import_example_screenshot(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    bundled = os.path.join(os.path.dirname(__file__), "..", "static", "trailer-import-example.png")
    legacy  = r"C:\Users\micge\Documents\Burt Costing Model\Trailer Import example.png"
    path = bundled if os.path.exists(bundled) else legacy
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Screenshot not found")
    return FileResponse(path, media_type="image/png")


@router.get("/admin/import/sample-template.xlsx")
async def admin_import_sample_template(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.excel_importer import DEFAULT_GRP_PATH

    try:
        src = load_workbook(DEFAULT_GRP_PATH, data_only=True, read_only=True)
        example_name = next((n for n in src.sheetnames if n.strip().upper() == "EXAMPLE"), None)
        if example_name:
            src_ws = src[example_name]
            out = Workbook()
            out_ws = out.active
            out_ws.title = "EXAMPLE"
            for row in src_ws.iter_rows(values_only=True):
                out_ws.append(list(row))
            buf = io.BytesIO()
            out.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="trailer_import_example.xlsx"'},
            )
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = "SAMPLE TRAILER"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    note_font = Font(italic=True, color="808080", size=9)
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    headers = ["A (ignored)", "Section", "Item Name", "Qty", "Unit",
               "Unit Price", "Total / Ratio", "H (ignored)", "Active (1=on,0=off)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ratios = [
        ("Default ratio 1", 0.05),
        ("Default ratio 2", 0.10),
        ("Default ratio 3", 0.15),
    ]
    for i, (label, val) in enumerate(ratios):
        ws.cell(row=8 + i, column=1, value=label).font = note_font
        ws.cell(row=8 + i, column=7, value=val)

    ws.cell(row=20, column=2, value="FLOOR").font = bold
    ws.cell(row=20, column=2).fill = section_fill
    ws.cell(row=21, column=3, value="2*300 INT LAMINATION")
    ws.cell(row=21, column=4, value=2)
    ws.cell(row=21, column=5, value="EA")
    ws.cell(row=21, column=6, value=150.00)
    ws.cell(row=21, column=7, value=300.00)
    ws.cell(row=21, column=9, value=1)

    ws.cell(row=22, column=3, value="ANTI-SKID FINAL COAT")
    ws.cell(row=22, column=4, value=1)
    ws.cell(row=22, column=5, value="L")
    ws.cell(row=22, column=6, value=85.50)
    ws.cell(row=22, column=7, value=85.50)
    ws.cell(row=22, column=9, value=1)

    ws.cell(row=24, column=2, value="ROOF").font = bold
    ws.cell(row=24, column=2).fill = section_fill
    ws.cell(row=25, column=3, value="ROOF SHEET 1.6MM")
    ws.cell(row=25, column=4, value=4)
    ws.cell(row=25, column=5, value="SHT")
    ws.cell(row=25, column=6, value=420.00)
    ws.cell(row=25, column=7, value=1680.00)
    ws.cell(row=25, column=9, value=1)

    ws.cell(row=27, column=3, value="(disabled item example)").font = note_font
    ws.cell(row=27, column=9, value=0)

    notes = [
        "",
        "NOTES:",
        "• Col B = Section header (e.g. FLOOR, ROOF). Items below it belong to that section.",
        "• Col C = Item name.  Col D = Qty.  Col E = Unit.",
        "• Col F = Unit price (formulas allowed; calculated value is imported).",
        "• Col G (rows 8-18) = Default ratio values for the trailer template.",
        "• Col G (item rows) = Excel total for that item (used for reconciliation).",
        "• Col I = 1 to import the row, 0 to skip.",
        "• Rows with blank price AND blank total are skipped automatically.",
    ]
    for i, txt in enumerate(notes):
        ws.cell(row=30 + i, column=2, value=txt).font = note_font

    widths = [16, 22, 32, 8, 10, 14, 16, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="trailer_import_sample.xlsx"'},
    )


@router.post("/api/import/sheet_preview")
async def api_import_sheet_preview(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    body = await request.json()
    from app.excel_importer import parse_sheet, list_sheets
    upload_id = body.get("upload_id")
    grp_path, formulas_path = _resolve_upload_paths(upload_id)
    if grp_path:
        sheets = list_sheets(grp_path=grp_path)
    else:
        sheets = list_sheets()

    sheet_index = body.get("sheet_index")
    if sheet_index is not None:
        try:
            sheet_name = sheets[int(sheet_index)]
        except (IndexError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid sheet index")
    else:
        sheet_name = body.get("sheet_name", "")
        if not sheet_name:
            raise HTTPException(status_code=400, detail="sheet_name or sheet_index is required")
        if sheet_name not in sheets:
            match = next((s for s in sheets if s.strip() == sheet_name.strip()), None)
            if not match:
                raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found")
            sheet_name = match

    try:
        kwargs = {}
        if grp_path:      kwargs["grp_path"] = grp_path
        if formulas_path: kwargs["formulas_path"] = formulas_path
        ps = parse_sheet(sheet_name, **kwargs)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse failed: {e}")

    sections = []
    for s in ps.sections:
        skin_count = sum(1 for it in s.items if it.is_formula_skin)
        disabled_count = sum(1 for it in s.items if not it.is_enabled)
        sections.append({
            "name": s.name,
            "raw_name": s.raw_name,
            "start_row": s.start_row,
            "end_row": s.end_row,
            "multiplier": s.multiplier,
            "excel_total": s.excel_total,
            "item_count": len(s.items),
            "skin_count": skin_count,
            "disabled_count": disabled_count,
            "items": [
                {
                    "name": it.name,
                    "qty": it.qty,
                    "unit_price": it.unit_price,
                    "excel_total": it.excel_total,
                    "symbolic_formula": it.symbolic_formula,
                    "source_cell": it.source_cell,
                    "is_enabled": it.is_enabled,
                    "is_formula_skin": it.is_formula_skin,
                    "skin_parent": it.skin_parent,
                    "notes": it.notes,
                }
                for it in s.items
            ],
        })

    grand_excel = ps.grand_total_excel
    diff = (ps.computed_total - grand_excel) if grand_excel is not None else None
    pct = (abs(diff) / grand_excel * 100) if (grand_excel and diff is not None) else None
    reconciled = (diff is not None and abs(diff) < 0.01)

    return {
        "sheet_name": ps.sheet_name,
        "trailer_type_hint": ps.trailer_type_hint,
        "length": ps.length,
        "width": ps.width,
        "height": ps.height,
        "markup": ps.markup,
        "constants": ps.constants,
        "grand_total_cell": ps.grand_total_cell,
        "grand_total_excel": grand_excel,
        "computed_total": ps.computed_total,
        "reconciled": reconciled,
        "diff": diff,
        "diff_pct": pct,
        "sections": sections,
        "skipped_sections": ps.skipped_sections,
        "warnings": ps.warnings,
        "section_count": len(ps.sections),
        "item_count": sum(len(s.items) for s in ps.sections),
        "skin_count": sum(1 for s in ps.sections for it in s.items if it.is_formula_skin),
    }


@router.post("/api/import/grp/preview")
async def api_import_grp_preview(request: Request, db: Session = Depends(get_db)):
    """Phase C of the GRP importer rework. Runs the new
    excel_grp_importer.discover() against the chosen sheet and returns
    the WritePlan as JSON, ready to be rendered as a tree. No DB writes.

    Body:
        upload_id        — uploaded workbook session (required for prod;
                           dev falls back to the configured DEFAULT_GRP_PATH)
        sheet_name       — sheet name to scan
        trailer_name     — optional override for the trailer template name
    """
    require_admin(request, db)
    body = await request.json()
    grp_path = _resolve_scan_grp_path((body.get("upload_id") or "").strip() or None)
    sheet_name = (body.get("sheet_name") or "").strip()
    if not sheet_name:
        raise HTTPException(status_code=400, detail="sheet_name is required")
    trailer_name = (body.get("trailer_name") or "").strip() or None

    from app.excel_grp_importer import discover
    try:
        plan = discover(grp_path, sheet_name, trailer_name_override=trailer_name)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Discovery failed: {type(e).__name__}: {e}")

    return _serialise_plan(plan)


@router.post("/api/import/grp/commit")
async def api_import_grp_commit(request: Request, db: Session = Depends(get_db)):
    """Phase D: take a freshly-discovered WritePlan and apply it to the
    DB inside one transaction. The discover() call is repeated server-side
    rather than trusting a client-submitted plan, so users can't smuggle
    in arbitrary BOM rows.

    Body:
        upload_id        — uploaded workbook session
        sheet_name       — sheet name to import
        trailer_name     — optional override
        replace_existing — if True, archive + delete any existing trailer
                           with the chosen name first
    """
    from datetime import datetime, timezone

    require_admin(request, db)
    body = await request.json()
    grp_path = _resolve_scan_grp_path((body.get("upload_id") or "").strip() or None)
    sheet_name = (body.get("sheet_name") or "").strip()
    if not sheet_name:
        raise HTTPException(status_code=400, detail="sheet_name is required")
    trailer_name_override = (body.get("trailer_name") or "").strip() or None
    replace_existing = bool(body.get("replace_existing", False))

    from app.excel_grp_importer import discover
    try:
        plan = discover(grp_path, sheet_name, trailer_name_override=trailer_name_override)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Discovery failed: {type(e).__name__}: {e}")

    if plan.errors:
        raise HTTPException(
            status_code=422,
            detail={"message": "Plan has errors — fix the workbook or override and re-run preview",
                    "errors": [e.__dict__ for e in plan.errors]},
        )

    # Pre-existing trailer handling
    existing_tt = db.query(TrailerType).filter_by(name=plan.trailer_name).first()
    if existing_tt and not replace_existing:
        raise HTTPException(
            status_code=409,
            detail=f"Trailer type {plan.trailer_name!r} already exists. "
                   f"Pass replace_existing=true to wipe and re-import.",
        )
    if existing_tt and replace_existing:
        archive_trailer_template_binding(existing_tt, db)
        db.query(BillOfMaterial).filter_by(trailer_type_id=existing_tt.id).delete()
        db.delete(existing_tt)
        db.flush()

    # ── Create the trailer row ──
    tt = TrailerType(
        name=plan.trailer_name,
        description=f"Imported from GRP sheet: {plan.sheet_name}",
        is_active=True,
        default_length=plan.dimensions.get("length"),
        default_width=plan.dimensions.get("width"),
        default_height=plan.dimensions.get("height"),
    )
    db.add(tt)
    db.flush()
    restored_orphan = restore_orphan_for_trailer(tt, db)

    # ── Body options first (linked items reference them by name) ──
    body_opt_section_id = _resolve_bom_section(db, "BODY OPTIONS")
    sort_idx = 0
    name_to_material_id: dict[str, int] = {}
    body_opt_cat = (db.query(MaterialCategory).filter_by(name="Body Options").first())
    if not body_opt_cat:
        body_opt_cat = MaterialCategory(name="Body Options")
        db.add(body_opt_cat); db.flush()

    # Names handled by the calculator's legacy DRD/SRD machinery
    # (calculator.js _DRDSR_TOGGLE_GROUPS). When an imported body option
    # matches one of these, write it with body_option_group=name so the
    # legacy mutex-toggle UI renders + the section-name-based gate at
    # getBomWithSelectedOptions() correctly enforces "items in 'DRD '
    # section only show when DRD master is on", overriding any per-line
    # body_option_linked the row may carry.
    _CALC_NATIVE_TOGGLES = {"DRD", "SRD"}

    for opt in plan.body_options:
        sort_idx += 1
        # Re-use a Material with that name if one exists, else create
        mat = db.query(Material).filter_by(name=opt.name).first()
        if not mat:
            mat = Material(
                name=opt.name,
                category_id=body_opt_cat.id,
                price_per_unit=0.0,
                unit_of_measure="ea",
                is_active=True,
                last_updated=datetime.now(timezone.utc),
            )
            db.add(mat); db.flush()
        name_to_material_id[opt.name] = mat.id
        # Decide where this option lives in the body-options panel:
        #   • DRD/SRD masters → group=<name> (legacy machinery + native mutex)
        #   • Other auto-detected mutex pairs → multi w/ subgroup so the
        #     existing radio-cluster rule kicks in
        #   • Plain options → multi (independent checkbox)
        if opt.name.upper() in _CALC_NATIVE_TOGGLES:
            grp_label = opt.name.upper()
            sel_mode = "multi"           # legacy DRDSR enforces mutex itself
            sel_group = None
            subgroup = None
        elif opt.radio_group:
            grp_label = "BODY OPTIONS"
            sel_mode = "single"
            sel_group = opt.radio_group
            subgroup = opt.radio_group
        else:
            grp_label = "BODY OPTIONS"
            sel_mode = "multi"
            sel_group = None
            subgroup = None
        bom = BillOfMaterial(
            trailer_type_id=tt.id,
            material_id=mat.id,
            formula_expression="1",
            waste_percentage=0,
            bom_section="BODY OPTIONS",
            bom_section_id=body_opt_section_id,
            is_body_option=True,
            body_option_group=grp_label,
            body_option_subgroup=subgroup,
            body_option_default=opt.default_yn,
            sort_order=sort_idx,
            variable_value=opt.quantity,
            selection_mode=sel_mode,
            selection_group=sel_group,
        )
        db.add(bom)

    # ── Section items ──
    for line in plan.bom_lines:
        sort_idx += 1
        sec_id = _resolve_bom_section(db, line.section)

        # Material — re-use if a row with that name already exists
        mat = db.query(Material).filter_by(name=line.item_name).first()
        if not mat:
            cat = db.query(MaterialCategory).filter_by(name=line.section).first()
            if not cat:
                cat = MaterialCategory(name=line.section)
                db.add(cat); db.flush()
            mat = Material(
                name=line.item_name,
                category_id=cat.id,
                price_per_unit=float(line.price_value or line.price_fallback or 0),
                unit_of_measure="ea",
                is_active=True,
                last_updated=datetime.now(timezone.utc),
            )
            db.add(mat); db.flush()

        # Pick effective unit price
        unit_price_override = None
        skin_id = taping_id = floor_id = cleat_id = None
        if line.price_kind == "external_formulas_2018" and line.price_ref_sheet and line.price_ref_cell:
            from app.excel_formula_matcher import SHEET_MAP
            sheet_key = line.price_ref_sheet.strip().upper()
            sm = SHEET_MAP.get(sheet_key)
            if sm:
                table, fk_col, totals_map, extras = sm
                opt_name = totals_map.get(line.price_ref_cell)
                if opt_name:
                    fk_row = db.execute(
                        f"SELECT id FROM {table} WHERE UPPER(name)=:n",  # type: ignore
                        {"n": opt_name.upper()},
                    ).fetchone() if False else None
                    # Use ORM lookup via raw SQL safely:
                    from sqlalchemy import text as _sql
                    fk_id = db.execute(
                        _sql(f"SELECT id FROM {table} WHERE UPPER(name)=:n LIMIT 1"),
                        {"n": opt_name.upper()},
                    ).scalar()
                    if fk_id:
                        if   fk_col == "skin_formula_id":   skin_id  = fk_id
                        elif fk_col == "taping_block_id":   taping_id = fk_id
                        elif fk_col == "floor_plate_id":    floor_id = fk_id
                        elif fk_col == "mounting_cleat_id": cleat_id = fk_id
        else:
            v = line.price_value if line.price_value is not None else line.price_fallback
            if v is not None:
                unit_price_override = float(v)

        # Linked option — translate body_option_linked + linked_id
        linked_id = name_to_material_id.get(line.gate_option_name) if line.gate_option_name else None

        bom = BillOfMaterial(
            trailer_type_id=tt.id,
            material_id=mat.id,
            formula_expression=line.qty_formula or "1",
            waste_percentage=0,
            notes=f"src:{line.source_addr} qty:{line.qty_source_cell} price:{line.price_kind}",
            bom_section=line.section,
            bom_section_id=sec_id,
            sort_order=sort_idx,
            unit_price_override=unit_price_override,
            skin_formula_id=skin_id,
            taping_block_id=taping_id,
            floor_plate_id=floor_id,
            mounting_cleat_id=cleat_id,
            body_option_linked=line.gate_option_name,
            body_option_linked_id=linked_id,
            excel_formula=line.price_raw_formula,
            source_cell=line.source_addr,
        )
        db.add(bom)

    db.commit()

    return {
        "ok": True,
        "trailer_type_id": tt.id,
        "trailer_name": plan.trailer_name,
        "body_options_imported": len(plan.body_options),
        "bom_lines_imported": len(plan.bom_lines),
        "sections": len(plan.sections),
        "warnings_count": len(plan.warnings),
        "restored_template_binding": bool(restored_orphan),
    }


def _serialise_plan(plan) -> dict:
    """Convert a WritePlan dataclass tree into JSON-friendly dicts."""
    return {
        "sheet_name":        plan.sheet_name,
        "trailer_name":      plan.trailer_name,
        "source_path":       plan.source_path,
        "dimensions":        plan.dimensions,
        "default_margin":    plan.default_margin,
        "default_ratio":     plan.default_ratio,
        "grand_total_excel": plan.grand_total_excel,
        "body_options": [
            {
                "name":         o.name,
                "default_yn":   o.default_yn,
                "quantity":     o.quantity,
                "source_addr":  o.source_addr,
                "radio_group":  o.radio_group,
            } for o in plan.body_options
        ],
        "sections": [
            {
                "name":           s.name,
                "header_row":     s.header_row,
                "total_row":      s.total_row,
                "master_option":  s.master_option,
                "j_multiplier":   s.j_multiplier,
            } for s in plan.sections
        ],
        "bom_lines": [
            {
                "section":             l.section,
                "item_name":           l.item_name,
                "source_addr":         l.source_addr,
                "qty_formula":         l.qty_formula,
                "qty_source_cell":     l.qty_source_cell,
                "price_kind":          l.price_kind,
                "price_value":         l.price_value,
                "price_ref_sheet":     l.price_ref_sheet,
                "price_ref_cell":      l.price_ref_cell,
                "price_fallback":      l.price_fallback,
                "price_raw_formula":   l.price_raw_formula,
                "gate_option_name":    l.gate_option_name,
                "inherited_from_section": l.inherited_from_section,
            } for l in plan.bom_lines
        ],
        "warnings": [{"code": w.code, "message": w.message, "cell": w.cell}
                     for w in plan.warnings],
        "errors":   [{"code": e.code, "message": e.message, "cell": e.cell}
                     for e in plan.errors],
    }


@router.post("/api/import/sheet")
async def api_import_sheet(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    body = await request.json()
    overwrite = bool(body.get("overwrite", False))

    from app.excel_importer import import_sheet, list_sheets
    upload_id = body.get("upload_id")
    grp_path, formulas_path = _resolve_upload_paths(upload_id)
    if grp_path:
        sheets = list_sheets(grp_path=grp_path)
    else:
        sheets = list_sheets()

    sheet_index = body.get("sheet_index")
    if sheet_index is not None:
        try:
            sheet_name = sheets[int(sheet_index)]
        except (IndexError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid sheet index")
    else:
        sheet_name = body.get("sheet_name", "")
        if not sheet_name:
            raise HTTPException(status_code=400, detail="sheet_name or sheet_index is required")
        if sheet_name not in sheets:
            match = next((s for s in sheets if s.strip() == sheet_name.strip()), None)
            if not match:
                raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found")
            sheet_name = match

    trailer_name = body.get("trailer_name", "").strip() or sheet_name.strip()
    try:
        kwargs = {}
        if grp_path:      kwargs["grp_path"] = grp_path
        if formulas_path: kwargs["formulas_path"] = formulas_path
        result = import_sheet(db, trailer_name, sheet_name, overwrite=overwrite, **kwargs)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Formula-reference scanner ─────────────────────────────────────────────
# Read-only scan of the GRP Costings workbook, surfaces which BOM rows draw
# their PRICE from FORMULAS 2018.xls (and which don't), with DB-link
# suggestions wired against existing skin/taping/cleat/floor lookup tables.

def _resolve_scan_grp_path(upload_id: Optional[str]) -> str:
    """Return a path to scan: prefer the uploaded workbook, fall back to the
    admin-configured DEFAULT_GRP_PATH (dev mode only).

    Re-applies the admin-configured excel_folder setting before reading
    DEFAULT_GRP_PATH so a path change made via the admin UI takes effect
    without restarting the server.
    """
    if upload_id:
        info = _UPLOADS.get(upload_id)
        if not info:
            raise HTTPException(
                status_code=404,
                detail="Upload session not found or expired. Please re-upload the workbook.",
            )
        path = info.get("costing")
        if not path or not os.path.exists(path):
            raise HTTPException(
                status_code=410,
                detail="Cached upload is no longer available. Please re-upload.",
            )
        return path

    # Fall back to the admin-configured workbook (only safe in local/dev mode)
    _apply_excel_folder_setting()
    from app.excel_importer import DEFAULT_GRP_PATH as _DGP
    if _DGP and os.path.isfile(_DGP):
        return _DGP
    raise HTTPException(
        status_code=400,
        detail="No workbook available. Upload a GRP Costings .xlsx first.",
    )


@router.get("/admin/templates/formula-scan", response_class=HTMLResponse)
async def admin_formula_scan_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login")
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return templates.TemplateResponse("admin_formula_scan.html", {
        "request": request, "user": user,
    })


_FORMULA_SCAN_CACHE_KEY = "formula_scan_pending"


def _cache_pending_proposals(db: Session, source: str, proposals: list[dict]) -> None:
    """Persist actionable + audit-flagged proposals to admin_settings so the
    Body Templates page can render pills against them.

    Three statuses are kept (all require a matched bom_id):
      • set         — FK currently null, scan suggests a clean link
      • overwrite   — FK exists but mismatches the suggestion
      • unknown_ref — chain touches FORMULAS 2018 but the cell isn't in
                      the lookup table (e.g. col N alternate-price refs).
                      No auto-apply; the pill flags it for manual audit.

    'ok' rows are excluded — the FK is already correct, the row will show
    the regular linked pill via the BOM payload.
    """
    import json
    from datetime import datetime, timezone

    keep: dict[str, dict] = {}
    for p in proposals:
        if not p.get("bom_id"):
            continue
        if p["status"] not in ("set", "overwrite", "unknown_ref"):
            continue
        bom_key = str(p["bom_id"])
        # When a single BOM row has multiple proposals, prefer actionable
        # ones (set > overwrite > unknown_ref) so Apply works as expected
        # and the pill colour reflects the most useful state.
        prio = {"set": 0, "overwrite": 1, "unknown_ref": 2}
        if bom_key in keep and prio[keep[bom_key]["status"]] <= prio[p["status"]]:
            continue
        keep[bom_key] = {
            "trailer_type_id":    p["trailer_type_id"],
            "trailer_type_name":  p["trailer_type_name"],
            "target_table":       p["target_table"],
            "target_fk":          p["target_fk"],
            "target_option_id":   p["target_option_id"],
            "target_option_name": p["target_option_name"],
            "current_link_id":    p["current_link_id"],
            "status":             p["status"],
            "extras":             p.get("extras") or {},
            "item":               p["item"],
            "section":            p.get("section"),
            "ref_sheet":          p.get("ref_sheet"),
            "ref_cell":           p.get("ref_cell"),
            "chain":              p.get("chain"),
        }

    payload = {
        "scanned_at": datetime.now(timezone.utc).isoformat(),
        "source":     source,
        "items":      keep,
    }
    s = db.query(AdminSetting).filter_by(key=_FORMULA_SCAN_CACHE_KEY).first()
    if s:
        s.value = json.dumps(payload)
    else:
        db.add(AdminSetting(key=_FORMULA_SCAN_CACHE_KEY, value=json.dumps(payload)))
    db.commit()


def _load_pending_cache(db: Session) -> dict:
    import json
    s = db.query(AdminSetting).filter_by(key=_FORMULA_SCAN_CACHE_KEY).first()
    if not s or not s.value:
        return {"scanned_at": None, "source": None, "items": {}}
    try:
        return json.loads(s.value)
    except json.JSONDecodeError:
        return {"scanned_at": None, "source": None, "items": {}}


@router.post("/api/admin/formula-scan/run")
async def api_formula_scan_run(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    body = await request.json()
    upload_id = (body.get("upload_id") or "").strip() or None
    only_sheet = (body.get("sheet") or "").strip() or None

    grp_path = _resolve_scan_grp_path(upload_id)

    from app.excel_formula_scanner import scan_workbook
    from app.excel_formula_matcher import build_proposals

    try:
        scan = scan_workbook(grp_path, only_sheet=only_sheet)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scan failed: {type(e).__name__}: {e}")

    matched = build_proposals(scan, db)

    # Cache pending proposals so Body Templates can show migration badges.
    # Only do this on a full-workbook scan; a single-sheet scan would wipe
    # legitimate proposals from other sheets.
    if not only_sheet:
        _cache_pending_proposals(db, scan["source"], matched["proposals"])

    return {
        "source":            scan["source"],
        "formulas_link_ids": scan["formulas_link_ids"],
        "sheets":            scan["sheets"],
        "skipped_sheets":    scan["skipped_sheets"],
        "linked":            scan["linked"],
        "proposals":         matched["proposals"],
        "counts":            matched["counts"],
        "cached":            not only_sheet,
    }


@router.post("/api/admin/formula-scan/apply")
async def api_formula_scan_apply(request: Request, db: Session = Depends(get_db)):
    """Bulk-apply pending suggestions: write FK columns on bill_of_materials
    using the most recent cached scan output.

    Body:
        include_overwrites (bool, default false)
            false — only write 'set' rows where current FK is null
            true  — also overwrite mismatched links
        trailer_type_id (int, optional)
            If provided, restrict the apply to one trailer type. Useful for
            testing the migration on one body before applying everywhere.
            The remainder of the cache stays intact for a later run.
    """
    from sqlalchemy import text as _sql

    require_admin(request, db)
    body = await request.json()
    include_overwrites = bool(body.get("include_overwrites", False))
    only_tt = body.get("trailer_type_id")
    if only_tt is not None:
        try:
            only_tt = int(only_tt)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="trailer_type_id must be an integer")

    cache = _load_pending_cache(db)
    items = cache.get("items") or {}
    if not items:
        raise HTTPException(status_code=400,
            detail="No pending suggestions cached. Run a full scan first.")

    written = 0
    skipped_overwrite = 0
    skipped_unknown = 0
    skipped_other = 0
    skipped_other_tt = 0
    applied_keys: list[str] = []
    failures: list[str] = []

    for bom_id_str, sug in items.items():
        try:
            bom_id = int(bom_id_str)
        except ValueError:
            continue
        if only_tt is not None and sug.get("trailer_type_id") != only_tt:
            skipped_other_tt += 1
            continue
        status = sug.get("status")
        target_fk = sug.get("target_fk")
        target_id = sug.get("target_option_id")
        if status == "unknown_ref" or not target_fk or not target_id:
            # 'unknown_ref' is intentionally not auto-applied — the user
            # must manually link these via the Body Templates page.
            if status == "unknown_ref":
                skipped_unknown += 1
            else:
                skipped_other += 1
            continue
        if status == "overwrite" and not include_overwrites:
            skipped_overwrite += 1
            continue

        sets = [f"{target_fk} = :v"]
        params = {"v": target_id, "id": bom_id}
        for k, v in (sug.get("extras") or {}).items():
            sets.append(f"{k} = :{k}")
            params[k] = v
        sql = _sql(f"UPDATE bill_of_materials SET {', '.join(sets)} WHERE id = :id")
        try:
            db.execute(sql, params)
            written += 1
            applied_keys.append(bom_id_str)
        except Exception as e:
            failures.append(f"bom_id={bom_id}: {type(e).__name__}: {e}")

    db.commit()

    # Cache hygiene:
    #   • Full apply: wipe the cache entirely (re-scan to refresh).
    #   • Scoped apply: drop only the items we wrote so the rest stays
    #     visible as pending pills on other body types.
    setting = db.query(AdminSetting).filter_by(key=_FORMULA_SCAN_CACHE_KEY).first()
    if setting:
        if only_tt is None:
            db.delete(setting)
        else:
            import json
            payload = json.loads(setting.value) if setting.value else {"items": {}}
            for k in applied_keys:
                payload.get("items", {}).pop(k, None)
            setting.value = json.dumps(payload)
        db.commit()

    return {
        "written":            written,
        "skipped_overwrite":  skipped_overwrite,
        "skipped_unknown":    skipped_unknown,
        "skipped_other":      skipped_other,
        "skipped_other_tt":   skipped_other_tt,
        "scope":              "trailer_type_id={}".format(only_tt) if only_tt is not None else "all",
        "failures":           failures,
    }


@router.get("/api/admin/formula-scan/pending/{trailer_type_id}")
async def api_formula_scan_pending_for_trailer(
    trailer_type_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Return cached pending suggestions for one trailer type, keyed by bom_id.

    Used by the Body Templates page to render orange 'needs link' pills
    on rows the scan flagged but the user hasn't applied yet.
    """
    require_admin(request, db)
    cache = _load_pending_cache(db)
    items = cache.get("items") or {}
    out: dict[str, dict] = {}
    for bom_id, sug in items.items():
        if sug.get("trailer_type_id") == trailer_type_id:
            out[bom_id] = {
                "target_table":       sug.get("target_table"),
                "target_option_name": sug.get("target_option_name"),
                "status":             sug.get("status"),
                "ref_sheet":          sug.get("ref_sheet"),
                "ref_cell":           sug.get("ref_cell"),
                "chain":              sug.get("chain"),
            }
    return {
        "scanned_at": cache.get("scanned_at"),
        "items":      out,
    }
