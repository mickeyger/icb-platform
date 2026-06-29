import json
from datetime import datetime, timezone, timedelta
from io import BytesIO

from fastapi import Request, APIRouter, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db, CalculationRecord, TrailerType
from ..deps import get_current_user, user_can, active_branch, assert_calc_access
from ..services import resolve_report_template, strip_excluded_items

router = APIRouter()


@router.get("/results/{record_id}/export/excel")
async def export_excel(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    highlight: int = 0,
    branch=Depends(active_branch),
):
    """Export costing to Excel. highlight=1 → colour-code price changes."""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    if not user_can(user, "export.excel", db):
        raise HTTPException(status_code=403, detail="Permission denied: export.excel")

    rec = db.query(CalculationRecord).filter_by(id=record_id).first()
    if not rec:
        raise HTTPException(status_code=404)
    assert_calc_access(rec.branch_id, user, branch)  # WO v4.37 §3.1 D-3

    dims = json.loads(rec.dimensions_json)
    result = json.loads(rec.result_json)
    result = strip_excluded_items(result)  # only selected items on exports
    tt = db.query(TrailerType).filter_by(id=rec.trailer_type_id).first()
    trailer_name = tt.name if tt else "Trailer"
    is_repair = bool(rec.is_repair)
    customer_name = rec.customer.name if rec.customer else ""

    override_materials: set = set()
    recently_updated_mats: set = set()
    if highlight:
        override_materials = set((result.get("overrides_by_name") or {}).keys())
        cutoff = datetime.now(timezone.utc) - timedelta(days=7)
        for item in result.get("items", []):
            lu = item.get("last_updated")
            if lu and item["material"] not in override_materials:
                try:
                    lu_dt = datetime.fromisoformat(lu)
                    if lu_dt.tzinfo is None:
                        lu_dt = lu_dt.replace(tzinfo=timezone.utc)
                    if lu_dt >= cutoff:
                        recently_updated_mats.add(item["material"])
                except Exception:
                    pass

    import openpyxl
    import io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Breakdown"

    hdr_fill = PatternFill("solid", fgColor="1C2333")
    cat_fill = PatternFill("solid", fgColor="1F3A5F")
    total_fill = PatternFill("solid", fgColor="0D4A8A")
    grand_fill = PatternFill("solid", fgColor="388BFD")
    thin = Border(
        bottom=Side(style="thin", color="30363D"),
        right=Side(style="thin", color="30363D"),
    )

    def hdr(ws, row, col, text, bold=True, fill=None, align="left", num_fmt=None):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=bold, color="E6EDF3" if fill else "8B949E", name="Calibri")
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = thin
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = ("REPAIR QUOTE  —  TRAILER MANUFACTURING COST REPORT"
               if is_repair else "TRAILER MANUFACTURING COST REPORT")
    t.font = Font(bold=True, size=14, name="Calibri",
                  color="E02424" if is_repair else "58A6FF")
    t.fill = PatternFill("solid", fgColor="0D1117")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    s = ws["A2"]
    s.value = f"{trailer_name}  |  Report #{record_id}  |  {rec.created_at.strftime('%d %B %Y')}"
    s.font = Font(size=11, color="8B949E", name="Calibri")
    s.fill = PatternFill("solid", fgColor="161B22")
    s.alignment = Alignment(horizontal="center")

    if customer_name:
        ws.merge_cells("A3:I3")
        c3 = ws["A3"]
        c3.value = f"Client:  {customer_name}"
        c3.font = Font(bold=True, size=11, color="0D1117", name="Calibri")
        c3.alignment = Alignment(horizontal="center")

    ws["A4"] = "DIMENSIONS"
    ws["A4"].font = Font(bold=True, color="388BFD", name="Calibri")
    dim_row1 = [
        ("Length (m)", dims.get("length")),
        ("Width (m)", dims.get("width")),
        ("Height (m)", dims.get("height")),
        ("Num Axles", dims.get("num_axles")),
    ]
    dim_row2 = [
        ("Num Doors", dims.get("num_doors")),
        ("Insulation Thickness (m)", dims.get("insulation_thickness")),
    ]
    for i, (lbl, val) in enumerate(dim_row1):
        ws.cell(row=5, column=i * 2 + 1, value=lbl).font = Font(color="444444", name="Calibri")
        ws.cell(row=5, column=i * 2 + 2, value=val).font = Font(bold=True, color="000000", name="Calibri")
    for i, (lbl, val) in enumerate(dim_row2):
        ws.cell(row=6, column=i * 2 + 1, value=lbl).font = Font(color="444444", name="Calibri")
        ws.cell(row=6, column=i * 2 + 2, value=val).font = Font(bold=True, color="000000", name="Calibri")

    row = 8
    cols = ["Category", "Material", "SAP Code", "Formula", "Quantity", "Unit", "Unit Price (R)", "Waste %", "Line Cost (R)"]
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=col)
        cell.font = Font(bold=True, color="E6EDF3", name="Calibri")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center" if c > 4 else "left", vertical="center")
        cell.border = thin
    ws.row_dimensions[row].height = 18
    row += 1

    # Categories belonging to is_optional BOM sections (e.g. OPTIONAL EXTRAS)
    # render in red — mirrors the on-screen results page and live calculator.
    optional_cats = {
        it["category"] for it in result.get("items", [])
        if it.get("section_is_optional")
    }
    current_cat = None
    # Track each section's header row + its detail-row span so we can add a
    # collapsible outline after all rows are written. Spans are derived
    # dynamically (item counts vary per report) — never hard-coded.
    sections: list[dict] = []
    current_section = None
    for item in result.get("items", []):
        cat = item["category"]
        if cat != current_cat:
            if current_section is not None:
                current_section["last"] = row - 1
                sections.append(current_section)
            ws.merge_cells(f"A{row}:H{row}")   # leave column I free for the section total
            _is_opt_cat = cat in optional_cats
            c = ws.cell(row=row, column=1, value=cat.upper())
            c.font = Font(bold=True,
                          color=("F85149" if _is_opt_cat else "58A6FF"),
                          name="Calibri")
            c.fill = cat_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 16
            current_section = {"header": row, "first": row + 1, "last": None,
                               "kind": "body", "cat": cat, "optional": _is_opt_cat}
            row += 1
            current_cat = cat

        mat_name = item["material"]
        if highlight and mat_name in override_materials:
            price_colour = "CC0000"
            row_tint = "FFF5F5"
        elif highlight and mat_name in recently_updated_mats:
            price_colour = "1A6FBF"
            row_tint = "F0F6FF"
        else:
            price_colour = "000000"
            row_tint = None

        cells_data = [
            ("", "left"), (item["material"], "left"), (item["material_code"], "left"),
            (item["formula"], "left"),
            (item["quantity"], "right"), (item["unit"], "center"),
            (item["unit_price"], "right"), (item["waste_pct"], "right"),
            (item["line_cost"], "right"),
        ]
        for c, (val, align) in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            if c == 7:
                cell.font = Font(color=price_colour, name="Calibri", size=10,
                                 bold=(price_colour != "000000"))
            else:
                cell.font = Font(color="000000", name="Calibri", size=10)
            if row_tint and c != 7:
                cell.fill = PatternFill("solid", fgColor=row_tint)
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = thin
            if c in (7, 9) and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            if c == 8 and isinstance(val, (int, float)) and val:
                cell.value = f"{val}%"
        ws.row_dimensions[row].height = 15
        row += 1

    if current_section is not None:
        current_section["last"] = row - 1
        sections.append(current_section)
        current_section = None

    chassis = result.get("chassis") or {}
    if chassis.get("items"):
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(
            row=row,
            column=1,
            value=(
                f"CHASSIS  ({chassis.get('axle_count')}-axle · "
                f"{chassis.get('tyre_style')} · "
                f"{chassis.get('tyre_count')} tyres · {chassis.get('length')} m)"
            ),
        )
        c.font = Font(bold=True, color="58A6FF", name="Calibri")
        c.fill = cat_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16
        ch_header = row
        row += 1
        ch_first = row
        for it in chassis["items"]:
            cells_data = [
                (it.get("kind", ""), "left"), (it.get("label", ""), "left"), ("", "left"),
                ("", "left"),
                (it.get("qty", 0), "right"), ("ea", "center"),
                (it.get("unit_price", 0), "right"), ("", "right"),
                (it.get("line_cost", 0), "right"),
            ]
            for cn, (val, align) in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=cn, value=val)
                cell.font = Font(color="000000", name="Calibri", size=10)
                cell.alignment = Alignment(horizontal=align, vertical="center")
                cell.border = thin
                if cn in (5, 7, 9) and isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
            ws.row_dimensions[row].height = 15
            row += 1
        if row > ch_first:
            sections.append({"header": ch_header, "first": ch_first, "last": row - 1, "kind": "chassis"})
        ws.merge_cells(f"A{row}:H{row}")
        lc = ws.cell(row=row, column=1, value="CHASSIS SUBTOTAL")
        lc.font = Font(bold=True, color="C9D1D9", name="Calibri")
        lc.fill = total_fill
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws.cell(row=row, column=9, value=chassis.get("subtotal", 0))
        vc.font = Font(bold=True, color="58A6FF", name="Calibri")
        vc.fill = total_fill
        vc.number_format = "#,##0.00"
        vc.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="CATEGORY TOTALS").font = Font(bold=True, color="000000", name="Calibri")
    row += 1
    for cat, total in result.get("category_totals", {}).items():
        _opt = cat in optional_cats
        ws.cell(row=row, column=1, value=cat).font = Font(
            color=("F85149" if _opt else "444444"), name="Calibri",
            bold=_opt,
        )
        cell = ws.cell(row=row, column=9, value=total)
        cell.font = Font(bold=True, color=("F85149" if _opt else "000000"), name="Calibri")
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right")
        row += 1

    row += 1
    summary_rows = [("Cost per m²", result.get("cost_per_sqm", 0))]
    if chassis.get("items"):
        summary_rows.insert(0, ("Body Materials Subtotal", result.get("materials_total", 0)))
        summary_rows.insert(1, ("Chassis Subtotal", chassis.get("subtotal", 0)))
    summary_rows.append(("Total Manufacturing", result.get("grand_total", 0)))
    if result.get("profit_amount"):
        summary_rows.append(
            (f"Profit Margin ({result.get('profit_margin', 0)}%)", result.get("profit_amount", 0))
        )
    if result.get("ratio_amount"):
        summary_rows.append(
            (f"Ratio ({result.get('ratio_label') or ''})", result.get("ratio_amount", 0))
        )
    if result.get("selling_price"):
        summary_rows.append(("SELLING PRICE", result.get("selling_price", 0)))
    else:
        summary_rows.append(("TOTAL MANUFACTURING COST", result.get("grand_total", 0)))
    _disc_amt = float(result.get("discount_amount") or 0)
    if _disc_amt > 0:
        _dlabel = (f"Discount ({result.get('discount_input'):g}%)"
                   if result.get("discount_kind") == "percent" else "Discount")
        summary_rows.append((_dlabel, -_disc_amt))
        summary_rows.append(("NET TOTAL", float(result.get("net_total") or 0)))

    for label, val in summary_rows:
        is_grand = label in ("TOTAL MANUFACTURING COST", "SELLING PRICE", "NET TOTAL")
        ws.merge_cells(f"A{row}:H{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(
            bold=True,
            color="E6EDF3" if is_grand else "C9D1D9",
            size=12 if is_grand else 11,
            name="Calibri",
        )
        lc.fill = grand_fill if is_grand else total_fill
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws.cell(row=row, column=9, value=val)
        vc.font = Font(
            bold=True,
            color="FFFFFF" if is_grand else "58A6FF",
            size=13 if is_grand else 11,
            name="Calibri",
        )
        vc.fill = grand_fill if is_grand else total_fill
        vc.number_format = "#,##0.00"
        vc.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 22 if is_grand else 18
        row += 1

    widths = [14, 42, 18, 32, 12, 8, 14, 9, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if highlight and (override_materials or recently_updated_mats):
        wl = wb.create_sheet("Legend")
        wl.column_dimensions["A"].width = 18
        wl.column_dimensions["B"].width = 52
        wl.column_dimensions["C"].width = 22

        wl.merge_cells("A1:C1")
        lt = wl["A1"]
        lt.value = "PRICE HIGHLIGHT LEGEND"
        lt.font = Font(bold=True, size=13, color="58A6FF", name="Calibri")
        lt.fill = PatternFill("solid", fgColor="0D1117")
        lt.alignment = Alignment(horizontal="center", vertical="center")
        wl.row_dimensions[1].height = 24

        for col, (lbl, clr) in enumerate([("Sample Price", "58A6FF"), ("Meaning", "8B949E"), ("Applies to", "8B949E")], 1):
            hc = wl.cell(row=2, column=col, value=lbl)
            hc.font = Font(bold=True, color=clr, name="Calibri")
            hc.fill = PatternFill("solid", fgColor="1C2333")
            hc.border = thin
            hc.alignment = Alignment(horizontal="center")
        wl.row_dimensions[2].height = 16

        leg_data = []
        if recently_updated_mats:
            leg_data.append((
                "1A6FBF", "F0F6FF", "R 123.45",
                "Price permanently updated in the material database (within last 7 days)",
                ", ".join(sorted(recently_updated_mats)[:5]) +
                (f" + {len(recently_updated_mats)-5} more" if len(recently_updated_mats) > 5 else ""),
            ))
        if override_materials:
            leg_data.append((
                "CC0000", "FFF5F5", "R 99.00",
                "Quote-only price override — not saved to database, applies to this quote only",
                ", ".join(sorted(override_materials)[:5]) +
                (f" + {len(override_materials)-5} more" if len(override_materials) > 5 else ""),
            ))

        for i, (fc, bg, sample, meaning, mats) in enumerate(leg_data, 3):
            wl.cell(row=i, column=1, value=sample).font = Font(bold=True, color=fc, name="Calibri", size=11)
            wl.cell(row=i, column=1).fill = PatternFill("solid", fgColor=bg)
            wl.cell(row=i, column=1).border = thin
            wl.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="center")
            wl.cell(row=i, column=2, value=meaning).font = Font(color="111111", name="Calibri", size=10)
            wl.cell(row=i, column=2).fill = PatternFill("solid", fgColor=bg)
            wl.cell(row=i, column=2).border = thin
            wl.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="center")
            wl.cell(row=i, column=3, value=mats).font = Font(color="555555", name="Calibri", size=9, italic=True)
            wl.cell(row=i, column=3).fill = PatternFill("solid", fgColor=bg)
            wl.cell(row=i, column=3).border = thin
            wl.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="center")
            wl.row_dimensions[i].height = 36

        ws["A2"].value = (ws["A2"].value or "") + "  |  ⬤ Highlighted: price changes colour-coded  →  see Legend tab"

    # ── Collapsible row grouping (outline) — opens EXPANDED ───────────────────
    # Each section's detail rows form an outline group under its always-visible
    # header (summaryBelow=False → the +/- sits beside the header). The sheet
    # opens fully expanded; the user collapses on demand. Headers, the CATEGORY
    # TOTALS block and the pricing rows are never grouped. Each body-section
    # header also carries =SUM of its Line Cost column (col I) so a collapsed
    # section still shows its total (SUM spans hidden rows, so it shows either way).
    if ws.sheet_properties.outlinePr is None:
        from openpyxl.worksheet.properties import Outline
        ws.sheet_properties.outlinePr = Outline()
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.applyStyles = False
    ws.sheet_view.showOutlineSymbols = True
    for sec in sections:
        first, last, header = sec["first"], sec.get("last"), sec["header"]
        empty = last is None or last < first
        if not empty:
            for r in range(first, last + 1):
                ws.row_dimensions[r].outline_level = 1   # detail rows only; never hidden
        if sec.get("kind") == "body":
            tot = ws.cell(row=header, column=9,
                          value=(f"=SUM(I{first}:I{last})" if not empty else 0))
            tot.font = Font(bold=True,
                            color=("F85149" if sec.get("optional") else "58A6FF"),
                            name="Calibri")
            tot.fill = cat_fill
            tot.number_format = "#,##0.00"
            tot.alignment = Alignment(horizontal="right", vertical="center")
            tot.border = thin

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    username = rec.user.username if rec.user else "unknown"
    filename = f"Costing_{trailer_name.replace(' ', '_')}_{record_id}_{username}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _cost_breakdown_pdf_reportlab(ctx: dict) -> bytes:
    """Pure-Python (ReportLab) fallback for the cost-breakdown PDF.

    WeasyPrint is deliberately absent from this deployment (needs native
    GTK/Pango/cairo; can't build on the HostAfrica cPanel/CageFS prod host —
    see requirements.txt and ADR 0017). When WeasyPrint can't be imported,
    the /export/pdf endpoint renders the *same* ``ctx`` data through ReportLab
    instead. Layout mirrors templates/reports/cost_breakdown.html (A4 landscape)
    but is faithful, not a pixel match.
    """
    from xml.sax.saxutils import escape

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfgen import canvas
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    def _fmt2(v):
        """Mirror the template's ``'{:,.2f}'.format(x) if x is number`` rule."""
        try:
            return "{:,.2f}".format(float(v))
        except (TypeError, ValueError):
            return "" if v in (None, "") else str(v)

    page_w, page_h = landscape(A4)
    left_margin = right_margin = 10 * mm
    avail_w = page_w - left_margin - right_margin

    optional_cats = set(ctx.get("optional_cats") or [])
    BLUE = colors.HexColor("#58A6FF")
    RED = colors.HexColor("#F85149")

    body = ParagraphStyle("cell", fontName="Helvetica", fontSize=8, leading=9.5)
    title_style = ParagraphStyle(
        "title", fontName="Helvetica-Bold", fontSize=14, leading=18,
        alignment=TA_CENTER, textColor=BLUE,
    )
    sub_style = ParagraphStyle(
        "sub", fontName="Helvetica", fontSize=9.5, leading=12,
        alignment=TA_CENTER, textColor=colors.HexColor("#C9D1D9"),
    )
    client_style = ParagraphStyle(
        "client", fontName="Helvetica-Bold", fontSize=11, leading=14,
        alignment=TA_CENTER, textColor=colors.HexColor("#0D1117"),
    )

    elements = []

    # --- Title bar -------------------------------------------------------
    title_txt = "TRAILER MANUFACTURING COST REPORT"
    if ctx.get("is_repair"):
        title_txt += '<font color="#E02424">  &mdash;  REPAIR QUOTE</font>'
    title_tbl = Table([[Paragraph(title_txt, title_style)]], colWidths=[avail_w])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0D1117")),
        ("TOPPADDING", (0, 0), (-1, -1), 6 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6 * mm),
    ]))
    elements.append(title_tbl)

    sub_txt = "%s  |  Report #%s  |  %s" % (
        escape(str(ctx.get("trailer_name") or "")),
        escape(str(ctx.get("record_id") or "")),
        escape(str(ctx.get("created_at_human") or "")),
    )
    sub_tbl = Table([[Paragraph(sub_txt, sub_style)]], colWidths=[avail_w])
    sub_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#161B22")),
        ("TOPPADDING", (0, 0), (-1, -1), 2 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2 * mm),
    ]))
    elements.append(sub_tbl)
    elements.append(Spacer(1, 4 * mm))

    if ctx.get("customer_name"):
        elements.append(Paragraph(
            "Client: " + escape(str(ctx["customer_name"])), client_style))
        elements.append(Spacer(1, 4 * mm))

    # --- Dimensions ------------------------------------------------------
    dims = ctx.get("dims") or {}

    def _d(key):
        v = dims.get(key)
        return "" if v in (None, "") else str(v)

    elements.append(Paragraph(
        '<font color="#1F3A5F"><b>DIMENSIONS</b></font>',
        ParagraphStyle("dh", fontSize=10, leading=12)))
    elements.append(Spacer(1, 1 * mm))
    dim_rows = [
        ["Length (m)", _d("length"), "Width (m)", _d("width"),
         "Height (m)", _d("height"), "Num Axles", _d("num_axles")],
        ["Num Doors", _d("num_doors"), "Insulation Thk (m)",
         _d("insulation_thickness"), "", "", "", ""],
    ]
    lbl_w = avail_w * 0.14
    val_w = avail_w * 0.11
    dim_tbl = Table(dim_rows, colWidths=[lbl_w, val_w] * 4)
    dim_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#555555")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#555555")),
        ("TEXTCOLOR", (4, 0), (4, -1), colors.HexColor("#555555")),
        ("TEXTCOLOR", (6, 0), (6, -1), colors.HexColor("#555555")),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("FONTNAME", (5, 0), (5, -1), "Helvetica-Bold"),
        ("FONTNAME", (7, 0), (7, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 0.8 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8 * mm),
    ]))
    elements.append(dim_tbl)
    elements.append(Spacer(1, 4 * mm))

    # --- BOM table (grouped by category) --------------------------------
    bom_pct = [0.11, 0.26, 0.09, 0.19, 0.07, 0.05, 0.08, 0.06, 0.09]
    bom_w = [avail_w * p for p in bom_pct]
    header = ["Category", "Material", "SAP Code", "Formula", "Quantity",
              "Unit", "Unit Price (R)", "Waste %", "Line Cost (R)"]
    data = [header]
    cat_rows, optional_rows = [], []
    current = object()
    for it in ctx.get("items") or []:
        cat = it.get("category")
        if cat != current:
            data.append([str(cat or ""), "", "", "", "", "", "", "", ""])
            r = len(data) - 1
            cat_rows.append(r)
            if cat in optional_cats:
                optional_rows.append(r)
            current = cat
        waste = it.get("waste_pct")
        data.append([
            "",
            Paragraph(escape(str(it.get("material") or "")), body),
            it.get("material_code") or "",
            Paragraph(escape(str(it.get("formula") or "")), body),
            _fmt2(it.get("quantity")),
            it.get("unit") or "",
            _fmt2(it.get("unit_price")),
            ("%s%%" % waste) if waste else "",
            _fmt2(it.get("line_cost")),
        ])

    bom_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C2333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#E6EDF3")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ALIGN", (5, 0), (5, -1), "CENTER"),
        ("ALIGN", (6, 0), (8, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
    ]
    for r in cat_rows:
        bom_style.append(("SPAN", (0, r), (-1, r)))
        bom_style.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#1F3A5F")))
        bom_style.append(("TEXTCOLOR", (0, r), (-1, r), RED if r in optional_rows else BLUE))
        bom_style.append(("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"))
        bom_style.append(("ALIGN", (0, r), (-1, r), "LEFT"))
    bom_tbl = Table(data, colWidths=bom_w, repeatRows=1)
    bom_tbl.setStyle(TableStyle(bom_style))
    elements.append(bom_tbl)

    # --- Chassis (optional) ---------------------------------------------
    chassis = ctx.get("chassis")
    if chassis and chassis.get("items"):
        chead = "CHASSIS — %s-axle · %s · %s tyres · %s m" % (
            chassis.get("axle_count"), chassis.get("tyre_style"),
            chassis.get("tyre_count"), chassis.get("length"))
        cdata = [[chead, "", "", "", "", "", "", "", ""]]
        for it in chassis["items"]:
            cdata.append([
                it.get("kind") or "",
                Paragraph(escape(str(it.get("label") or "")), body),
                "", "",
                _fmt2(it.get("qty")), "ea", _fmt2(it.get("unit_price")),
                "", _fmt2(it.get("line_cost")),
            ])
        cdata.append(["Chassis Subtotal", "", "", "", "", "", "", "",
                      _fmt2(chassis.get("subtotal"))])
        last = len(cdata) - 1
        ctbl = Table(cdata, colWidths=bom_w)
        ctbl.setStyle(TableStyle([
            ("SPAN", (0, 0), (-1, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), BLUE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("SPAN", (0, last), (-2, last)),
            ("ALIGN", (0, last), (-2, last), "RIGHT"),
            ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#EFF3F8")),
            ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ALIGN", (4, 1), (4, last - 1), "RIGHT"),
            ("ALIGN", (5, 1), (5, last - 1), "CENTER"),
            ("ALIGN", (6, 1), (8, last), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
        ]))
        elements.append(Spacer(1, 4 * mm))
        elements.append(ctbl)

    # --- Category totals (optional) -------------------------------------
    cat_totals = ctx.get("category_totals") or {}
    if cat_totals:
        ct_data = [["Category", "Subtotal (R)"]]
        ct_optional = []
        for cat, total in cat_totals.items():
            ct_data.append([str(cat), _fmt2(total)])
            if cat in optional_cats:
                ct_optional.append(len(ct_data) - 1)
        ct_tbl = Table(ct_data, colWidths=[avail_w * 0.45, avail_w * 0.15])
        ct_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C2333")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#E6EDF3")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),
            ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#D0D7DE")),
            ("TOPPADDING", (0, 0), (-1, -1), 1.2 * mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2 * mm),
        ]
        for r in ct_optional:
            ct_style.append(("TEXTCOLOR", (0, r), (-1, r), RED))
        ct_tbl.setStyle(TableStyle(ct_style))
        elements.append(Spacer(1, 4 * mm))
        elements.append(ct_tbl)

    # --- Summary ---------------------------------------------------------
    def _summary_row(rows, label, value, grand=False):
        rows.append((label, value, grand))

    srows = []
    _summary_row(srows, "Cost per m²", _fmt2(ctx.get("cost_per_sqm") or 0))
    _summary_row(srows, "Total Manufacturing", _fmt2(ctx.get("grand_total") or 0))
    if ctx.get("profit_amount"):
        _summary_row(srows, "Profit Margin (%s%%)" % ctx.get("profit_margin"),
                     _fmt2(ctx.get("profit_amount") or 0))
    if ctx.get("ratio_amount"):
        _summary_row(srows, "Ratio (%s)" % ctx.get("ratio_label"),
                     _fmt2(ctx.get("ratio_amount") or 0))

    selling = ctx.get("selling_price")
    grand_total = ctx.get("grand_total") or 0
    discount = ctx.get("discount_amount") or 0
    try:
        discount = float(discount)
    except (TypeError, ValueError):
        discount = 0
    if discount > 0:
        _summary_row(srows, "SELLING PRICE" if selling else "TOTAL",
                     "R " + _fmt2(selling if selling else grand_total))
        disc_lbl = "Discount"
        if ctx.get("discount_kind") == "percent":
            try:
                disc_lbl = "Discount (%g%%)" % float(ctx.get("discount_input"))
            except (TypeError, ValueError):
                pass
        _summary_row(srows, disc_lbl, "− R " + _fmt2(discount))
        _summary_row(srows, "NET TOTAL", "R " + _fmt2(ctx.get("net_total") or 0),
                     grand=True)
    else:
        _summary_row(
            srows,
            "SELLING PRICE" if selling else "TOTAL MANUFACTURING COST",
            "R " + _fmt2(selling if selling else grand_total),
            grand=True,
        )

    sdata = [[lbl, val] for (lbl, val, _g) in srows]
    summary_tbl = Table(sdata, colWidths=[avail_w * 0.75, avail_w * 0.25])
    s_style = [
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
        ("TOPPADDING", (0, 0), (-1, -1), 2 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2 * mm),
    ]
    for i, (_lbl, _val, grand) in enumerate(srows):
        if grand:
            s_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#388BFD")))
            s_style.append(("TEXTCOLOR", (0, i), (-1, i), colors.white))
            s_style.append(("FONTSIZE", (0, i), (0, i), 11))
            s_style.append(("FONTSIZE", (1, i), (1, i), 12))
        else:
            s_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#0D4A8A")))
            s_style.append(("TEXTCOLOR", (0, i), (0, i), colors.HexColor("#C9D1D9")))
            s_style.append(("TEXTCOLOR", (1, i), (1, i), BLUE))
    summary_tbl.setStyle(TableStyle(s_style))
    elements.append(Spacer(1, 5 * mm))
    elements.append(summary_tbl)

    # --- Footer (Generated … / Page X of Y) ------------------------------
    gen_text = "Generated " + str(ctx.get("generated_at") or "")
    foot_left_x = left_margin
    foot_right_x = page_w - right_margin

    class _NumberedCanvas(canvas.Canvas):
        def __init__(self, *a, **k):
            canvas.Canvas.__init__(self, *a, **k)
            self._saved_pages = []

        def showPage(self):
            self._saved_pages.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved_pages)
            for state in self._saved_pages:
                self.__dict__.update(state)
                self.setFont("Helvetica", 8)
                self.setFillColor(colors.HexColor("#555555"))
                self.drawString(foot_left_x, 8 * mm, gen_text)
                self.drawRightString(
                    foot_right_x, 8 * mm,
                    "Page %d of %d" % (self._pageNumber, total))
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=left_margin, rightMargin=right_margin,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title="Cost Breakdown — Report #%s" % ctx.get("record_id"),
    )
    doc.build(elements, canvasmaker=_NumberedCanvas)
    return buf.getvalue()


@router.get("/results/{record_id}/export/pdf")
async def export_pdf(record_id: int, request: Request, db: Session = Depends(get_db), branch=Depends(active_branch)):
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    if not user_can(user, "export.pdf", db):
        raise HTTPException(status_code=403, detail="Permission denied: export.pdf")

    rec = db.query(CalculationRecord).filter_by(id=record_id).first()
    if not rec:
        raise HTTPException(status_code=404)
    assert_calc_access(rec.branch_id, user, branch)  # WO v4.37 §3.1 D-3

    dims = json.loads(rec.dimensions_json)
    result = json.loads(rec.result_json)
    result = strip_excluded_items(result)  # only selected items on exports
    tt = db.query(TrailerType).filter_by(id=rec.trailer_type_id).first()
    trailer_name = tt.name if tt else "Trailer"
    customer_info = None
    if rec.customer:
        customer_info = {
            "name": rec.customer.name or "",
            "email": rec.customer.email or "",
            "telephone": rec.customer.telephone or "",
        }

    # Categories belonging to is_optional BOM sections — flagged once so the
    # template/renderer can render their headers + totals in red, matching the
    # on-screen results page.
    _items = result.get("items", [])
    optional_cats = sorted({it["category"] for it in _items if it.get("section_is_optional")})

    ctx = {
        "trailer_name": trailer_name,
        "record_id": record_id,
        "created_at_human": rec.created_at.strftime("%d %B %Y") if rec.created_at else "",
        "generated_at": datetime.now().strftime("%d %b %Y %H:%M"),
        "customer_name": (customer_info or {}).get("name") or "",
        "is_repair": bool(rec.is_repair),
        "dims": dims,
        "items": _items,
        "category_totals": result.get("category_totals", {}) or {},
        "optional_cats": optional_cats,
        "cost_per_sqm": result.get("cost_per_sqm", 0),
        "profit_margin": result.get("profit_margin", 0),
        "profit_amount": result.get("profit_amount", 0),
        "ratio_value": result.get("ratio_value"),
        "ratio_label": result.get("ratio_label"),
        "ratio_amount": result.get("ratio_amount", 0),
        "selling_price": result.get("selling_price"),
        "grand_total": result.get("grand_total", 0),
        "chassis": result.get("chassis"),
        "materials_total": result.get("materials_total"),
        "discount_kind": result.get("discount_kind"),
        "discount_input": result.get("discount_input"),
        "discount_amount": result.get("discount_amount", 0),
        "net_total": result.get("net_total"),
    }

    try:
        # Prefer WeasyPrint (high-fidelity HTML/CSS render) when it's available.
        # It needs native GTK/Pango/cairo libs that this deployment deliberately
        # omits (see requirements.txt + ADR 0017), so fall back to the pure-Python
        # ReportLab renderer when the import fails — rendering the same ctx data.
        try:
            from weasyprint import HTML  # noqa: WPS433 (lazy, optional dep)
            from ..report_engine import _env

            html_str = _env.get_template("cost_breakdown.html").render(**ctx)
            pdf_bytes = HTML(string=html_str).write_pdf()
        except (ImportError, OSError):
            pdf_bytes = _cost_breakdown_pdf_reportlab(ctx)

        username = user.username if user else "unknown"
        filename = f"Costing_{trailer_name.replace(' ', '_')}_{record_id}_{username}.pdf"

        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception:
        # WO v4.36d §3.1 — log the full exception server-side; return a GENERIC client message. The prior
        # detail=f"...: {exc}" leaked internal error text to the caller (Subagent B finding). 500 preserved.
        import logging
        logging.getLogger(__name__).exception("export_pdf failed (record %s)", record_id)
        raise HTTPException(status_code=500, detail="PDF generation failed")


@router.get("/results/{record_id}/report")
async def report_for_record(record_id: int, request: Request, db: Session = Depends(get_db), branch=Depends(active_branch)):
    """Render the report PDF using the trailer's resolved ReportTemplate."""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    if not user_can(user, "quote.generate", db):
        raise HTTPException(status_code=403, detail="Permission denied: quote.generate")

    rec = db.query(CalculationRecord).filter_by(id=record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Costing record not found")
    assert_calc_access(rec.branch_id, user, branch)  # WO v4.37 §3.1 D-3

    tt = db.query(TrailerType).filter_by(id=rec.trailer_type_id).first()
    tmpl = resolve_report_template(tt)
    if not tmpl:
        raise HTTPException(status_code=400, detail="No report template is assigned to this trailer type.")

    dims = json.loads(rec.dimensions_json or "{}")
    result = json.loads(rec.result_json or "{}")
    result = strip_excluded_items(result)  # only selected items on the report
    customer = None
    if rec.customer:
        customer = {
            "name": rec.customer.name or "",
            "email": rec.customer.email or "",
            "telephone": rec.customer.telephone or "",
        }

    try:
        from ..report_engine import render_by_slug
        pdf_bytes = render_by_slug(
            slug=tmpl.slug,
            record_id=record_id,
            customer=customer,
            dimensions=dims,
            result=result,
            created_at=rec.created_at,
            quote_number=rec.quote_number,
        )
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report generation failed: {e}")

    safe_customer = (
        "".join(ch for ch in (rec.customer.name if rec.customer else "Customer")
                if ch.isalnum() or ch in (" ", "_", "-"))
        .strip()
        .replace(" ", "_")
    ) or "Customer"
    safe_slug = tmpl.slug.replace("/", "_")
    filename = f"{safe_slug}_{record_id}_{safe_customer}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/results/{record_id}/report/explosive-quote")
async def report_explosive_quote_compat(record_id: int, request: Request, db: Session = Depends(get_db), branch=Depends(active_branch)):
    return await report_for_record(record_id=record_id, request=request, db=db, branch=branch)
