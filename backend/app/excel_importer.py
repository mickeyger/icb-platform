"""
excel_importer.py  –  v2 importer for GRP Costings 2018.xlsx

Produces a structured ParsedSheet object that can be inspected (dry-run)
or committed to the database.  Replaces import_excel_sheet.py once wired
into the admin UI.

Rules captured from the source workbook:
  • Dimensions (default_length / width / height) come from C4, C5, C6
    (labels LENGTH / WIDTH / HEIGHT in col A — tolerant of row offset).
  • Markup is G5; trailer constants are A8:A18 / C8:C18 → TrailerRatio.
  • Sections are Column B labels; their boundary is the next Column B label
    (or the GRAND TOTAL row).
  • A section is skipped entirely if its aggregated Column J value is 0.
  • DOOR FITTINGS is disambiguated into SRD DOOR FITTINGS / DRD DOOR FITTINGS
    based on the most-recently-seen SRD or DRD header.
  • Column I holds a per-line on/off flag (cached 0 or 1).  If cached value
    is 0 the line is excluded.
  • The price column (G) may reference the FORMULA SKINS sheet of
    FORMULAS 2018.xls — those items are expanded into child BOM lines.
  • Symbolic formula is derived by translating cell refs in the H-column
    formula (C4→length, C5→width, C6→height), with external / aggregate
    references falling back to cached numeric values.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field, asdict
from typing import Optional

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

DEFAULT_GRP_PATH = (
    os.environ.get("EXCEL_PATH")
    or r"C:\Users\micge\Documents\Burt Costing Model\GRP Costings 2018.xlsx"
)
DEFAULT_FORMULAS_PATH = (
    os.environ.get("FORMULAS_XLS_PATH")
    or r"C:\Users\micge\Documents\Burt Costing Model\FORMULAS 2018.xls"
)


# ── Data classes ───────────────────────────────────────────────────────────

@dataclass
class ParsedItem:
    name: str
    section: str
    qty: Optional[float]
    unit_price: float
    excel_total: Optional[float]
    excel_formula: Optional[str]
    symbolic_formula: str
    source_cell: str                # e.g. "H27"
    is_enabled: bool = True
    is_formula_skin: bool = False
    highlight_color: Optional[str] = None
    skin_parent: Optional[str] = None
    notes: Optional[str] = None
    # BODY OPTIONS fields
    is_body_option: bool = False
    body_option_group: Optional[str] = None
    body_option_subgroup: Optional[str] = None   # within-group mutual-exclusion label
    body_option_default: bool = False
    # For regular section items: the body-option name that activates this row.
    body_option_linked: Optional[str] = None


@dataclass
class ParsedSection:
    name: str                       # display name (may include SRD/DRD prefix)
    raw_name: str                   # original Column B text
    start_row: int
    end_row: int
    total_row: Optional[int]
    multiplier: float
    excel_total: float              # cached value in column J
    items: list[ParsedItem] = field(default_factory=list)


@dataclass
class ParsedSheet:
    sheet_name: str
    trailer_type_hint: str          # value in E1 (CHILLER/FREEZER/EXPLOSIVE)
    length: Optional[float]
    width: Optional[float]
    height: Optional[float]
    markup: Optional[float]
    constants: list[dict]           # [{label, value, cell}]
    grand_total_cell: Optional[str]
    grand_total_excel: Optional[float]
    computed_total: float           # sum of section (excel_total × multiplier)
    sections: list[ParsedSection]
    skipped_sections: list[str]
    warnings: list[str]
    body_options: list[ParsedItem] = field(default_factory=list)


# ── Helpers ────────────────────────────────────────────────────────────────

SIZE_LABELS = {"LENGTH", "WIDTH", "HEIGHT"}
CELL_REF_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")
EXTERNAL_REF_RE = re.compile(r"'?\[[^\]]*\][^'!]+'?![^,\s)]+|\[[^\]]*\][^!]+![^,\s)]+")
SKIN_REF_RE = re.compile(
    r"""['"]?(?:\[\d+\])?FORMULA\s+SKINS['"]?!\$?([A-Z]+)\$?(\d+)""", re.I
)
AGGREGATE_RE = re.compile(
    r"^(SUM|IF|SUMPRODUCT|OFFSET|AVERAGE|MAX|MIN|VLOOKUP|INDEX|MATCH)\(", re.I
)

# Matches an external workbook reference like:
#   '[1]EPS'!$C$12        '[FORMULAS 2018.xls]EPS'!$C$12
#   [1]EPS!$C$12          'EPS'!$C$12  (foreign sheet on same wb is NOT this)
EXT_PRICE_REF_RE = re.compile(
    r"""
    '?                     # optional opening quote
    \[(?P<book>[^\]]+)\]   # [book]
    (?P<sheet>[^'!]+)      # sheet
    '?                     # optional closing quote
    !
    \$?(?P<col>[A-Z]+)\$?(?P<row>\d+)
    """,
    re.VERBOSE,
)


def _resolve_price_formula(price_formula, ws_d, formulas_path: str) -> Optional[float]:
    """Evaluate a price-cell formula, resolving external refs against the
    FORMULAS workbook so that multipliers (e.g. ``=FORMULAS!B12 * 1.15``) are
    applied. Returns ``None`` if the formula cannot be safely evaluated.
    """
    if not isinstance(price_formula, str) or not price_formula.startswith("="):
        return None
    expr = price_formula[1:].strip()
    # Bail on aggregate functions / lookups — those need a real engine.
    if AGGREGATE_RE.match(expr):
        return None

    # Resolve external workbook references first.
    formulas_wb_cache: dict = {}

    def _ext_value(book: str, sheet: str, col: str, row: int) -> float:
        # In this workbook, external refs always point at FORMULAS 2018.xls
        # (whether spelled '[1]' or '[FORMULAS 2018.xls]').
        wb = formulas_wb_cache.get("wb")
        if wb is None:
            try:
                wb = openpyxl.load_workbook(formulas_path, data_only=True, read_only=True)
            except Exception:
                return 0.0
            formulas_wb_cache["wb"] = wb
        try:
            ws = wb[sheet] if sheet in wb.sheetnames else wb[sheet.strip("'")]
            v = _safe_float(ws.cell(int(row), openpyxl.utils.column_index_from_string(col)).value)
            return float(v) if v is not None else 0.0
        except Exception:
            return 0.0

    def _ext_sub(m: re.Match) -> str:
        v = _ext_value(m.group("book"), m.group("sheet"),
                       m.group("col"), int(m.group("row")))
        return _fmt_num(v)

    expr2 = EXT_PRICE_REF_RE.sub(_ext_sub, expr)

    # Any remaining `[..]..!..` refs we can't parse — give up.
    if "[" in expr2 and "!" in expr2:
        return None

    # Resolve same-workbook cell refs to cached numeric values.
    def _local_sub(m: re.Match) -> str:
        key = f"{m.group(1)}{m.group(2)}"
        try:
            v = _safe_float(ws_d[key].value)
        except Exception:
            v = None
        return _fmt_num(v or 0.0)
    expr3 = CELL_REF_RE.sub(_local_sub, expr2)

    # Strip sheet-qualified local refs that may remain (e.g. Sheet!A1 → A1 already handled above)
    # Then evaluate the pure-arithmetic expression.
    try:
        result = eval(expr3, {"__builtins__": {}}, {})  # noqa: S307
        return max(0.0, float(result))
    except Exception:
        return None


def _safe_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _fmt_num(v: float) -> str:
    if v == 0:
        return "0"
    s = f"{v:.8f}".rstrip("0").rstrip(".")
    return s or "0"


def _needs_parens(expr: str) -> bool:
    if re.fullmatch(r"[\w.]+", expr):
        return False
    depth = 0
    for ch in expr:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch in "+-" and depth == 0:
            return True
    return False


def _maybe_wrap(expr: str) -> str:
    return f"({expr})" if _needs_parens(expr) else expr


def _simplify_mul_by_one(expr: str) -> str:
    """Collapse `* 1` / `1 *` artefacts left over when the unit-price cell
    is substituted with 1 during formula translation."""
    prev = None
    cur = expr
    while prev != cur:
        prev = cur
        cur = re.sub(r"\*\s*1(?![\d.])", "", cur)       # trailing *1
        cur = re.sub(r"(?<![\d.])1\s*\*", "", cur)      # leading 1*
        cur = re.sub(r"\(\s*([^()]+?)\s*\)", lambda m: (
            f"({m.group(1)})" if _needs_parens(m.group(1)) else m.group(1)
        ), cur)                                         # strip redundant parens
    return cur.strip() or "1"


# ── Formula translation ───────────────────────────────────────────────────

def _build_cell_var_map(size_rows: dict) -> dict[str, str]:
    """Map base cells (C{r}) of LENGTH/WIDTH/HEIGHT rows to dim variables."""
    m: dict[str, str] = {}
    for label, row in size_rows.items():
        m[f"C{row}"] = label
        # derivatives seen in the sheets: D{r}=C{r}+0.05, E{r}=(C{r}+0.05)*N
        m[f"D{row}"] = f"({label} + 0.05)"
    return m


def _translate_formula(
    excel_formula: str,
    ws_d,
    ws_f,
    cell_var_map: dict[str, str],
    visited: Optional[set] = None,
    depth: int = 0,
    exclude_cells: Optional[set] = None,
) -> str:
    """Translate an Excel formula into a Python expression using
    length/width/height variables where possible.  External refs and
    aggregates collapse to their cached numeric values."""
    if visited is None:
        visited = set()
    if depth > 8:
        return "0"
    if not isinstance(excel_formula, str) or not excel_formula.startswith("="):
        return ""

    expr = excel_formula[1:].strip()

    # external workbook ref (e.g. '[1]EPS'!$C$12, '[2]FORMULA SKINS'!$D$49)
    if EXTERNAL_REF_RE.search(expr) or ("[" in expr and "!" in expr):
        return None  # caller falls back to cached numeric

    if AGGREGATE_RE.match(expr):
        return None

    def sub_ref(m: re.Match) -> str:
        key = f"{m.group(1)}{m.group(2)}"
        # Unit-price cells get factored out: return "1" so the saved formula
        # is pure quantity (runtime multiplies by material.price_per_unit).
        if exclude_cells and key in exclude_cells:
            return "1"
        if key in cell_var_map:
            return cell_var_map[key]
        if key in visited:
            # cycle fallback
            v = ws_d[key].value
            return _fmt_num(_safe_float(v) or 0.0)
        visited.add(key)
        raw = ws_f[key].value
        if isinstance(raw, str) and raw.startswith("="):
            translated = _translate_formula(raw, ws_d, ws_f, cell_var_map,
                                            visited, depth + 1,
                                            exclude_cells=exclude_cells)
            if translated is None:
                v = _safe_float(ws_d[key].value)
                return _fmt_num(v or 0.0)
            return _maybe_wrap(translated)
        v = _safe_float(raw)
        return _fmt_num(v or 0.0)

    translated = CELL_REF_RE.sub(sub_ref, expr)
    # collapse `+0` / `-0`
    translated = re.sub(r"([+\-])0(?!\.\d)", "", translated).lstrip("+").strip()
    return translated or "0"


# ── Formula Skin expansion ────────────────────────────────────────────────

def _is_formula_skin_ref(g_formula: str) -> Optional[tuple[str, int]]:
    """Return (col_letter, row) if g_formula references FORMULA SKINS, else None."""
    if not isinstance(g_formula, str):
        return None
    m = SKIN_REF_RE.search(g_formula)
    if m:
        return m.group(1).upper(), int(m.group(2))
    return None


def _expand_formula_skin(total_col: str, total_row: int,
                         formulas_path: str) -> list[dict]:
    """Open FORMULAS 2018.xls, locate the block whose TOTAL is at
    (total_col, total_row), and return a list of sub-items."""
    try:
        import xlrd  # type: ignore
    except ImportError:
        return []

    if not os.path.exists(formulas_path):
        return []

    try:
        wb = xlrd.open_workbook(formulas_path)
        ws = wb.sheet_by_name("FORMULA SKINS")
    except Exception:
        return []

    col_idx = column_index_from_string(total_col) - 1  # 0-based for xlrd
    # The block layout is: blank row | ITEM header | blank | data... | blank | TOTAL
    # Walk upwards from (total_row-2) collecting rows until we hit the ITEM header.
    children: list[dict] = []
    r = total_row - 2  # 1-based → 0-based, one above the TOTAL row
    while r >= 0:
        item = ws.cell_value(r, 0)
        if isinstance(item, str) and item.strip().upper() == "ITEM":
            break
        if isinstance(item, str) and item.strip():
            qty = ws.cell_value(r, 1) if ws.ncols > 1 else None
            price = ws.cell_value(r, 2) if ws.ncols > 2 else None
            line_total = ws.cell_value(r, col_idx) if ws.ncols > col_idx else None
            sap = ws.cell_value(r, 4) if ws.ncols > 4 else None
            try:
                qty_f = float(qty) if qty not in (None, "") else None
                price_f = float(price) if price not in (None, "") else None
                total_f = float(line_total) if line_total not in (None, "") else None
            except (TypeError, ValueError):
                qty_f = price_f = total_f = None
            children.append({
                "name": item.strip(),
                "qty": qty_f,
                "unit_price": price_f or 0.0,
                "total": total_f,
                "sap_code": sap.strip() if isinstance(sap, str) else None,
            })
        r -= 1

    children.reverse()
    return children


# ── Sheet parser ──────────────────────────────────────────────────────────

def _find_size_rows(ws_d, max_scan: int = 30) -> dict:
    """Return {"length": row, "width": row, "height": row}."""
    rows = {}
    for r in range(1, min(max_scan, ws_d.max_row) + 1):
        a = ws_d.cell(r, 1).value
        if isinstance(a, str):
            lbl = a.strip().upper()
            if lbl in SIZE_LABELS:
                rows[lbl.lower()] = r
    return rows


def _find_column_b_sections(ws_d, size_rows_set: set[int]) -> list[tuple[int, str]]:
    """Return [(row, section_name)] for every Column B label row."""
    out = []
    for r in range(1, ws_d.max_row + 1):
        if r in size_rows_set:
            continue
        a = ws_d.cell(r, 1).value
        b = ws_d.cell(r, 2).value
        if b and isinstance(b, str) and b.strip() and a is None:
            txt = b.strip()
            if txt.upper() in ("TOTAL", "GRAND TOTAL"):
                continue
            out.append((r, txt))
    return out


def _extract_multiplier(j_formula: str) -> float:
    if not isinstance(j_formula, str):
        return 1.0
    m = re.search(r"\*\s*(\d+(?:\.\d+)?)\s*$", j_formula)
    if m:
        v = float(m.group(1))
        return int(v) if v == int(v) else v
    return 1.0


def _detect_sheet_totals_column(ws_d, ws_f) -> tuple[int, int]:
    """Detect the totals column and flag column for this sheet.

    Most sheets use the standard layout (totals in column J, flag in column I).
    Wide-format sheets add SRD/DRD variant columns between J and the final
    total (e.g. column U), with the flag shifted left accordingly.

    Returns (totals_col, flag_col) as 1-based column indices.
    Standard: (10, 9)  →  columns J, I
    Wide:     (21, 20) →  columns U, T  (example)

    Detection strategy:
    1. Find the GRAND TOTAL row (scan bottom-up for the label).
    2. In that row scan right-to-left for the rightmost =SUM(…) formula —
       the grand-total cell always aggregates section totals via a SUM formula.
    3. Fallback: rightmost non-zero numeric value in that row.
    4. Fallback: default (10, 9).
    """
    max_scan_col = min(ws_d.max_column or 30, 40)
    for r in range(ws_d.max_row, max(1, ws_d.max_row - 50), -1):
        label_found = False
        for c in range(1, max_scan_col + 1):
            v = ws_d.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "GRAND TOTAL":
                label_found = True
                break
        if not label_found:
            continue
        # Primary: rightmost =SUM(…) formula in this row
        for cc in range(max_scan_col, 0, -1):
            fv = ws_f.cell(r, cc).value
            if isinstance(fv, str) and fv.strip().lstrip().upper().startswith("=SUM"):
                return cc, cc - 1
        # Fallback: rightmost non-zero numeric value
        for cc in range(max_scan_col, 0, -1):
            val = _safe_float(ws_d.cell(r, cc).value)
            if val and val > 0:
                return cc, cc - 1
        break  # found label row but no numeric — stop looking
    return 10, 9  # default: column J, column I


def _find_section_total(ws_d, ws_f, start: int, end: int,
                        grand_total_row: Optional[int],
                        totals_col: int = 10) -> tuple[Optional[int], float, float]:
    """Return (total_row, excel_section_total, multiplier).

    Section contribution to the grand total is always computed by summing
    cached totals-column values in the section range (this matches Excel's
    =SUM(col_x:col_y) grand-total formula and handles sheets where a section
    has no explicit TOTAL row, or where the total is spread across two rows
    via SRD/DRD conditional flags).

    `total_row` and `multiplier` are still extracted (for display / runtime
    scaling) by locating a row that holds "TOTAL" text in col A/D/G.
    """
    upper = end
    if grand_total_row and grand_total_row <= end:
        upper = grand_total_row - 1

    total_row = None
    multiplier = 1.0
    for r in range(start + 1, upper + 1):
        for c in (1, 4, 7):
            v = ws_d.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "TOTAL":
                total_row = r
                break
        if total_row:
            j_raw = ws_f.cell(total_row, totals_col).value
            multiplier = _extract_multiplier(str(j_raw or ""))
            # Some sheets store the multiplier on the row just after TOTAL
            if multiplier == 1.0:
                for r2 in (total_row + 1, total_row + 2):
                    jr2 = ws_f.cell(r2, totals_col).value
                    m = _extract_multiplier(str(jr2 or ""))
                    if m != 1.0:
                        multiplier = m
                        break
            break

    # Sum the totals column across the section range (mirrors the grand-total
    # =SUM formula so we always match what Excel reports).
    section_total = 0.0
    for r in range(start + 1, upper + 1):
        v = _safe_float(ws_d.cell(r, totals_col).value)
        if v:
            section_total += v

    return total_row, section_total, multiplier


def _find_grand_total(ws_d, ws_f, totals_col: int = 10) -> tuple[Optional[int], Optional[str], Optional[float]]:
    col_letter = get_column_letter(totals_col)
    for r in range(ws_d.max_row, max(1, ws_d.max_row - 30), -1):
        for c in range(1, min(ws_d.max_column or 30, 40) + 1):
            v = ws_d.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "GRAND TOTAL":
                gt_val = _safe_float(ws_d.cell(r, totals_col).value)
                return r, f"{col_letter}{r}", gt_val
    return None, None, None


# Column-header tokens seen in sheets.
HEADER_TOKENS = {"LENGTH", "WIDTH", "HEIGHT", "M2", "PRICE", "TOTAL",
                 "TOTAL MTR", "QUANTI", "QUANTITY", "QTY"}


def _detect_section_columns(ws_d, start: int, end: int,
                            totals_col: int = 10) -> dict:
    """Detect the column layout by finding a header row near `start`.

    Returns a dict with keys: layout ('area'|'count'), price_col, total_col,
    qty_col, header_row.  Falls back to area layout when no header row is
    found within a few rows of the section start.

    `totals_col` is the sheet-level detected total column (e.g. 10=J for
    standard sheets, 21=U for wide sheets with SRD/DRD variant columns).
    The scan extends to include all columns up to totals_col so that TOTAL
    headers placed beyond column J are correctly detected.
    """
    scan_limit = max(11, totals_col + 2)
    for r in range(start + 1, min(start + 6, end + 1)):
        row_headers: dict[str, int] = {}
        # Track all TOTAL tokens; on wide sheets there may be several — keep
        # the rightmost one (the final per-row total, not a variant sub-total).
        total_candidates: list[int] = []
        for c in range(1, scan_limit):
            v = ws_d.cell(r, c).value
            if isinstance(v, str):
                txt = v.strip().upper()
                if txt in HEADER_TOKENS:
                    if txt in ("TOTAL", "TOTAL MTR"):
                        total_candidates.append(c)
                    else:
                        row_headers[txt] = c
        if total_candidates:
            # Prefer the sheet-level totals_col if it falls within the
            # candidates, otherwise take the rightmost detected TOTAL header.
            if totals_col in total_candidates:
                row_headers["TOTAL"] = totals_col
            else:
                row_headers["TOTAL"] = total_candidates[-1]
        if len(row_headers) < 2:
            continue
        # Count layout: QUANTI column present
        if "QUANTI" in row_headers or "QUANTITY" in row_headers or "QTY" in row_headers:
            qty_col = row_headers.get("QUANTI") or row_headers.get("QUANTITY") or row_headers.get("QTY")
            return {
                "layout":    "count",
                "qty_col":   qty_col,
                "price_col": row_headers.get("PRICE", qty_col + 1),
                "total_col": row_headers.get("TOTAL", totals_col),
                "header_row": r,
            }
        # Area layout: M2 + PRICE + TOTAL
        return {
            "layout":    "area",
            "qty_col":   row_headers.get("M2", 6),
            "price_col": row_headers.get("PRICE", 7),
            "total_col": row_headers.get("TOTAL", totals_col),
            "header_row": r,
        }
    # Default to area layout; use totals_col for the total column
    return {"layout": "area", "qty_col": 6, "price_col": 7,
            "total_col": totals_col, "header_row": None}


def _rename_door_fittings(sections: list[tuple[int, str]]) -> list[tuple[int, str, str]]:
    """Tag each Column-B label with a display name.  Consecutive DOOR FITTINGS
    occurrences are prefixed with SRD / DRD based on the preceding header."""
    out: list[tuple[int, str, str]] = []
    last_door_kind = None
    for row, raw in sections:
        up = raw.upper()
        if up == "SRD":
            last_door_kind = "SRD"
            out.append((row, raw, raw))
        elif up == "DRD":
            last_door_kind = "DRD"
            out.append((row, raw, raw))
        elif up == "DOOR FITTINGS":
            display = f"{last_door_kind} DOOR FITTINGS" if last_door_kind else raw
            out.append((row, raw, display))
        else:
            out.append((row, raw, raw))
    return out


def parse_sheet(sheet_name: str,
                grp_path: str = DEFAULT_GRP_PATH,
                formulas_path: str = DEFAULT_FORMULAS_PATH) -> ParsedSheet:
    wb_d = openpyxl.load_workbook(grp_path, data_only=True)
    wb_f = openpyxl.load_workbook(grp_path, data_only=False)
    ws_d = wb_d[sheet_name]
    ws_f = wb_f[sheet_name]

    warnings: list[str] = []

    # Header metadata
    trailer_hint = str(ws_d["E1"].value or "").strip()
    size_rows = _find_size_rows(ws_d)
    length = _safe_float(ws_d.cell(size_rows["length"], 3).value) if "length" in size_rows else None
    width  = _safe_float(ws_d.cell(size_rows["width"], 3).value)  if "width"  in size_rows else None
    height = _safe_float(ws_d.cell(size_rows["height"], 3).value) if "height" in size_rows else None
    markup = _safe_float(ws_d["G5"].value)

    # Trailer ratios: column G rows 8–18 (only populated numeric cells).
    # Different sheets put the ratios at different starting rows — some use
    # only G8, others fill G9–G17 (or up to G18).  Reading the full 8–18
    # band captures every sheet's defaults.  The dropdown label is the
    # numeric G-cell value itself (not the column-A text).
    constants: list[dict] = []
    for r in range(8, 19):
        g_val = _safe_float(ws_d.cell(r, 7).value)
        if g_val is None or g_val == 0:
            continue
        # Format with up to 4 decimals, trimming trailing zeros for cleanliness.
        label = f"{g_val:.4f}".rstrip("0").rstrip(".") or "0"
        constants.append({"label": label, "value": g_val,
                          "c_value": _safe_float(ws_d.cell(r, 3).value),
                          "cell": f"G{r}"})

    cell_var_map = _build_cell_var_map(size_rows)
    size_row_set = set(size_rows.values())

    # Section boundaries
    raw_sections = _find_column_b_sections(ws_d, size_row_set)
    labelled = _rename_door_fittings(raw_sections)

    # Detect the totals and flag columns once for the whole sheet.
    # Standard layout: totals=J (10), flag=I (9).
    # Wide-format sheets (e.g. extra SRD/DRD variant columns): totals=U (21),
    # flag=T (20), etc.  All downstream functions receive these values so no
    # column index is hardcoded.
    totals_col, flag_col = _detect_sheet_totals_column(ws_d, ws_f)
    if totals_col != 10:
        warnings.append(
            f"Wide-format sheet detected: totals in column "
            f"{get_column_letter(totals_col)}, flag in column "
            f"{get_column_letter(flag_col)} (not the standard J/I layout)."
        )

    grand_total_row, grand_total_cell, grand_total_excel = _find_grand_total(ws_d, ws_f, totals_col)

    parsed_sections: list[ParsedSection] = []
    parsed_body_options: list[ParsedItem] = []
    # group → [body option items] — populated when BODY OPTIONS section is parsed
    # and then used by subsequent section iterations to link section items.
    _bo_by_group: dict[str, list] = {}
    skipped: list[str] = []

    for idx, (start_row, raw_name, display_name) in enumerate(labelled):
        next_start = (labelled[idx + 1][0] - 1
                      if idx + 1 < len(labelled) else ws_d.max_row)
        # Cap at grand-total row so last section doesn't swallow it.
        end_row = next_start
        if grand_total_row and grand_total_row <= end_row:
            end_row = grand_total_row - 1

        # BODY OPTIONS section: parse before the sec_total == 0 skip (it always
        # has a zero J-column total in the workbook).
        if "BODY OPT" in raw_name.upper():
            opts = _parse_body_options(ws_d, start_row, end_row)
            parsed_body_options.extend(opts)
            # Build group lookup so subsequent sections can link their items.
            for _opt in opts:
                _grp = _opt.body_option_group or "MISC"
                _bo_by_group.setdefault(_grp, []).append(_opt)
            continue

        total_row, sec_total, multiplier = _find_section_total(
            ws_d, ws_f, start_row, end_row, grand_total_row, totals_col
        )

        # Import all sections regardless of J-column total so that alternate
        # body-option sections (e.g. DRD when SRD is active, or vice versa)
        # are available for the user to switch to at costing time.
        # Genuinely empty sections produce no items and are harmless.

        cols = _detect_section_columns(ws_d, start_row, end_row, totals_col)
        price_col = cols["price_col"]
        total_col = cols["total_col"]
        qty_col   = cols["qty_col"]
        header_row = cols["header_row"]
        layout = cols["layout"]

        section = ParsedSection(
            name=display_name,
            raw_name=raw_name,
            start_row=start_row,
            end_row=end_row,
            total_row=total_row,
            multiplier=multiplier,
            excel_total=sec_total,
        )

        # Body-option group that controls this section (e.g. "FRONT" for the FRONT section).
        sec_grp = _section_to_body_opt_group(display_name)
        sec_body_opts = _bo_by_group.get(sec_grp, []) if sec_grp else []

        # For DRD/SRD sections every item is unconditionally linked to the group
        # name (not scored by name, because items like "DOOR HINGE" don't mention
        # DRD/SRD).  The group name is checked against selectedGroups in the UI so
        # items show whenever any option in that group is selected.
        # forced_link short-circuits the normal I-flag / name-scoring logic.
        forced_link: Optional[str] = None
        up_disp = display_name.upper().strip()
        if up_disp.startswith("DRD"):
            forced_link = "DRD"
        elif up_disp.startswith("SRD"):
            forced_link = "SRD"

        for r in range(start_row + 1, end_row + 1):
            if r == total_row or r == header_row:
                continue
            a_val = ws_d.cell(r, 1).value
            if not isinstance(a_val, str) or not a_val.strip():
                continue
            name = a_val.strip()
            up = name.upper()
            if up in SIZE_LABELS or up in ("ITEM", "TOTAL", "QUANTITY", "QUANTI"):
                continue

            cached_price = _safe_float(ws_d.cell(r, price_col).value) or 0.0
            total  = _safe_float(ws_d.cell(r, total_col).value)
            qty    = _safe_float(ws_d.cell(r, qty_col).value)
            i_cached = _safe_float(ws_d.cell(r, flag_col).value)

            # Resolve the price formula ourselves whenever one exists. Excel's
            # cached value can be stale (e.g. cross-workbook refs not refreshed
            # before save), so the freshly-evaluated formula result — with all
            # multipliers applied — is preferred. Cached value is the fallback.
            price_formula_raw = ws_f.cell(r, price_col).value
            resolved_price = _resolve_price_formula(
                price_formula_raw, ws_d, formulas_path
            )
            if resolved_price and resolved_price > 0:
                price = resolved_price
            else:
                price = cached_price

            # Determine body-option linkage for this item.
            # Rules:
            #  • forced_link set  → entire section belongs to one specific option
            #    (DRD or SRD); bypass I-flag and name scoring entirely.
            #  • col I = 0  → alternative variant. Include it if a body-option
            #    group exists for this section; link by name match, fall back to
            #    the first alternative option. Skip if no group at all.
            #  • col I = 1  → check name against all options in the group. Only
            #    link those that clearly belong to a specific option (e.g. 'PU' →
            #    'FRONT PU'); items with no name match are always-present and get
            #    no link. This prevents structural parts from vanishing when the
            #    user switches to EPS.
            bo_linked: Optional[str] = None
            if forced_link:
                # Try name scoring first (keeps EPS/PU items on their specific option);
                # fall back to the group name for structural items with no name match.
                scored = _item_body_opt_link(name, sec_body_opts, fallback_to_alt=False) if sec_body_opts else None
                bo_linked = scored if scored else forced_link
            elif i_cached is not None and i_cached == 0:
                if sec_body_opts and any(ch.isalpha() for ch in name):
                    bo_linked = _item_body_opt_link(
                        name, sec_body_opts, fallback_to_alt=True
                    )
                    if not bo_linked:
                        warnings.append(
                            f"Row {r} ({display_name}): '{name}' skipped — "
                            "col I=0 and no matching body option found."
                        )
                        continue
                    # Fall through: include this alternative-variant item.
                else:
                    if any(ch.isalpha() for ch in name):
                        warnings.append(
                            f"Row {r} ({display_name}): '{name}' skipped — "
                            f"column {get_column_letter(flag_col)} flag is 0 (disabled in sheet)."
                        )
                    continue
            elif sec_body_opts:
                # col I = 1: link only if the item name clearly belongs to a
                # specific body option. No match → always-present, no link.
                bo_linked = _item_body_opt_link(
                    name, sec_body_opts, fallback_to_alt=False
                )

            # Skip rows with neither a price nor a total AND no body-option link
            # (linked alternative items may have cached price 0 when disabled).
            if price == 0 and (total is None or total == 0) and bo_linked is None:
                if any(ch.isalpha() for ch in name):
                    warnings.append(f"Row {r} ({display_name}): '{name}' skipped — both price (col {get_column_letter(price_col)}) and total (col {get_column_letter(total_col)}) are blank or zero.")
                continue

            total_formula = ws_f.cell(r, total_col).value
            price_formula = ws_f.cell(r, price_col).value

            # Formula Skin references on the price cell are NOT expanded into
            # child items — the parent row is kept as a single BOM line using
            # the cached price (resolved by Excel from the SKINS sheet) and
            # this row's own total.  Tracked as a SKIN-tagged item for the UI.
            skin_ref = (_is_formula_skin_ref(price_formula)
                        if isinstance(price_formula, str) else None)
            is_skin_parent = bool(skin_ref)
            skin_source = (f"FORMULA SKINS!{skin_ref[0]}{skin_ref[1]}"
                           if skin_ref else None)

            # Translate the total-col formula to a symbolic QUANTITY-ONLY
            # expression.  The unit-price cell (price_col, row) is factored
            # out so runtime can multiply by material.price_per_unit.
            price_cell_key = f"{get_column_letter(price_col)}{r}"
            symbolic = None

            # Hardcoded rule: 130*62MM TAPPING BLOCKS — qty = total / 68.
            # "*" in the source sheet is the literal char, not multiplication.
            _norm_name = re.sub(r"\s+", " ", (name or "").upper()).strip()
            if _norm_name == "130*62MM TAPPING BLOCKS" and total:
                symbolic = _fmt_num(round(total / 68, 6))

            if not symbolic and isinstance(total_formula, str) and total_formula.startswith("="):
                symbolic = _translate_formula(
                    total_formula, ws_d, ws_f, cell_var_map,
                    exclude_cells={price_cell_key},
                )
                if symbolic:
                    symbolic = _simplify_mul_by_one(symbolic)
            if not symbolic:
                if price and total:
                    symbolic = _fmt_num(round(total / price, 6))
                elif qty:
                    symbolic = _fmt_num(qty)
                elif total and price:
                    symbolic = _fmt_num(round(total / price, 6))
                else:
                    symbolic = "1"

            section.items.append(ParsedItem(
                name=name,
                section=display_name,
                qty=qty,
                unit_price=price,
                excel_total=total,
                excel_formula=str(total_formula) if isinstance(total_formula, str) else None,
                symbolic_formula=symbolic,
                source_cell=skin_source or f"{chr(64+total_col)}{r}",
                is_enabled=True,
                is_formula_skin=is_skin_parent,
                highlight_color="red" if is_skin_parent else None,
                notes=(f"FORMULA SKINS ref · layout={layout}"
                       if is_skin_parent else f"layout={layout}"),
                body_option_linked=bo_linked,
            ))

        parsed_sections.append(section)

    # Each section's excel_total is the J-cell value — multipliers are
    # already baked in by Excel.  Don't re-apply them here.
    computed_total = sum(s.excel_total for s in parsed_sections)

    wb_d.close()
    wb_f.close()

    return ParsedSheet(
        sheet_name=sheet_name,
        trailer_type_hint=trailer_hint,
        length=length,
        width=width,
        height=height,
        markup=markup,
        constants=constants,
        grand_total_cell=grand_total_cell,
        grand_total_excel=grand_total_excel,
        computed_total=round(computed_total, 2),
        sections=parsed_sections,
        skipped_sections=skipped,
        warnings=warnings,
        body_options=parsed_body_options,
    )


def parsed_to_dict(ps: ParsedSheet) -> dict:
    return asdict(ps)


# ── Body-options helpers ──────────────────────────────────────────────────

# Maps name prefixes → mutual-exclusion group name.  Order matters: longer /
# more-specific prefixes should come first so shorter ones don't shadow them.
_BODY_OPT_PREFIXES: list[tuple[str, str]] = [
    ("1ST ROW",   "FLOOR"),
    ("2ND ROW",   "FLOOR"),
    ("3CR12",     "FLOOR"),
    ("18MM",      "FLOOR"),
    ("12MM",      "FLOOR"),
    ("19MM",      "FLOOR"),
    ("BIRCH",     "FLOOR"),
    ("FINN",      "FLOOR"),
    ("PHENO",     "FLOOR"),
    ("KICK",      "FLOOR"),
    ("TREAD",     "FLOOR"),
    ("RICE",      "FLOOR"),
    ("ALU",       "FLOOR"),
    ("FRONT",     "FRONT"),
    ("DRD",       "DRD"),
    ("SRD",       "SRD"),
    ("SIDES",     "SIDES"),
    ("SIDE",      "SIDES"),
    ("ROOF",      "ROOF"),
    ("FLOOR",     "FLOOR"),
]


# Same prefix table used both for option-group inference and for mapping
# section names → the body-option group that controls them.
_SECTION_OPT_GROUP_PREFIXES: list[tuple[str, str]] = [
    ("FRONT",  "FRONT"),
    ("DRD",    "DRD"),
    ("SRD",    "SRD"),
    ("SIDES",  "SIDES"),
    ("SIDE",   "SIDES"),
    ("ROOF",   "ROOF"),
    ("FLOOR",  "FLOOR"),
]


def _section_to_body_opt_group(section_name: str) -> Optional[str]:
    """Return the body-option group that controls this section, or None."""
    up = section_name.upper().strip()
    for prefix, grp in _SECTION_OPT_GROUP_PREFIXES:
        if up.startswith(prefix):
            return grp
    return None


# Words that appear in every option name for a group and carry no discriminating
# signal (e.g. "FRONT" appears in "FRONT EPS" and "FRONT PU" equally, so it
# tells us nothing about which option an item belongs to).
_BODY_OPT_LINK_STOP_WORDS = frozenset({
    "FRONT", "DRD", "SRD", "SIDES", "SIDE", "ROOF", "FLOOR",
    "BODY", "OPTIONS", "AND", "THE", "OF", "WITH",
})


def _item_body_opt_link(item_name: str,
                         body_opts: "list[ParsedItem]",
                         fallback_to_alt: bool = False) -> "Optional[str]":
    """Return the body-option name this section item belongs to, or None (always shown).

    Uses best-match scoring: count how many key words from each option name
    appear in the item name, then return the option with the most hits.
    Ties are broken by the option that appears first.
    """
    item_words = set(re.sub(r"\W+", " ", item_name.upper()).split())
    best_name: Optional[str] = None
    best_score = 0
    for opt in body_opts:
        opt_words = set(re.sub(r"\W+", " ", opt.name.upper()).split())
        key_words = opt_words - _BODY_OPT_LINK_STOP_WORDS
        if not key_words:
            continue
        score = sum(
            1
            for kw in key_words
            for iw in item_words
            if len(kw) >= 2 and len(iw) >= 2 and (kw in iw or iw in kw)
        )
        if score > best_score:
            best_score = score
            best_name = opt.name
    if best_score > 0:
        return best_name
    if fallback_to_alt:
        alts = [o for o in body_opts if not o.body_option_default]
        return alts[0].name if alts else None
    return None


def _infer_body_option_group(name: str, last_group: str) -> str:
    up = name.upper().strip()
    for prefix, group in _BODY_OPT_PREFIXES:
        if up.startswith(prefix):
            return group
    return last_group or "MISC"


# Patterns for assigning a sub-group label within a body-option group.
# Checked in order; first match wins.  The sub-group name is what the user sees
# in the Body Options panel as a labelled cluster (radio if >1 item shares it,
# toggle checkbox if only 1 item).
def _infer_body_option_subgroup(name: str) -> Optional[str]:
    """Return a sub-group label for a body-option item based on its name.

    Items sharing the same sub-group become a radio set (pick one).
    Items alone in their sub-group become standalone toggle checkboxes.
    Kick plates get unique per-item sub-groups so they are independent toggles.
    """
    u = name.upper()
    if any(k in u for k in ("EPS", "PU", "INSULATION")):
        return "INSULATION"
    if any(k in u for k in ("BIRCH", "FINN", "PLYWOOD", "PLYWWOD", "SHATTER", "PHENO")):
        return "PLYWOOD"
    if any(k in u for k in ("RICE", "ALU FLOOR", "ALU EXTRUSION", "ALU EXTRUTION",
                              "54199", "VINYL", "TREAD")):
        return "SURFACE"
    # Kick plates: give each its own unique sub-group so they're independent toggles
    if "KICK" in u or "3CR12" in u or "3CR 12" in u:
        # Use the full name (normalised) as the sub-group key → always a single-item group
        slug = re.sub(r"\W+", " ", u).strip()
        return f"KICK - {slug}"
    if any(k in u for k in ("BOLT", "WASHER", "NUT", "SCREW", "ANCHOR")):
        return "HARDWARE"
    return None


def _parse_body_options(ws_d, start_row: int, end_row: int) -> list[ParsedItem]:
    """Parse the BODY OPTIONS section from the Excel sheet.

    Layout (per Burt's workbook):
      col A = item name
      col D = Y / N  (default selection)
      col G = unit price
    Items within the same inferred group are mutually exclusive (radio-button
    behaviour in the UI).
    """
    items: list[ParsedItem] = []
    last_group = "MISC"
    for r in range(start_row + 1, end_row + 1):
        a_val = ws_d.cell(r, 1).value
        if not isinstance(a_val, str) or not a_val.strip():
            continue
        name = a_val.strip()
        up = name.upper()
        if up in SIZE_LABELS or up in ("ITEM", "TOTAL", "QUANTITY", "QUANTI", "BODY OPTIONS"):
            continue

        d_val = ws_d.cell(r, 4).value
        is_default = str(d_val or "").strip().upper() == "Y"

        unit_price = _safe_float(ws_d.cell(r, 7).value) or 0.0

        group = _infer_body_option_group(name, last_group)
        last_group = group
        subgroup = _infer_body_option_subgroup(name)

        items.append(ParsedItem(
            name=name,
            section="BODY OPTIONS",
            qty=None,
            unit_price=unit_price,
            excel_total=None,
            excel_formula=None,
            symbolic_formula="1",
            source_cell=f"G{r}",
            is_enabled=True,
            is_body_option=True,
            body_option_group=group,
            body_option_subgroup=subgroup,
            body_option_default=is_default,
        ))
    return items


# ── DB-writer (replaces the old import_excel_sheet.import_sheet) ───────────

def list_sheets(grp_path: str = DEFAULT_GRP_PATH) -> list[str]:
    """Return all sheet names from the source workbook."""
    wb = openpyxl.load_workbook(grp_path, data_only=True, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _material_category_name(section: str, is_formula_skin: bool) -> str:
    """Map a section name to a MaterialCategory display name.  Formula-skin
    children always live in the Resins & Adhesives category."""
    if is_formula_skin:
        return "Resins & Adhesives"
    return section  # one category per section (same as the richer /api/import/execute path)


def import_sheet(db, trailer_name: str, sheet_name: str,
                 overwrite: bool = False,
                 grp_path: str = DEFAULT_GRP_PATH,
                 formulas_path: str = DEFAULT_FORMULAS_PATH) -> dict:
    """Parse `sheet_name` from the GRP workbook and write the result to the
    database.  Returns a summary dict compatible with the existing admin
    import page."""
    # Deferred imports so this module stays importable outside the web app.
    try:
        from app.database import (
            Material, MaterialCategory, TrailerType, BillOfMaterial,
            TrailerRatio, BOMSection, BodyOptionGroup, BodyOptionSubgroup,
        )
    except ImportError:  # when run from app/ directory
        from database import (  # type: ignore
            Material, MaterialCategory, TrailerType, BillOfMaterial,
            TrailerRatio, BOMSection, BodyOptionGroup, BodyOptionSubgroup,
        )
    from datetime import datetime, timezone

    ps = parse_sheet(sheet_name, grp_path=grp_path, formulas_path=formulas_path)

    # ── TrailerType: create or overwrite ─────────────────────────────────
    existing = db.query(TrailerType).filter_by(name=trailer_name,
                                               is_active=True).first()
    if existing:
        if not overwrite:
            raise ValueError(
                f'Trailer type "{trailer_name}" already exists. '
                f"Use overwrite=True to replace its BOM."
            )
        # Wipe its BOM and ratios, keep the row itself
        db.query(BillOfMaterial).filter_by(trailer_type_id=existing.id).delete()
        db.query(TrailerRatio).filter_by(trailer_type_id=existing.id).delete()
        tt = existing
    else:
        tt = TrailerType(
            name=trailer_name,
            description=f"Imported from {sheet_name}",
            is_active=True,
        )
        db.add(tt)
        db.flush()
        # Restore any prior template binding archived under this name
        # (covers the soft-delete → re-import case).
        try:
            from .main import restore_orphan_for_trailer
            restore_orphan_for_trailer(tt, db)
        except Exception as _e:
            # Non-fatal: import should not fail because of restore plumbing.
            print(f"Template-binding restore skipped: {_e}")

    tt.default_length = ps.length
    tt.default_width  = ps.width
    tt.default_height = ps.height
    if ps.markup is not None:
        tt.markup_percentage = ps.markup

    # ── Trailer constants (A8:A18 / C8:C18) → TrailerRatio ────────────────
    for i, c in enumerate(ps.constants):
        db.add(TrailerRatio(
            trailer_type_id=tt.id,
            ratio_value=c["value"],
            label=c["label"],
            sort_order=i,
        ))

    items_created = 0
    materials_created = 0
    sort_idx = 0

    # ── Sections + items ──────────────────────────────────────────────────
    for s_idx, section in enumerate(ps.sections):
        # Register / update the section with its multiplier
        bom_sec = db.query(BOMSection).filter_by(name=section.name).first()
        if not bom_sec:
            bom_sec = BOMSection(name=section.name, sort_order=s_idx,
                                 multiplier=section.multiplier)
            db.add(bom_sec)
        else:
            bom_sec.multiplier = section.multiplier
        db.flush()

        for item in section.items:
            cat_name = _material_category_name(section.name, item.is_formula_skin)
            cat = db.query(MaterialCategory).filter_by(name=cat_name).first()
            if not cat:
                cat = MaterialCategory(name=cat_name)
                db.add(cat)
                db.flush()

            mat = (db.query(Material)
                   .filter_by(name=item.name, category_id=cat.id, is_active=True)
                   .first())
            if not mat:
                mat = Material(
                    name=item.name,
                    category_id=cat.id,
                    price_per_unit=item.unit_price,
                    unit_of_measure="each",
                    last_updated=datetime.now(timezone.utc),
                    is_active=True,
                )
                db.add(mat)
                db.flush()
                materials_created += 1
            else:
                # Refresh price on every import so re-imports pick up the
                # latest calculated unit price (e.g. resolved formula values
                # that bypass stale Excel caches).
                if item.unit_price and item.unit_price > 0 and \
                   abs((mat.price_per_unit or 0) - item.unit_price) > 1e-6:
                    mat.price_per_unit = item.unit_price
                    mat.last_updated = datetime.now(timezone.utc)

            db.add(BillOfMaterial(
                trailer_type_id=tt.id,
                material_id=mat.id,
                formula_expression=item.symbolic_formula or "1",
                waste_percentage=0.0,
                notes=item.notes or section.name,
                bom_section=section.name,
                bom_section_id=bom_sec.id,
                sort_order=sort_idx,
                excel_formula=item.excel_formula,
                unit_price_snapshot=item.unit_price,
                source_cell=(item.source_cell or "")[:64],
                is_formula_skin=item.is_formula_skin,
                highlight_color=item.highlight_color,
                body_option_linked=item.body_option_linked,
            ))
            sort_idx += 1
            items_created += 1

    # ── Body options ──────────────────────────────────────────────────────────
    if ps.body_options:
        body_opt_cat = db.query(MaterialCategory).filter_by(name="BODY OPTIONS").first()
        if not body_opt_cat:
            body_opt_cat = MaterialCategory(name="BODY OPTIONS")
            db.add(body_opt_cat)
            db.flush()

        body_opt_sec = db.query(BOMSection).filter_by(name="BODY OPTIONS").first()
        if not body_opt_sec:
            body_opt_sec = BOMSection(name="BODY OPTIONS", sort_order=999)
            db.add(body_opt_sec)
            db.flush()

        # Cache group/subgroup IDs for this import to avoid repeated lookups
        _grp_id_cache: dict[str, int] = {}
        _sub_id_cache: dict[tuple, int] = {}

        def _get_grp_id(gname: str) -> int | None:
            if not gname:
                return None
            up = gname.upper()
            if up not in _grp_id_cache:
                g = db.query(BodyOptionGroup).filter_by(name=up).first()
                if not g:
                    g = BodyOptionGroup(name=up, sort_order=0)
                    db.add(g)
                    db.flush()
                _grp_id_cache[up] = g.id
            return _grp_id_cache[up]

        def _get_sub_id(gid: int | None, sname: str | None) -> int | None:
            if not gid or not sname:
                return None
            up = sname.upper()
            key = (gid, up)
            if key not in _sub_id_cache:
                s = db.query(BodyOptionSubgroup).filter_by(group_id=gid, name=up).first()
                if not s:
                    s = BodyOptionSubgroup(group_id=gid, name=up, sort_order=0)
                    db.add(s)
                    db.flush()
                _sub_id_cache[key] = s.id
            return _sub_id_cache[key]

        for opt in ps.body_options:
            mat = (db.query(Material)
                   .filter_by(name=opt.name, category_id=body_opt_cat.id, is_active=True)
                   .first())
            if not mat:
                mat = Material(
                    name=opt.name,
                    category_id=body_opt_cat.id,
                    price_per_unit=opt.unit_price,
                    unit_of_measure="each",
                    last_updated=datetime.now(timezone.utc),
                    is_active=True,
                )
                db.add(mat)
                db.flush()
                materials_created += 1
            else:
                if opt.unit_price and opt.unit_price > 0 and \
                   abs((mat.price_per_unit or 0) - opt.unit_price) > 1e-6:
                    mat.price_per_unit = opt.unit_price
                    mat.last_updated = datetime.now(timezone.utc)

            grp_id = _get_grp_id(opt.body_option_group)
            sub_id = _get_sub_id(grp_id, opt.body_option_subgroup)
            db.add(BillOfMaterial(
                trailer_type_id=tt.id,
                material_id=mat.id,
                formula_expression="1",
                waste_percentage=0.0,
                notes="BODY OPTIONS",
                bom_section="BODY OPTIONS",
                bom_section_id=body_opt_sec.id,
                sort_order=sort_idx,
                unit_price_snapshot=opt.unit_price,
                source_cell=(opt.source_cell or "")[:64],
                is_body_option=True,
                body_option_group=opt.body_option_group,
                body_option_group_id=grp_id,
                body_option_subgroup=opt.body_option_subgroup,
                body_option_subgroup_id=sub_id,
                body_option_default=opt.body_option_default,
            ))
            sort_idx += 1
            items_created += 1

    db.commit()

    # Second-pass: resolve body_option_linked string names → body_option_linked_id FK.
    # Done after commit so all body-option materials created above are visible.
    link_resolved = 0
    link_unresolved = []
    linked_rows = (db.query(BillOfMaterial)
                   .filter(BillOfMaterial.trailer_type_id == tt.id,
                           BillOfMaterial.body_option_linked.isnot(None),
                           BillOfMaterial.body_option_linked != "",
                           BillOfMaterial.body_option_linked_id.is_(None))
                   .all())
    for bom_row in linked_rows:
        lm = db.query(Material).filter_by(name=bom_row.body_option_linked).first()
        if lm:
            bom_row.body_option_linked_id = lm.id
            link_resolved += 1
        else:
            link_unresolved.append(bom_row.body_option_linked)
    if linked_rows:
        db.commit()
    if link_unresolved:
        for name in set(link_unresolved):
            ps.warnings.append(
                f"body_option_linked '{name}' not found in materials — "
                "link kept as string only (group-level or missing material)"
            )

    return {
        "trailer_id":        tt.id,
        "trailer_name":      tt.name,
        "bom_items":         items_created,
        "new_materials":     materials_created,
        "markup_percentage": tt.markup_percentage,
        "ratios":            [c["value"] for c in ps.constants],
        "sheet":             sheet_name,
        # Extras the new UI can surface
        "dimensions":        {"length": ps.length, "width": ps.width, "height": ps.height},
        "grand_total_excel": ps.grand_total_excel,
        "computed_total":    ps.computed_total,
        "skipped_sections":  ps.skipped_sections,
        "warnings":          ps.warnings,
        "sections":          [
            {"name": s.name, "multiplier": s.multiplier,
             "total": s.excel_total, "items": len(s.items)}
            for s in ps.sections
        ],
        "body_options":      len(ps.body_options),
    }


__all__ = [
    "parse_sheet", "parsed_to_dict",
    "ParsedSheet", "ParsedSection", "ParsedItem",
    "list_sheets", "import_sheet",
    "DEFAULT_GRP_PATH", "DEFAULT_FORMULAS_PATH",
]
