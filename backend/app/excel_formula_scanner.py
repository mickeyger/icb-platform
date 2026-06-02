"""
app/excel_formula_scanner.py

Scan a GRP Costings xlsx workbook to identify, for every BOM row, whether
the PRICE column ultimately resolves (directly OR through a chain of
intra-sheet cell references) to a cell in FORMULAS 2018.xls.

Used by:
  • tools/scan_formula_references.py   — CLI / CSV export
  • app/routers/import_excel.py        — admin UI

The scanner is purely read-only and has no DB dependencies. The companion
matcher in app/excel_formula_matcher.py joins these results to DB rows.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Iterable
from urllib.parse import unquote
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────

DEFAULT_GRP_PATH = (
    r"C:\Users\micge\Documents\Burt Costing Model"
    r"\Latest price list\GRP Costings 2018.xlsx"
)

SKIP_SHEETS: frozenset[str] = frozenset({
    "TRAILER UNITS SOLD", "CHASSIS COSTINGS", "SHEET1", "SHEET PLANNING",
    "VACUUM PLANNING HEIDELBERG", "SIDE TIPPER COSTINGS", "TRAILER PRICE LIST",
    "EXAMPLE",
})

# External reference token: '[<linkId>]<sheet>'!<cell>
_FORMULA_RE = re.compile(r"'\[(\d+)\]([^']+)'!\$?([A-Z]+)\$?(\d+)")

# Intra-sheet cell reference (after external tokens have been stripped out)
_CELL_REF_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")

_AGG_RE = re.compile(
    r"\b(SUM|SUMPRODUCT|VLOOKUP|HLOOKUP|INDEX|MATCH|IF|IFS|OFFSET|"
    r"AVERAGE|MAX|MIN|COUNT|COUNTIF|SUMIF)\s*\(",
    re.IGNORECASE,
)

_HEADER_KEYWORDS: frozenset[str] = frozenset({
    "WIDTH", "HEIGHT", "M2", "PRICE", "TOTAL", "TOTAL M",
    "QUANT", "QUANTI", "LENGTH", "QTY", "DESCRIPTION", "MATERIAL",
    "ITEM", "COST", "GRAND TOTAL",
})

_SIZE_LABELS: frozenset[str] = frozenset({"LENGTH", "WIDTH", "HEIGHT"})

_MAX_DEPTH = 12
_MAX_HEADER_SCAN_COL = 20  # covers chiller-style col-L offset layouts


# ── External link discovery ────────────────────────────────────────────────

def discover_formulas_link_ids(xlsx_path: str | Path) -> set[int]:
    """Return externalLink IDs in the workbook that point at FORMULAS 2018.xls.

    Reads xl/externalLinks/_rels/externalLink<N>.xml.rels directly from the
    xlsx zip — openpyxl doesn't preserve the link table.
    """
    out: set[int] = set()
    with zipfile.ZipFile(str(xlsx_path)) as z:
        for name in z.namelist():
            m = re.match(r"xl/externalLinks/_rels/externalLink(\d+)\.xml\.rels", name)
            if not m:
                continue
            link_id = int(m.group(1))
            root = ET.fromstring(z.read(name))
            for rel in root:
                target = unquote(rel.attrib.get("Target", "")).upper()
                if "FORMULAS 2018" in target:
                    out.add(link_id)
    return out


# ── Per-row PRICE column detection ─────────────────────────────────────────

def build_price_column_map(ws) -> dict[int, int]:
    """Return {row_num: price_col_idx} for every row in the sheet.

    A header row is any row that has 'PRICE' in cols 1..MAX_HEADER_SCAN_COL.
    Multiple headers per sheet are allowed — sections with different layouts
    (e.g. door-style at col E vs body at col G) get different PRICE columns.
    """
    headers: list[tuple[int, int]] = []
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        for c in range(1, _MAX_HEADER_SCAN_COL + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "PRICE":
                headers.append((r, c))
                break

    out: dict[int, int] = {}
    if not headers:
        return out

    idx = 0
    current = headers[0][1]
    for r in range(1, max_row + 1):
        while idx < len(headers) and headers[idx][0] <= r:
            current = headers[idx][1]
            idx += 1
        out[r] = current
    return out


# ── Recursive formula trace ────────────────────────────────────────────────

def trace_external_refs(
    ws,
    row: int,
    col: int,
    formulas_link_ids: set[int],
    depth: int = 0,
    seen: frozenset | None = None,
    chain: tuple[str, ...] = (),
) -> list[dict]:
    """Walk the formula tree at (row, col).

    Returns a list of result dicts:
        {kind: 'external', ref_sheet, ref_cell, raw_formula, chain}
        {kind: 'aggregate', raw_formula, chain}
        {kind: 'literal', value, chain}    — terminal non-formula cells

    Cycles and runaway depth are guarded.
    """
    if seen is None:
        seen = frozenset()
    key = (row, col)
    if key in seen or depth > _MAX_DEPTH:
        return []
    seen = seen | {key}

    cell = ws.cell(row, col)
    coord = cell.coordinate
    raw = cell.value

    if raw is None:
        return []
    if not isinstance(raw, str) or not raw.startswith("="):
        return [{"kind": "literal", "value": raw, "chain": chain + (coord,)}]

    expr = raw[1:]
    chain_here = chain + (coord,)
    results: list[dict] = []

    # Direct external refs to FORMULAS 2018.xls
    for m in _FORMULA_RE.finditer(expr):
        link_id = int(m.group(1))
        if link_id not in formulas_link_ids:
            continue
        results.append({
            "kind": "external",
            "ref_sheet": m.group(2).strip(),
            "ref_cell": f"{m.group(3)}{m.group(4)}",
            "raw_formula": raw,
            "chain": chain_here,
        })

    if _AGG_RE.search(expr):
        results.append({
            "kind": "aggregate",
            "raw_formula": raw,
            "chain": chain_here,
        })

    # Strip external tokens before scanning for intra-sheet refs so we don't
    # accidentally recurse into e.g. F24 from '[1]TAPING BLOCKS'!$F$24.
    expr_local = _FORMULA_RE.sub("", expr)
    expr_local = re.sub(r'"[^"]*"', "", expr_local)

    for m in _CELL_REF_RE.finditer(expr_local):
        sub_col = column_index_from_string(m.group(1))
        sub_row = int(m.group(2))
        if (sub_row, sub_col) == key:
            continue
        results.extend(trace_external_refs(
            ws, sub_row, sub_col, formulas_link_ids,
            depth + 1, seen, chain_here,
        ))

    return results


# ── Section header lookup ──────────────────────────────────────────────────

def section_for_row(ws, row: int) -> str:
    """Most-recent non-empty column-B value above this row."""
    for r in range(row - 1, max(0, row - 60), -1):
        b = ws.cell(r, 2).value
        if not (isinstance(b, str) and b.strip()):
            continue
        up = b.strip().upper()
        if up in _HEADER_KEYWORDS or up == "BODY OPTIONS":
            continue
        return b.strip()
    return ""


# ── Sheet & workbook scan ──────────────────────────────────────────────────

def _is_item_row(value) -> bool:
    if not isinstance(value, str):
        return False
    name = value.strip()
    if not name:
        return False
    up = name.upper()
    if up in _HEADER_KEYWORDS or up in _SIZE_LABELS:
        return False
    return True


def scan_sheet(ws, formulas_link_ids: set[int]) -> list[dict]:
    """Scan one worksheet, returning only rows whose PRICE column ultimately
    resolves to a cell in FORMULAS 2018.xls.

    A single Excel cell may produce multiple entries when its formula chain
    reaches more than one distinct external reference.
    """
    out: list[dict] = []
    price_map = build_price_column_map(ws)
    if not price_map:
        return out

    max_row = ws.max_row
    for r in range(1, max_row + 1):
        a_val = ws.cell(r, 1).value
        if not _is_item_row(a_val):
            continue
        item_name = a_val.strip()
        price_col = price_map.get(r)
        if not price_col:
            continue
        if ws.cell(r, price_col).value is None:
            continue

        results = trace_external_refs(ws, r, price_col, formulas_link_ids)
        externals = [x for x in results if x["kind"] == "external"]
        if not externals:
            continue

        had_aggregate = any(x["kind"] == "aggregate" for x in results)
        section = section_for_row(ws, r)
        price_col_letter = get_column_letter(price_col)

        seen_refs: set[tuple[str, str]] = set()
        for ext in externals:
            key = (ext["ref_sheet"], ext["ref_cell"])
            if key in seen_refs:
                continue
            seen_refs.add(key)
            chain_str = (
                " -> ".join(ext["chain"])
                + f" -> [FORMULAS 2018.xls]{ext['ref_sheet']}!{ext['ref_cell']}"
            )
            out.append({
                "section": section,
                "item": item_name,
                "row": r,
                "price_col": price_col_letter,
                "ref_sheet": ext["ref_sheet"],
                "ref_cell": ext["ref_cell"],
                "raw_formula": ext["raw_formula"],
                "chain": chain_str,
                "had_aggregate": had_aggregate,
            })

    return out


def scan_workbook(
    xlsx_path: str | Path,
    *,
    only_sheet: str | None = None,
    skip_sheets: Iterable[str] | None = None,
) -> dict:
    """Scan an entire workbook.

    Returns:
        {
            "source": absolute path,
            "formulas_link_ids": [int, ...],
            "sheets": [{"name": str, "linked_count": int}, ...],
            "linked": [row, ...],   # body_type added
            "skipped_sheets": [str, ...],
        }
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    formulas_link_ids = discover_formulas_link_ids(str(xlsx_path))
    if not formulas_link_ids:
        raise ValueError("No external link to FORMULAS 2018.xls found in workbook.")

    skip_set = frozenset(s.upper() for s in (skip_sheets or SKIP_SHEETS))

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    linked_all: list[dict] = []
    sheets_summary: list[dict] = []
    skipped: list[str] = []

    for sn in wb.sheetnames:
        if sn.strip().upper() in skip_set:
            skipped.append(sn)
            continue
        if only_sheet and sn != only_sheet:
            continue
        ws = wb[sn]
        rows = scan_sheet(ws, formulas_link_ids)
        for r in rows:
            r["body_type"] = sn
        linked_all.extend(rows)
        sheets_summary.append({"name": sn, "linked_count": len(rows)})

    return {
        "source": str(xlsx_path.resolve()),
        "formulas_link_ids": sorted(formulas_link_ids),
        "sheets": sheets_summary,
        "linked": linked_all,
        "skipped_sheets": skipped,
    }
