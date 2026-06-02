"""
app/excel_cell_resolver.py

Classify a cell in a GRP Costings worksheet by what it ultimately resolves
to. The complex sheets (RIGID DRY FREIGHT, etc.) chain intra-sheet refs
several layers deep before reaching a literal value or an external
reference, and they encode conditional inclusion as IF(Cxx="Y",1,0)
patterns. This module is the single source of truth for walking those
chains during import and during the formula-scan migration tools.

Usage
─────
    from app.excel_cell_resolver import (
        Resolver, ResolvedKind, discover_external_link_ids,
    )

    wb       = openpyxl.load_workbook(path, data_only=False)
    wb_data  = openpyxl.load_workbook(path, data_only=True)
    ext_links = discover_external_link_ids(path)
    r = Resolver(wb["RIGID DRY FREIGHT"], wb_data["RIGID DRY FREIGHT"], ext_links)

    res = r.resolve("G54")
    if res.kind == ResolvedKind.EXTERNAL_FORMULAS_2018:
        # link to skin/taping/floor/cleat via SHEET_MAP
        ...
    elif res.kind == ResolvedKind.LITERAL:
        unit_price = res.value
    elif res.kind == ResolvedKind.IF_CONDITION:
        # gate cell — res.if_test_cell is the BODY OPTIONS row to look up
        ...

Design notes
────────────
• Pure data structures + a small Resolver class. No DB writes, no I/O
  beyond the openpyxl worksheets passed in.
• Memoised per (Resolver instance, cell address). Cycle-safe via a
  per-call frozenset; depth-limited at MAX_DEPTH.
• Terminal kinds are mutually exclusive — a cell resolves to exactly
  one ResolvedKind. The Resolver picks the most specific one available.
"""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Iterable
from urllib.parse import unquote
from xml.etree import ElementTree as ET

from openpyxl.utils import column_index_from_string, get_column_letter

# ── Regexes ─────────────────────────────────────────────────────────────────

# External reference: '[<linkId>]<sheet>'!<cell>  e.g. '[2]SRD FLOOR PLATE'!$F$24
_EXT_REF = re.compile(r"'\[(\d+)\]([^']+)'!\$?([A-Z]+)\$?(\d+)")
# External reference shorthand: [<linkId>]<name>!<cell>  e.g. [1]RIVETS!$C$5
_EXT_REF_SHORT = re.compile(r"\[(\d+)\]([A-Za-z0-9 +&_/]+)!\$?([A-Z]+)\$?(\d+)")
# Intra-sheet cell reference. Used only after external tokens are stripped.
_CELL_REF = re.compile(r"\$?([A-Z]+)\$?(\d+)")
# Aggregate / lookup functions whose interior we don't expand.
_AGG_RE = re.compile(
    r"\b(SUM|SUMPRODUCT|VLOOKUP|HLOOKUP|INDEX|MATCH|IFS|OFFSET|"
    r"AVERAGE|MAX|MIN|COUNT|COUNTIF|SUMIF)\s*\(",
    re.IGNORECASE,
)
# IF(<test_cell>="<expected>",<true>,<false>)  — captures the four pieces.
# Uses string comparison so it matches the gate-cell idiom used throughout
# the workbook: =IF(C8="Y",1,0). Tolerates whitespace and trailing-space
# typos like IF(C8 ="Y",1,0).
_IF_GATE = re.compile(
    r"^\s*IF\s*\(\s*"
    r"\$?([A-Z]+)\$?(\d+)"          # test cell
    r"\s*=\s*"
    r"\"([^\"]*)\""                  # expected value (string in quotes)
    r"\s*,\s*"
    r"([^,()]+|\([^()]*\))"          # true branch (simple value or one-level paren)
    r"\s*,\s*"
    r"([^()]+|\([^()]*\))"           # false branch
    r"\s*\)\s*$",
    re.IGNORECASE,
)

MAX_DEPTH = 16


# ── Public API ─────────────────────────────────────────────────────────────

class ResolvedKind(str, Enum):
    LITERAL = "literal"
    EXTERNAL_FORMULAS_2018 = "external_formulas_2018"
    EXTERNAL_OTHER = "external_other"
    AGGREGATE = "aggregate"
    IF_CONDITION = "if_condition"
    EXPRESSION = "expression"
    EMPTY = "empty"
    UNRESOLVED = "unresolved"


@dataclass(frozen=True)
class Resolved:
    """The classification of one cell.

    Fields are populated based on `kind` — most are None for irrelevant kinds:

      LITERAL                 → value
      EXTERNAL_FORMULAS_2018  → ref_sheet, ref_cell
      EXTERNAL_OTHER          → ref_link_id, ref_sheet, ref_cell, fallback_value
      AGGREGATE               → fallback_value
      IF_CONDITION            → if_test_cell, if_expected, if_true_branch,
                                if_false_branch (raw text)
      EXPRESSION              → fallback_value (we couldn't reduce to a single
                                literal but it's not an aggregate either)
      EMPTY                   → (everything None)
      UNRESOLVED              → reason
    """
    kind: ResolvedKind
    value: float | None = None
    ref_sheet: str | None = None
    ref_cell: str | None = None
    ref_link_id: int | None = None
    if_test_cell: str | None = None
    if_expected: str | None = None
    if_true_branch: str | None = None
    if_false_branch: str | None = None
    fallback_value: float | None = None
    chain: tuple[str, ...] = field(default_factory=tuple)
    raw_formula: str | None = None
    reason: str | None = None


# ── External-link discovery ────────────────────────────────────────────────

def discover_external_link_ids(xlsx_path: str | Path) -> dict[int, str]:
    """Return {externalLink_id: target_filename} for every external reference
    in the workbook. The Resolver uses this to classify the link IDs
    (e.g. [2] → 'FORMULAS 2018.xls' → ResolvedKind.EXTERNAL_FORMULAS_2018).
    """
    out: dict[int, str] = {}
    with zipfile.ZipFile(str(xlsx_path)) as z:
        for name in z.namelist():
            m = re.match(r"xl/externalLinks/_rels/externalLink(\d+)\.xml\.rels", name)
            if not m:
                continue
            link_id = int(m.group(1))
            root = ET.fromstring(z.read(name))
            for rel in root:
                target = unquote(rel.attrib.get("Target", ""))
                out[link_id] = target.split("/")[-1]
    return out


def _classify_link(link_id: int, ext_links: dict[int, str]) -> ResolvedKind:
    target = (ext_links.get(link_id) or "").upper()
    if "FORMULAS 2018" in target:
        return ResolvedKind.EXTERNAL_FORMULAS_2018
    return ResolvedKind.EXTERNAL_OTHER


# ── Helpers ────────────────────────────────────────────────────────────────

def _safe_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_if_gate(formula: str) -> tuple[str, str, str, str] | None:
    """If `formula` matches =IF(<cell>="<v>",<t>,<f>) return
    (test_cell, expected, true_branch, false_branch). Else None.
    The leading '=' is stripped before matching; whitespace tolerated.
    """
    if not isinstance(formula, str):
        return None
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:]
    m = _IF_GATE.match(expr)
    if not m:
        return None
    col, row, expected, t, f = m.groups()
    return f"{col}{row}", expected, t.strip(), f.strip()


# ── Resolver ───────────────────────────────────────────────────────────────

class Resolver:
    """Classify cells in one worksheet (paired with its data_only twin).

    `ws_data` is the same workbook opened with data_only=True so we can
    read the cached value Excel last computed for any cell — used as a
    fallback when we can't (or don't want to) evaluate a formula
    ourselves (aggregates, external refs, complex expressions).

    A single Resolver instance is single-sheet by design; the importer
    creates one per sheet to keep the memo cache scoped.
    """

    def __init__(self, ws, ws_data, ext_links: dict[int, str]):
        self.ws = ws
        self.ws_data = ws_data
        self.ext_links = ext_links
        self._cache: dict[str, Resolved] = {}

    # ── Public ────────────────────────────────────────────────────────────

    def resolve(self, addr: str) -> Resolved:
        return self._resolve(addr.upper(), depth=0, seen=frozenset())

    # ── Core ──────────────────────────────────────────────────────────────

    def _resolve(self, addr: str, *, depth: int, seen: frozenset) -> Resolved:
        cached = self._cache.get(addr)
        if cached is not None:
            return cached
        if addr in seen:
            return self._set(addr, Resolved(
                kind=ResolvedKind.UNRESOLVED,
                reason=f"cycle through {addr}",
                chain=tuple(seen) + (addr,),
            ))
        if depth > MAX_DEPTH:
            return self._set(addr, Resolved(
                kind=ResolvedKind.UNRESOLVED,
                reason=f"depth limit ({MAX_DEPTH}) reached at {addr}",
                chain=tuple(seen) + (addr,),
            ))

        col_letters, row_num = self._split(addr)
        cell = self.ws.cell(row_num, column_index_from_string(col_letters))
        raw = cell.value
        chain = tuple(seen) + (addr,)

        # Empty cell
        if raw is None or raw == "":
            return self._set(addr, Resolved(kind=ResolvedKind.EMPTY, chain=chain))

        # Pure number
        if not isinstance(raw, str):
            num = _safe_float(raw)
            return self._set(addr, Resolved(
                kind=ResolvedKind.LITERAL, value=num, chain=chain,
            ))

        # Pure number written as a string
        if not raw.startswith("="):
            num = _safe_float(raw)
            if num is not None:
                return self._set(addr, Resolved(
                    kind=ResolvedKind.LITERAL, value=num, chain=chain,
                ))
            # Otherwise it's a text label — neither value nor formula
            return self._set(addr, Resolved(
                kind=ResolvedKind.LITERAL, value=None, chain=chain, raw_formula=raw,
            ))

        # ── It's a formula ──
        expr = raw[1:]

        # 1) IF gate idiom (most specific — has to win over EXPRESSION)
        gate = parse_if_gate(raw)
        if gate is not None:
            test_cell, expected, t_branch, f_branch = gate
            return self._set(addr, Resolved(
                kind=ResolvedKind.IF_CONDITION,
                if_test_cell=test_cell,
                if_expected=expected,
                if_true_branch=t_branch,
                if_false_branch=f_branch,
                fallback_value=_safe_float(self._data_only(addr)),
                chain=chain,
                raw_formula=raw,
            ))

        # 2) Direct external reference (single token, possibly with simple math)
        m_ext = _EXT_REF.search(expr) or _EXT_REF_SHORT.search(expr)
        if m_ext:
            link_id = int(m_ext.group(1))
            ref_sheet = m_ext.group(2).strip()
            ref_cell = f"{m_ext.group(3)}{m_ext.group(4)}"
            kind = _classify_link(link_id, self.ext_links)
            return self._set(addr, Resolved(
                kind=kind,
                ref_link_id=link_id,
                ref_sheet=ref_sheet,
                ref_cell=ref_cell,
                fallback_value=_safe_float(self._data_only(addr)),
                chain=chain,
                raw_formula=raw,
            ))

        # 3) Aggregate / lookup function
        if _AGG_RE.search(expr):
            return self._set(addr, Resolved(
                kind=ResolvedKind.AGGREGATE,
                fallback_value=_safe_float(self._data_only(addr)),
                chain=chain,
                raw_formula=raw,
            ))

        # 4) Single cell reference — chain through it
        single_ref = re.fullmatch(r"\$?([A-Z]+)\$?(\d+)", expr.strip())
        if single_ref:
            target = f"{single_ref.group(1)}{single_ref.group(2)}"
            sub = self._resolve(target, depth=depth + 1, seen=seen | {addr})
            # Extend the chain to record the hop through this cell
            extended = Resolved(
                kind=sub.kind, value=sub.value,
                ref_sheet=sub.ref_sheet, ref_cell=sub.ref_cell,
                ref_link_id=sub.ref_link_id,
                if_test_cell=sub.if_test_cell, if_expected=sub.if_expected,
                if_true_branch=sub.if_true_branch, if_false_branch=sub.if_false_branch,
                fallback_value=(sub.fallback_value
                                if sub.fallback_value is not None
                                else _safe_float(self._data_only(addr))),
                chain=(addr,) + sub.chain,
                raw_formula=raw,
                reason=sub.reason,
            )
            return self._set(addr, extended)

        # 5) Anything else — a multi-term expression. We don't try to
        #    arithmetically evaluate it here (the importer's
        #    quantity-formula translator is the right tool for that).
        #    Record the data_only cached value as a fallback so the
        #    importer can use it when it doesn't need a symbolic form.
        return self._set(addr, Resolved(
            kind=ResolvedKind.EXPRESSION,
            fallback_value=_safe_float(self._data_only(addr)),
            chain=chain,
            raw_formula=raw,
        ))

    # ── Internals ─────────────────────────────────────────────────────────

    def _set(self, addr: str, res: Resolved) -> Resolved:
        self._cache[addr] = res
        return res

    def _data_only(self, addr: str):
        col, row = self._split(addr)
        return self.ws_data.cell(row, column_index_from_string(col)).value

    @staticmethod
    def _split(addr: str) -> tuple[str, int]:
        m = re.fullmatch(r"\$?([A-Z]+)\$?(\d+)", addr.upper())
        if not m:
            raise ValueError(f"Invalid cell address: {addr!r}")
        return m.group(1), int(m.group(2))
