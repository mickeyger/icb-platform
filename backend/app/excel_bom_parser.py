"""
Shared Excel BOM parsing utilities — used by both the web app (snapshot
capture) and the standalone audit tool.

Parses GRP Costings 2018.xlsx-style workbooks: detects section headers,
extracts material name + unit price + total per item.
"""

import re
import openpyxl

_SKIP_SHEETS = {
    "TRAILER UNITS SOLD", "CHASSIS COSTINGS", "SHEET1", "SHEET PLANNING",
    "VACUUM PLANNING HEIDELBERG", "SIDE TIPPER COSTINGS", "TRAILER PRICE LIST",
}

_COL_SKIP_HEADERS = {
    "WIDTH", "HEIGHT", "M2", "PRICE", "TOTAL", "GRAND TOTAL",
    "DESCRIPTION", "QTY", "COST", "QUANT", "MATERIAL", "ITEM",
}


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).upper().strip())


def norm_key(s: str) -> str:
    return re.sub(r"[^A-Z0-9 ]", "", norm(s))


def match_score(a: str, b: str) -> float:
    """Jaccard similarity between two name strings."""
    a_w = set(norm_key(a).split())
    b_w = set(norm_key(b).split())
    if not a_w and not b_w:
        return 0.0
    return len(a_w & b_w) / len(a_w | b_w)


def best_sheet_match(sheet_names: list[str], trailer_name: str) -> tuple[str | None, float]:
    """Return (best_sheet_name, score) for a given trailer name."""
    best, best_score = None, 0.0
    for s in sheet_names:
        sc = match_score(s, trailer_name)
        if sc > best_score:
            best, best_score = s, sc
    return best, best_score


def _safe_float(cell) -> float | None:
    v = cell.value if hasattr(cell, "value") else cell
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", ".").strip())
        return None if f == 0.0 else f
    except (ValueError, TypeError):
        return None


def _is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def _cell_str(cell) -> str:
    v = cell.value
    return "" if v is None else str(v).strip()


def detect_col_offset(ws) -> int:
    """Detect CHILLER-style sheets (data starts col L not col A). Returns 0 or 11."""
    ref_errors = non_empty = 0
    for row in ws.iter_rows(min_row=26, max_col=8, max_row=min(ws.max_row, 80)):
        h = row[7] if len(row) > 7 else None
        if h and h.value is not None:
            non_empty += 1
            if str(h.value).strip().upper() == "#REF!":
                ref_errors += 1
    if non_empty > 5 and ref_errors / non_empty > 0.4:
        return 11
    return 0


def parse_sheet(ws, col_offset: int) -> dict[str, list[dict]]:
    """
    Parse one worksheet.

    Returns { norm_section: [{"name", "unit_price", "total"}, ...] }
    Only includes items that have a total in Excel.
    """
    mat_col_idx  = col_offset
    hdr_col_idx  = col_offset + 1
    act_col_idx  = col_offset + 8 if col_offset == 0 else col_offset + 6
    unit_col_idx = 6             if col_offset == 0 else col_offset + 4
    tot_col_idx  = 7             if col_offset == 0 else col_offset + 5

    sections: dict[str, list[dict]] = {}
    current_section: str | None = None
    in_bom_zone = False

    for row in ws.iter_rows():
        if row[0].row < 26:
            continue
        if len(row) <= hdr_col_idx:
            continue

        col_mat  = row[mat_col_idx]
        col_hdr  = row[hdr_col_idx]
        col_act  = row[act_col_idx]  if len(row) > act_col_idx  else None
        col_unit = row[unit_col_idx] if len(row) > unit_col_idx else None
        col_tot  = row[tot_col_idx]  if len(row) > tot_col_idx  else None

        mat_val = _cell_str(col_mat)
        hdr_val = _cell_str(col_hdr)

        # Section header: col B is bold, col A is empty
        if not mat_val and hdr_val and _is_bold(col_hdr):
            if hdr_val.upper() in _COL_SKIP_HEADERS:
                continue
            try:
                float(hdr_val.replace(",", "."))
                continue
            except ValueError:
                pass
            current_section = norm(hdr_val)
            in_bom_zone = True
            if current_section not in sections:
                sections[current_section] = []
            continue

        if not in_bom_zone or not mat_val or current_section is None:
            continue
        if mat_val.upper() in _COL_SKIP_HEADERS:
            continue

        chk_a = row[col_offset + 5] if len(row) > col_offset + 5 else None
        chk_b = row[col_offset + 6] if len(row) > col_offset + 6 else None
        if any(_cell_str(c).upper() in {"TOTAL", "GRAND TOTAL"}
               for c in (chk_a, chk_b) if c is not None):
            continue

        active_flag = None
        if col_act and col_act.value is not None:
            try:
                active_flag = int(float(str(col_act.value))) if str(col_act.value).strip() else None
            except (ValueError, TypeError):
                pass
        if active_flag == 0:
            continue

        if active_flag is None and col_tot:
            raw = str(col_tot.value) if col_tot.value is not None else ""
            if raw.strip().upper() in {"", "#REF!", "0", "0.0"}:
                try:
                    if col_tot.value is not None and float(str(col_tot.value)) == 0:
                        continue
                except (ValueError, TypeError):
                    pass

        sections[current_section].append({
            "name":       norm(mat_val),
            "unit_price": _safe_float(col_unit),
            "total":      _safe_float(col_tot),
        })

    return sections


def parse_workbook(path: str) -> dict[str, dict[str, list[dict]]]:
    """
    Parse an entire workbook.

    Returns { sheet_name: { norm_section: [{"name", "unit_price", "total"}, ...] } }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, dict[str, list[dict]]] = {}
    for ws in wb.worksheets:
        if norm(ws.title) in _SKIP_SHEETS:
            continue
        offset   = detect_col_offset(ws)
        sections = parse_sheet(ws, offset)
        if sections:
            result[ws.title] = sections
    return result


def parse_sheet_for_trailer(path: str, trailer_name: str) -> tuple[dict[str, list[dict]], str, float]:
    """
    Open a workbook, find the best-matching sheet for `trailer_name`,
    and return (sections, matched_sheet_name, score).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    candidate_sheets = [ws.title for ws in wb.worksheets if norm(ws.title) not in _SKIP_SHEETS]
    matched, score = best_sheet_match(candidate_sheets, trailer_name)
    if not matched:
        return {}, "", 0.0
    ws = wb[matched]
    offset = detect_col_offset(ws)
    sections = parse_sheet(ws, offset)
    return sections, matched, score
