"""
app/excel_grp_importer.py

Phase B of the GRP Costings importer rework. Walks one worksheet (e.g.
RIGID DRY FREIGHT), uses the cell resolver from Phase A to classify each
relevant cell, and produces an in-memory WritePlan describing every BOM
row that would be created — without writing anything to the DB.

Designed to be deterministic and read-only. The same WritePlan shape is
consumed by:
  • tools/preview_grp_import.py — CLI dump for the admin to eyeball
  • the future /api/import/grp/preview endpoint (Phase C)
  • the future commit endpoint (Phase D), which writes the plan to the DB

Pattern recognised
──────────────────
  Dimensions     C3 / C4 / C5 (length / width / height)
  Margin / ratio G4 / G8
  Body options   block under B = "BODY OPTIONS" header. Each row has:
                 col A = option name, col C = default Y/N,
                 col D = optional quantity
  Section        col B = section name (anything else). Header row +
                 line items + TOTAL row (G = "TOTAL").
  Master toggle  TOTAL row's I-cell = =IF(C<n>="Y",1,0) — gates the
                 whole section by the BODY OPTIONS row at row n
  Per-line gate  Item row's I-cell = =IF(C<n>="Y",1,0) — overrides the
                 section master with a row-specific gate
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from app.excel_cell_resolver import (
    Resolver, ResolvedKind, Resolved, discover_external_link_ids, parse_if_gate,
)


# ── Tuning constants ────────────────────────────────────────────────────────

DIMENSION_CELLS = {"length": "C3", "width": "C4", "height": "C5"}
MARGIN_CELL = "G4"
RATIO_CELL  = "G8"

# What a header row in column B looks like — used to detect section
# boundaries. Anything else with content in B is treated as a section name.
RESERVED_B_HEADERS = {"BODY OPTIONS"}

# Words that indicate a "TOTAL" row (case-insensitive in column G).
TOTAL_LABELS = {"TOTAL", "GRAND TOTAL"}

# Cells inside a section header row that we expect to hold column labels
# (skipped during item discovery).
HEADER_KEYWORDS = {
    "WIDTH", "HEIGHT", "M2", "PRICE", "TOTAL", "QUANTITY", "QUAN", "QUANT",
    "LENGTH", "QTY", "DESCRIPTION", "MATERIAL", "ITEM",
}


# ── Data classes ────────────────────────────────────────────────────────────

@dataclass
class BodyOption:
    """One row in the BODY OPTIONS block (rows 8..34 typically)."""
    name: str
    default_yn: bool
    quantity: float | None      # column D — used as variable_value at import
    source_row: int
    source_addr: str            # e.g. "A8"
    # Filled in post-discovery when a mutually-exclusive group is detected.
    # When set, two or more options share the same radio_group and the
    # importer writes them with selection_mode='single' so the calculator's
    # existing radio constraint kicks in.
    radio_group: str | None = None


@dataclass
class Section:
    name: str                   # column B value at header_row
    header_row: int
    total_row: int | None       # row where col G == "TOTAL"
    master_option: str | None   # body option name (resolved from total row I)
    j_multiplier: float | None  # for sections like SIDES that have J=H*N
    raw_total_formula: str | None


@dataclass
class BomLine:
    section: str
    item_name: str
    source_row: int
    source_addr: str            # e.g. "A56"

    # Quantity expression (translated from D/E/F columns into our formula
    # vocabulary: length/width/height + literal numbers + arithmetic).
    qty_formula: str
    qty_source_cell: str        # the cell that drove the qty (typically Fxx)

    # Price classification, inherited from the cell resolver
    price_kind: str             # ResolvedKind value
    price_value: float | None
    price_ref_sheet: str | None
    price_ref_cell: str | None
    price_fallback: float | None
    price_chain: tuple[str, ...]
    price_raw_formula: str | None

    # Inclusion gate (column I)
    gate_option_name: str | None        # body option name; None = always
    gate_source_addr: str | None        # cell where the gate lives (or None)
    inherited_from_section: bool = False  # True if gate came from section master

    notes: str = ""


@dataclass
class Warning:
    code: str
    message: str
    cell: str | None = None


@dataclass
class WritePlan:
    """Everything needed to populate one trailer template's BOM, without
    actually doing it. The web/commit endpoints serialise this to JSON
    or apply it inside a single DB transaction."""
    sheet_name: str
    trailer_name: str
    source_path: str
    dimensions: dict[str, float]
    default_margin: float | None
    default_ratio: float | None
    body_options: list[BodyOption] = field(default_factory=list)
    sections: list[Section] = field(default_factory=list)
    bom_lines: list[BomLine] = field(default_factory=list)
    grand_total_excel: float | None = None
    warnings: list[Warning] = field(default_factory=list)
    errors: list[Warning] = field(default_factory=list)


# ── Discovery entry point ───────────────────────────────────────────────────

def discover(xlsx_path: str | Path, sheet_name: str,
             *, trailer_name_override: str | None = None) -> WritePlan:
    xlsx_path = str(Path(xlsx_path).resolve())
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise FileNotFoundError(f"Sheet not found: {sheet_name!r}")
    ext_links = discover_external_link_ids(xlsx_path)
    ws, ws_data = wb[sheet_name], wb_data[sheet_name]

    plan = WritePlan(
        sheet_name=sheet_name,
        trailer_name=(trailer_name_override or sheet_name).strip(),
        source_path=xlsx_path,
        dimensions={},
        default_margin=None,
        default_ratio=None,
    )

    r = Resolver(ws, ws_data, ext_links)

    _discover_dimensions_and_globals(r, plan)
    body_options_by_row = _discover_body_options(ws, plan)
    _discover_sections_and_lines(ws, r, plan, body_options_by_row)
    _discover_grand_total(r, plan)
    _emit_consistency_warnings(plan, body_options_by_row)
    return plan


# ── Stage 1: dimensions / margin / ratio ───────────────────────────────────

def _discover_dimensions_and_globals(r: Resolver, plan: WritePlan) -> None:
    for name, addr in DIMENSION_CELLS.items():
        res = r.resolve(addr)
        if res.kind == ResolvedKind.LITERAL and res.value is not None:
            plan.dimensions[name] = float(res.value)
        else:
            plan.warnings.append(Warning(
                code="dim_missing",
                message=f"Dimension {name} ({addr}) is not a literal number "
                        f"(got {res.kind.value})",
                cell=addr,
            ))

    margin = r.resolve(MARGIN_CELL)
    if margin.kind == ResolvedKind.LITERAL and margin.value is not None:
        plan.default_margin = float(margin.value)
    ratio = r.resolve(RATIO_CELL)
    if ratio.kind == ResolvedKind.LITERAL and ratio.value is not None:
        plan.default_ratio = float(ratio.value)
    elif ratio.fallback_value is not None:
        plan.default_ratio = float(ratio.fallback_value)


# ── Stage 2: BODY OPTIONS block ─────────────────────────────────────────────

def _discover_body_options(ws, plan: WritePlan) -> dict[int, BodyOption]:
    """Find the BODY OPTIONS heading in column B, then collect every row
    underneath where column A has an option name. Stops at the next
    section header in column B.

    Returns {row_num: BodyOption} for use during gate-cell resolution.
    """
    head_row = None
    for r in range(1, min(ws.max_row, 80) + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().upper() == "BODY OPTIONS":
            head_row = r
            break
    if head_row is None:
        plan.warnings.append(Warning(
            code="no_body_options_block",
            message="No BODY OPTIONS heading found in column B (rows 1-80)",
        ))
        return {}

    out: dict[int, BodyOption] = {}
    for r in range(head_row + 1, ws.max_row + 1):
        b_val = ws.cell(r, 2).value
        if isinstance(b_val, str) and b_val.strip() and b_val.strip().upper() not in RESERVED_B_HEADERS:
            break  # next section header — body options block is over
        a_val = ws.cell(r, 1).value
        if not (isinstance(a_val, str) and a_val.strip()):
            continue
        name = a_val.strip()
        c_val = ws.cell(r, 3).value
        default_yn = isinstance(c_val, str) and c_val.strip().upper() == "Y"
        d_val = ws.cell(r, 4).value
        try:
            qty = float(d_val) if d_val not in (None, "") else None
        except (TypeError, ValueError):
            qty = None
        opt = BodyOption(
            name=name,
            default_yn=default_yn,
            quantity=qty,
            source_row=r,
            source_addr=f"A{r}",
        )
        out[r] = opt
        plan.body_options.append(opt)
    return out


# ── Stage 3: sections + per-line discovery ─────────────────────────────────

def _discover_sections_and_lines(ws, r: Resolver, plan: WritePlan,
                                 body_options_by_row: dict[int, BodyOption]) -> None:
    """Walk every row from after the BODY OPTIONS block to the bottom.
    A row is a section header when column B has content not in
    RESERVED_B_HEADERS. A section runs until the next header (or the
    end of the sheet)."""

    # Identify section header rows
    body_block_end = max(body_options_by_row.keys()) if body_options_by_row else 0
    section_rows: list[tuple[int, str]] = []
    for row in range(body_block_end + 1, ws.max_row + 1):
        b = ws.cell(row, 2).value
        if isinstance(b, str) and b.strip() and b.strip().upper() not in RESERVED_B_HEADERS:
            section_rows.append((row, b.strip()))

    if not section_rows:
        plan.warnings.append(Warning(
            code="no_sections",
            message="No section headers found in column B after the BODY OPTIONS block",
        ))
        return

    # Process each section
    for idx, (start_row, sec_name) in enumerate(section_rows):
        end_row = section_rows[idx + 1][0] - 1 if idx + 1 < len(section_rows) else ws.max_row
        section = _process_one_section(
            ws, r, plan, sec_name, start_row, end_row, body_options_by_row,
        )
        if section is not None:
            plan.sections.append(section)

    # Disambiguate duplicate section names by prefixing with the master
    # option. A common pattern in the GRP workbook is two "DOOR FITTINGS"
    # sections — one gated by DRD, one by SRD — that should appear as
    # "DRD DOOR FITTINGS" / "SRD DOOR FITTINGS" so the calculator and Body
    # Templates pages don't smush them into one. Matches the legacy
    # importer's behaviour on the previously-imported tt=50.
    _disambiguate_duplicate_section_names(plan)


def _disambiguate_duplicate_section_names(plan: WritePlan) -> None:
    name_counts: dict[str, int] = {}
    for s in plan.sections:
        name_counts[s.name] = name_counts.get(s.name, 0) + 1

    rename_map: dict[tuple[str, int], str] = {}
    # Track which master options had a section renamed, keyed by the
    # ORIGINAL (pre-rename) section name. Two or more masters renamed
    # for the same original name → mutually-exclusive alternatives
    # (the DRD vs SRD pattern).
    masters_by_orig_section: dict[str, set[str]] = {}

    for s in plan.sections:
        if name_counts.get(s.name, 0) <= 1:
            continue
        if not s.master_option:
            plan.warnings.append(Warning(
                code="duplicate_section_name",
                message=f"Section name {s.name!r} appears more than once "
                        f"and the second one has no master toggle to "
                        f"disambiguate by. Rows may collide.",
                cell=f"B{s.header_row}",
            ))
            continue
        master = s.master_option.strip()
        if master.upper() in s.name.upper():
            continue
        new_name = f"{master} {s.name}".strip()
        rename_map[(s.name, s.header_row)] = new_name
        masters_by_orig_section.setdefault(s.name, set()).add(master)

    if rename_map:
        for s in plan.sections:
            new_name = rename_map.get((s.name, s.header_row))
            if new_name and s.name != new_name:
                old_name = s.name
                s.name = new_name
                plan.warnings.append(Warning(
                    code="section_renamed_for_disambiguation",
                    message=f"Section {old_name!r} (row {s.header_row}, master "
                            f"{s.master_option!r}) renamed to {new_name!r} to "
                            f"avoid collision with another section of the same "
                            f"name elsewhere in the sheet.",
                    cell=f"B{s.header_row}",
                ))
                for line in plan.bom_lines:
                    if line.section == old_name and line.source_row > s.header_row \
                       and (s.total_row is None or line.source_row < s.total_row):
                        line.section = new_name

    # Auto-detect mutually-exclusive body option groups from the rename
    # signal. When two or more master options each had a section renamed
    # for the SAME original name (e.g. "DOOR FITTINGS"), they're competing
    # alternatives — only one should be selected at a time.
    _detect_mutex_body_options(plan, masters_by_orig_section)


def _detect_mutex_body_options(plan: WritePlan,
                               masters_by_orig_section: dict[str, set[str]]) -> None:
    """For each original section name with 2+ master options that gated
    a renamed copy, mark those masters as a radio group on the
    BodyOption objects so the importer can write selection_mode='single'.
    """
    name_to_option = {opt.name: opt for opt in plan.body_options}
    for orig_section, masters in masters_by_orig_section.items():
        if len(masters) < 2:
            continue
        group_label = orig_section.strip().upper()  # e.g. "DOOR FITTINGS"
        applied: list[str] = []
        for master in masters:
            opt = name_to_option.get(master)
            if opt is None:
                continue
            opt.radio_group = group_label
            applied.append(master)
        if applied:
            plan.warnings.append(Warning(
                code="mutex_body_options_detected",
                message=f"Body options {sorted(applied)!r} each gate a "
                        f"renamed copy of {orig_section!r} — grouped as "
                        f"mutually-exclusive radio under {group_label!r}. "
                        f"On the calculator, selecting one will deselect "
                        f"the others.",
            ))


def _process_one_section(ws, r: Resolver, plan: WritePlan, sec_name: str,
                         start_row: int, end_row: int,
                         body_options_by_row: dict[int, BodyOption]) -> Section | None:
    # Find TOTAL row inside the section
    total_row = None
    for row in range(start_row, end_row + 1):
        g = ws.cell(row, 7).value
        if isinstance(g, str) and g.strip().upper() in TOTAL_LABELS:
            total_row = row
            break

    # Resolve section master from TOTAL row's I-cell
    master_option = None
    j_mult = None
    raw_total_formula = None
    if total_row is not None:
        h_addr = f"H{total_row}"
        h_res = r.resolve(h_addr)
        raw_total_formula = h_res.raw_formula
        i_addr = f"I{total_row}"
        i_res = r.resolve(i_addr)
        if i_res.kind == ResolvedKind.IF_CONDITION and i_res.if_test_cell:
            test_row = _row_of(i_res.if_test_cell)
            opt = body_options_by_row.get(test_row)
            if opt is not None:
                master_option = opt.name
            else:
                plan.warnings.append(Warning(
                    code="master_test_unmapped",
                    message=f"Section {sec_name!r} master toggle tests "
                            f"{i_res.if_test_cell} but no body option lives at "
                            f"row {test_row}",
                    cell=i_addr,
                ))
        # J multiplier: e.g. SIDES has J138 = =H138*2
        j_addr = f"J{total_row}"
        j_res = r.resolve(j_addr)
        if j_res.raw_formula and "*" in j_res.raw_formula:
            # Look for a literal multiplier (e.g. =H138*2 or =H280*I280*D27*D27).
            # Only flag the simple constant case here; richer parsing happens
            # later if needed.
            simple = re.fullmatch(
                r"=\s*H\d+\s*\*\s*([0-9]+(?:\.[0-9]+)?)\s*",
                j_res.raw_formula,
            )
            if simple:
                try:
                    j_mult = float(simple.group(1))
                except ValueError:
                    pass

    section = Section(
        name=sec_name,
        header_row=start_row,
        total_row=total_row,
        master_option=master_option,
        j_multiplier=j_mult,
        raw_total_formula=raw_total_formula,
    )

    # Discover line items in this section
    item_end = (total_row - 1) if total_row else end_row
    for row in range(start_row + 1, item_end + 1):
        line = _process_one_line(ws, r, plan, section, row, body_options_by_row)
        if line is not None:
            plan.bom_lines.append(line)

    return section


def _process_one_line(ws, r: Resolver, plan: WritePlan, section: Section,
                     row: int, body_options_by_row: dict[int, BodyOption]) -> BomLine | None:
    a_val = ws.cell(row, 1).value
    if not (isinstance(a_val, str) and a_val.strip()):
        return None
    item_name = a_val.strip()
    if item_name.upper() in HEADER_KEYWORDS:
        return None

    # Quantity expression — favour F (m2 / qty) when present, else D.
    qty_addr, qty_formula = _build_qty_formula(ws, r, row)

    # Price (column G)
    g_addr = f"G{row}"
    p = r.resolve(g_addr)

    # Per-line gate (column I)
    gate_option = None
    gate_addr = f"I{row}"
    gate_res = r.resolve(gate_addr)
    inherited = False
    if gate_res.kind == ResolvedKind.IF_CONDITION and gate_res.if_test_cell:
        test_row = _row_of(gate_res.if_test_cell)
        opt = body_options_by_row.get(test_row)
        gate_option = opt.name if opt else None
        if opt is None:
            plan.warnings.append(Warning(
                code="line_gate_unmapped",
                message=f"Line {item_name!r} gate at {gate_addr} tests "
                        f"{gate_res.if_test_cell} but no body option lives there",
                cell=gate_addr,
            ))
    elif gate_res.kind == ResolvedKind.EMPTY:
        # No per-line gate — inherit the section master (if any)
        if section.master_option:
            gate_option = section.master_option
            inherited = True

    # Skip empty/header rows (no quantity AND no price kind)
    if p.kind == ResolvedKind.EMPTY and not qty_formula:
        return None

    return BomLine(
        section=section.name,
        item_name=item_name,
        source_row=row,
        source_addr=f"A{row}",
        qty_formula=qty_formula or "1",
        qty_source_cell=qty_addr,
        price_kind=p.kind.value,
        price_value=p.value,
        price_ref_sheet=p.ref_sheet,
        price_ref_cell=p.ref_cell,
        price_fallback=p.fallback_value,
        price_chain=p.chain,
        price_raw_formula=p.raw_formula,
        gate_option_name=gate_option,
        gate_source_addr=gate_addr if gate_option else None,
        inherited_from_section=inherited,
    )


# ── Quantity formula translation ───────────────────────────────────────────

# Map for the dimension cells. Anything that resolves through these gets
# replaced by the symbolic dimension token instead of its numeric value.
_DIM_TOKEN_FOR_CELL = {
    "C3": "length", "C4": "width", "C5": "height",
}

_AGG_RE_LOCAL = re.compile(
    r"\b(SUM|SUMPRODUCT|VLOOKUP|HLOOKUP|INDEX|MATCH|IF|IFS|OFFSET|"
    r"AVERAGE|MAX|MIN|COUNT|COUNTIF|SUMIF)\s*\(",
    re.IGNORECASE,
)
_CELL_REF_LOCAL = re.compile(r"\$?([A-Z]+)\$?(\d+)")

# External ref tokens — '[2]TAPING BLOCKS'!$F$47 etc. We strip just the
# '[N]SHEET'! prefix and leave the cell ref so the local substitution
# pass can still resolve it. This matches user expectation: the external
# sheet name is noise inside a qty formula; the cell ref inside happens
# to translate to the right expression locally (e.g. F47 in this sheet
# computes the door perimeter formula). The actual price for cells that
# reference external workbooks comes from the price-cell resolver, not
# from this qty translator — so dropping the external prefix here doesn't
# affect costing accuracy.
_EXT_REF_PREFIX_LOCAL = re.compile(r"'\[\d+\][^']+'!")
_EXT_REF_PREFIX_SHORT = re.compile(r"\[\d+\][A-Za-z0-9 +&_/]+!")

# Cleanup patterns for the inevitable +0 / 0+ / -0 / 0- terms that show
# up after substitution. (Doesn't touch 0.05 etc. — only bare 0s.)
_REDUNDANT_ZERO_RE = re.compile(r"(?<![.\d])0(?![.\d])")


def _build_qty_formula(ws, r: Resolver, row: int) -> tuple[str, str]:
    """Return (source_cell_addr, translated_formula). Picks F first, then E,
    then D — the column that actually drives the line's quantity. The
    translator substitutes dimension cells (C3/C4/C5) with length/width/height
    and recurses into other intra-sheet refs with the same logic.
    """
    for col_letter in ("F", "E", "D"):
        addr = f"{col_letter}{row}"
        v = ws.cell(row, column_index_from_string(col_letter)).value
        if v is None or v == "":
            continue
        if not isinstance(v, str):
            return addr, _fmt_num(v)
        if not v.startswith("="):
            return addr, str(v)
        translated = _translate_formula(ws, addr, depth=0, seen=frozenset())
        if translated:
            return addr, translated
    return "", ""


def _translate_formula(ws, addr: str, *, depth: int, seen: frozenset) -> str:
    if addr in seen or depth > 12:
        return ""
    col_letters, row_num = _split_addr(addr)
    cell = ws.cell(row_num, column_index_from_string(col_letters))
    raw = cell.value
    # Treat empty cells + text labels as 0 inside an arithmetic expression.
    # Excel evaluates blank-ref + n as n; a header label like "TOTAL"
    # appearing in a qty formula is meaningless so reduce to 0 too. The
    # cleanup pass below strips redundant '+0' / '-0' afterwards.
    if raw is None or raw == "":
        return "0"
    if not isinstance(raw, str):
        return _fmt_num(raw)
    if not raw.startswith("="):
        try:
            return _fmt_num(float(raw))
        except (TypeError, ValueError):
            return "0"  # text label — pretend it's zero so callers can clean up

    expr = raw[1:]
    if _AGG_RE_LOCAL.search(expr):
        # Aggregate — leave it for the caller to reach via fallback_value
        return ""

    # Strip external '[N]SHEET'! prefix only, keep the cell ref so the
    # CELL_REF substitution can still translate it against this sheet.
    # See the constants above for the rationale.
    expr_stripped = _EXT_REF_PREFIX_LOCAL.sub("", expr)
    expr_stripped = _EXT_REF_PREFIX_SHORT.sub("", expr_stripped)

    def sub_ref(m: re.Match) -> str:
        ref = f"{m.group(1)}{m.group(2)}"
        if ref in _DIM_TOKEN_FOR_CELL:
            return _DIM_TOKEN_FOR_CELL[ref]
        sub = _translate_formula(ws, ref, depth=depth + 1, seen=seen | {addr})
        if not sub:
            return ref  # leave the raw ref so the caller knows it didn't reduce
        return _maybe_wrap(sub)

    translated = _CELL_REF_LOCAL.sub(sub_ref, expr_stripped)
    return _cleanup_zero_terms(translated)


def _cleanup_zero_terms(expr: str) -> str:
    """Remove the +0 / 0+ / -0 / 0- byproducts of substituting empty cells
    or stripped external tokens. Repeats until stable so chains like
    '0+0+height' collapse cleanly."""
    if not expr:
        return expr
    prev = None
    cur = expr
    while prev != cur:
        prev = cur
        # Match a bare 0 (not 0.05, not 10) preceded/followed by + or -
        # at zero arithmetic depth. Repeated regex passes are simpler than
        # parsing — workbook formulas rarely nest deeply.
        cur = re.sub(r"(?<![.\d])\+\s*0(?![.\d])",  "", cur)   # +0  →
        cur = re.sub(r"(?<![.\d])0(?![.\d])\s*\+",  "", cur)   # 0+  →
        cur = re.sub(r"(?<![.\d])-\s*0(?![.\d])",   "", cur)   # -0  →
        # Don't touch '0-' or '0*' or '0/' — those would change semantics.
        # Strip leading/trailing whitespace + balanced wrapper parens.
        cur = cur.strip()
        if cur.startswith("(") and cur.endswith(")"):
            inner = cur[1:-1]
            depth = 0
            balanced = True
            for ch in inner:
                if ch == "(": depth += 1
                elif ch == ")": depth -= 1
                if depth < 0: balanced = False; break
            if balanced and depth == 0:
                cur = inner.strip()
    return cur or "0"


def _maybe_wrap(expr: str) -> str:
    """Wrap in parens if expr contains a top-level + or -."""
    if re.fullmatch(r"[\w.]+", expr):
        return expr
    depth = 0
    for ch in expr:
        if ch == "(": depth += 1
        elif ch == ")": depth -= 1
        elif ch in "+-" and depth == 0:
            return f"({expr})"
    return expr


def _fmt_num(v) -> str:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ""
    if f == int(f):
        return str(int(f))
    return f"{f:.8f}".rstrip("0").rstrip(".")


def _split_addr(addr: str) -> tuple[str, int]:
    m = re.fullmatch(r"\$?([A-Z]+)\$?(\d+)", addr.upper())
    return (m.group(1), int(m.group(2))) if m else ("", 0)


def _row_of(addr: str) -> int:
    _, r = _split_addr(addr)
    return r


# ── Grand total cross-check ────────────────────────────────────────────────

def _discover_grand_total(r: Resolver, plan: WritePlan) -> None:
    # Convention used in the workbook: J317 (or thereabouts) = =SUM(J50:J316).
    # Walk up from a sane upper bound looking for an aggregate in the J column.
    for row in range(320, 250, -1):
        res = r.resolve(f"J{row}")
        if res.kind == ResolvedKind.AGGREGATE and res.fallback_value is not None:
            plan.grand_total_excel = float(res.fallback_value)
            return


# ── Cross-cutting consistency checks ───────────────────────────────────────

def _emit_consistency_warnings(plan: WritePlan,
                               body_options_by_row: dict[int, BodyOption]) -> None:
    """Surface things admins should eyeball before commit."""

    # Body options that no line item or section references — orphaned toggles.
    referenced = set()
    for line in plan.bom_lines:
        if line.gate_option_name:
            referenced.add(line.gate_option_name)
    for sec in plan.sections:
        if sec.master_option:
            referenced.add(sec.master_option)
    for opt in plan.body_options:
        if opt.name not in referenced:
            plan.warnings.append(Warning(
                code="orphan_body_option",
                message=f"Body option {opt.name!r} (row {opt.source_row}) "
                        f"is never referenced — toggling it on the calculator "
                        f"will not change the BOM",
                cell=opt.source_addr,
            ))

    # Lines that ended up with no resolved price + no fallback — usable but
    # the admin needs to know.
    for line in plan.bom_lines:
        if line.price_kind == ResolvedKind.EMPTY.value:
            continue
        has_value = (
            line.price_value is not None
            or line.price_fallback is not None
            or line.price_ref_cell is not None
        )
        if not has_value:
            plan.errors.append(Warning(
                code="line_price_unresolved",
                message=f"Line {line.item_name!r} ({line.source_addr}) — "
                        f"price ({line.price_kind}) has no value, "
                        f"fallback, or external ref",
                cell=f"G{line.source_row}",
            ))
