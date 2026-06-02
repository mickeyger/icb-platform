"""
tools/audit_bom_vs_excel.py
───────────────────────────
BOM price audit: compares Excel section totals against the app's own
calculated totals (formula engine + default trailer dimensions).

App totals are computed the same way as when the user opens a costing in
the app, using the trailer's saved default dimensions.  Skin formula prices,
taping-block costs, floor-plate costs and mounting-cleat costs are all
resolved correctly.  Only sections that exist in the Excel BOM are compared.

Usage
─────
    python tools/audit_bom_vs_excel.py                 # open in browser
    python tools/audit_bom_vs_excel.py --out p.html    # save to file
    python tools/audit_bom_vs_excel.py --all           # show matching items too
    DATABASE_URL="mysql+pymysql://..." python tools/audit_bom_vs_excel.py
"""

import json
import os
import sys
import re
import argparse
import webbrowser
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
import sqlalchemy as sa
from app.formula_engine import calculate_bom   # uses no FastAPI deps

# ── Paths ─────────────────────────────────────────────────────────────────────

GRP_PATH = os.environ.get(
    "GRP_PATH",
    r"C:\Users\micge\Documents\Burt Costing Model\Latest price list\GRP Costings 2018.xlsx",
)
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./costing.db")

_SKIP_SHEETS = {
    "TRAILER UNITS SOLD", "CHASSIS COSTINGS", "SHEET1", "SHEET PLANNING",
    "VACUUM PLANNING HEIDELBERG", "SIDE TIPPER COSTINGS", "TRAILER PRICE LIST",
}

# Default dimensions used when the trailer has no explicit saved values
_FALLBACK_DIMS = {
    "floor_thickness":      0.060,
    "panel_thickness":      0.042,
    "insulation_thickness": 0.060,
    "num_doors": 2,
    "num_axles": 2,
}

# ── Normalisation ──────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).upper().strip())

def _norm_key(s: str) -> str:
    return re.sub(r"[^A-Z0-9 ]", "", _norm(s))

def _match_score(a: str, b: str) -> float:
    a_w = set(_norm_key(a).split())
    b_w = set(_norm_key(b).split())
    if not a_w and not b_w:
        return 0.0
    return len(a_w & b_w) / len(a_w | b_w)

def _safe_float(cell) -> float | None:
    v = cell.value if hasattr(cell, "value") else cell
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", ".").strip())
        return None if f == 0.0 else f
    except (ValueError, TypeError):
        return None


# ── Excel parsing ──────────────────────────────────────────────────────────────

def _is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False

def _cell_str(cell) -> str:
    v = cell.value
    return "" if v is None else str(v).strip()

_COL_SKIP_HEADERS = {"WIDTH", "HEIGHT", "M2", "PRICE", "TOTAL", "GRAND TOTAL",
                     "DESCRIPTION", "QTY", "COST", "QUANT", "MATERIAL", "ITEM"}


def _detect_col_offset(ws) -> int:
    """Detect CHILLER-style sheets (data starts col L not col A). Returns 0 or 11."""
    ref_errors = non_empty = 0
    for row in ws.iter_rows(min_row=26, max_col=8, max_row=min(ws.max_row, 80)):
        h = row[7] if len(row) > 7 else None
        if h and h.value is not None:
            non_empty += 1
            if str(h.value).strip().upper() == "#REF!":
                ref_errors += 1
    if non_empty > 5 and ref_errors / non_empty > 0.4:
        return 11
    return 0


def _parse_sheet(ws, col_offset: int) -> dict[str, list[dict]]:
    """
    Parse one worksheet. Returns:
        { norm_section: [{"name", "unit_price", "total"}, ...] }
    Only includes items that have both unit_price and total in Excel.
    Column layout: standard offset=0 (A/B/G/H/I), col-L offset=11 (L/M/P/Q/R)
    """
    mat_col_idx  = col_offset
    hdr_col_idx  = col_offset + 1
    act_col_idx  = col_offset + 8 if col_offset == 0 else col_offset + 6
    unit_col_idx = 6             if col_offset == 0 else col_offset + 4
    tot_col_idx  = 7             if col_offset == 0 else col_offset + 5

    sections: dict[str, list[dict]] = {}
    current_section: str | None = None
    in_bom_zone = False

    for row in ws.iter_rows():
        if row[0].row < 26:
            continue
        if len(row) <= hdr_col_idx:
            continue

        col_mat  = row[mat_col_idx]
        col_hdr  = row[hdr_col_idx]
        col_act  = row[act_col_idx]  if len(row) > act_col_idx  else None
        col_unit = row[unit_col_idx] if len(row) > unit_col_idx else None
        col_tot  = row[tot_col_idx]  if len(row) > tot_col_idx  else None

        mat_val = _cell_str(col_mat)
        hdr_val = _cell_str(col_hdr)

        if not mat_val and hdr_val and _is_bold(col_hdr):
            if hdr_val.upper() in _COL_SKIP_HEADERS:
                continue
            try:
                float(hdr_val.replace(",", "."))
                continue
            except ValueError:
                pass
            current_section = _norm(hdr_val)
            in_bom_zone = True
            if current_section not in sections:
                sections[current_section] = []
            continue

        if not in_bom_zone or not mat_val or current_section is None:
            continue
        if mat_val.upper() in _COL_SKIP_HEADERS:
            continue

        chk_a = row[col_offset + 5] if len(row) > col_offset + 5 else None
        chk_b = row[col_offset + 6] if len(row) > col_offset + 6 else None
        if any(_cell_str(c).upper() in {"TOTAL", "GRAND TOTAL"}
               for c in (chk_a, chk_b) if c is not None):
            continue

        active_flag = None
        if col_act and col_act.value is not None:
            try:
                active_flag = int(float(str(col_act.value))) if str(col_act.value).strip() else None
            except (ValueError, TypeError):
                pass
        if active_flag == 0:
            continue

        if active_flag is None and col_tot:
            raw = str(col_tot.value) if col_tot.value is not None else ""
            if raw.strip().upper() in {"", "#REF!", "0", "0.0"}:
                try:
                    if col_tot.value is not None and float(str(col_tot.value)) == 0:
                        continue
                except (ValueError, TypeError):
                    pass

        sections[current_section].append({
            "name":       _norm(mat_val),
            "unit_price": _safe_float(col_unit),
            "total":      _safe_float(col_tot),
        })

    return sections


def parse_excel_bom(path: str) -> dict[str, dict[str, list[dict]]]:
    """{ sheet_name: { norm_section: [{"name", "unit_price", "total"}, ...] } }"""
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, dict[str, list[dict]]] = {}
    for ws in wb.worksheets:
        if _norm(ws.title) in _SKIP_SHEETS:
            continue
        offset   = _detect_col_offset(ws)
        sections = _parse_sheet(ws, offset)
        if sections:
            result[ws.title] = sections
    return result


# ── Database: resolve special pricing ────────────────────────────────────────

def _load_special_prices(engine) -> tuple[dict, dict, dict, dict]:
    """
    Compute effective unit prices for skin formulas, taping blocks,
    floor plates, and mounting cleats using the same logic as the app.

    Returns four dicts:
        skin_prices   {formula_id:  price_per_m2}
        taping_prices {block_id:    cost_per_block}
        floor_prices  {plate_id:    cost}
        cleat_prices  {cleat_id:    cost}
    """
    with engine.connect() as conn:
        # ── Skin formula prices ──────────────────────────────────────────────
        rows = conn.execute(sa.text("""
            SELECT
                sfi.formula_id,
                SUM(COALESCE(
                    CASE WHEN sfi.price_source = 'sap' THEN sc.last_purch_price
                         ELSE ing.price_standard END, 0
                ) * sfi.qty_per_m2) AS formula_cost
            FROM skin_formula_items sfi
            JOIN skin_formula_ingredients ing ON ing.id = sfi.ingredient_id
            LEFT JOIN sap_item_codes sc ON sc.id = ing.sap_item_code_id
            GROUP BY sfi.formula_id
        """)).fetchall()
        skin_prices = {r[0]: round(float(r[1] or 0), 4) for r in rows}

        # ── Taping block prices ──────────────────────────────────────────────
        rows = conn.execute(sa.text("""
            SELECT
                tbi.block_id,
                SUM(COALESCE(
                    CASE WHEN tbi.price_source = 'sap' THEN sc.last_purch_price
                         ELSE tbi.price_per_unit END, 0
                ) * tbi.m2 * tbi.quantity) AS block_cost
            FROM taping_block_items tbi
            LEFT JOIN sap_item_codes sc ON sc.id = tbi.sap_item_code_id
            WHERE tbi.quantity > 0
            GROUP BY tbi.block_id
        """)).fetchall()
        taping_prices = {r[0]: round(float(r[1] or 0), 4) for r in rows}

        # ── Floor plate prices (with JSON post-formula) ──────────────────────
        rows = conn.execute(sa.text("""
            SELECT
                fpi.plate_id,
                SUM(COALESCE(
                    CASE WHEN fpi.price_source = 'sap' THEN sc.last_purch_price
                         ELSE fpi.price_per_unit END, 0
                ) * fpi.m2 * fpi.quantity) AS raw_cost,
                fp.price_formula
            FROM floor_plate_items fpi
            JOIN floor_plates fp ON fp.id = fpi.plate_id
            LEFT JOIN sap_item_codes sc ON sc.id = fpi.sap_item_code_id
            WHERE fpi.quantity > 0
            GROUP BY fpi.plate_id, fp.price_formula
        """)).fetchall()
        floor_prices = {}
        for plate_id, raw_cost, pf_json in rows:
            cost = float(raw_cost or 0)
            if pf_json:
                try:
                    for s in json.loads(pf_json):
                        op, val = s.get("op"), float(s.get("val", 1))
                        if op == "/" and val:
                            cost /= val
                        elif op == "*" and val:
                            cost *= val
                except Exception:
                    pass
            floor_prices[plate_id] = round(cost, 4)

        # ── Mounting cleat prices ────────────────────────────────────────────
        rows = conn.execute(sa.text("""
            SELECT
                mci.cleat_id,
                SUM(COALESCE(
                    CASE WHEN mci.price_source = 'sap' THEN sc.last_purch_price
                         ELSE mci.price_per_unit END, 0
                ) * mci.m2 * mci.quantity) AS cleat_cost
            FROM mounting_cleat_items mci
            LEFT JOIN sap_item_codes sc ON sc.id = mci.sap_item_code_id
            WHERE mci.quantity > 0
            GROUP BY mci.cleat_id
        """)).fetchall()
        cleat_prices = {r[0]: round(float(r[1] or 0), 4) for r in rows}

    return skin_prices, taping_prices, floor_prices, cleat_prices


# ── Database: load BOM + compute actual costs ─────────────────────────────────

# { trailer_name: { norm_section: { norm_name: app_line_cost } } }
AppBOM = dict[str, dict[str, dict[str, float]]]


def load_app_bom(database_url: str) -> tuple[AppBOM, list[str]]:
    """
    For each trailer type, compute the actual line-item costs as shown in the
    app's calculator, using:
      • the trailer's saved default dimensions
      • resolved pricing (skin formulas, taping blocks, floor plates, cleats)
      • the formula engine (same code path as /api/calculate)

    Returns:
        ( { trailer_name: { norm_section: { norm_name: line_cost } } },
          [ trailer_name, ... ] )
    """
    engine = sa.create_engine(
        database_url,
        connect_args={"check_same_thread": False} if database_url.startswith("sqlite") else {},
    )

    skin_p, taping_p, floor_p, cleat_p = _load_special_prices(engine)

    with engine.connect() as conn:
        trailers = conn.execute(sa.text("""
            SELECT id, name,
                   default_length, default_width, default_height,
                   default_num_axles, default_num_doors, default_insulation
            FROM trailer_types
            ORDER BY name
        """)).fetchall()

        # All non-body-option BOM rows with formula + pricing metadata
        bom_rows = conn.execute(sa.text("""
            SELECT
                b.trailer_type_id,
                COALESCE(b.bom_section, c.name, 'Uncategorised') AS section,
                m.name          AS material_name,
                m.price_per_unit,
                b.formula_expression,
                b.waste_percentage,
                COALESCE(bs.multiplier, 1.0)                     AS section_multiplier,
                b.skin_formula_id,
                b.taping_block_id,
                b.floor_plate_id,
                b.mounting_cleat_id,
                b.unit_price_override,
                b.sort_order
            FROM bill_of_materials b
            JOIN materials m ON m.id = b.material_id
            LEFT JOIN material_categories c ON c.id = m.category_id
            LEFT JOIN bom_sections bs ON bs.id = b.bom_section_id
            WHERE b.is_body_option = 0
            ORDER BY b.trailer_type_id, b.sort_order, b.id
        """)).fetchall()

    # Group raw BOM rows by trailer_type_id, resolving effective unit price
    tt_bom_items: dict[int, list[dict]] = {}
    for row in bom_rows:
        (tid, section, mat_name, price_pu, formula_expr, waste_pct,
         sect_mult, sf_id, tb_id, fp_id, mc_id, price_ov, _sort) = row

        if sf_id and sf_id in skin_p:
            eff_price = skin_p[sf_id]
        elif tb_id and tb_id in taping_p:
            eff_price = taping_p[tb_id]
        elif fp_id and fp_id in floor_p:
            eff_price = floor_p[fp_id]
        elif mc_id and mc_id in cleat_p:
            eff_price = cleat_p[mc_id]
        elif price_ov is not None:
            eff_price = float(price_ov)
        else:
            eff_price = float(price_pu or 0)

        tt_bom_items.setdefault(tid, []).append({
            "material_name":      _norm(mat_name),
            "category_name":      _norm(section),
            "formula_expression": formula_expr or "1",
            "waste_percentage":   float(waste_pct or 0),
            "price_per_unit":     eff_price,
            "unit_of_measure":    "each",
            "material_code":      "",
            "section_multiplier": float(sect_mult or 1.0),
        })

    # Run the formula engine for each trailer using its default dimensions
    result: AppBOM = {}
    for (tid, tt_name, def_len, def_wid, def_hgt,
         def_axles, def_doors, def_ins) in trailers:

        bom_items = tt_bom_items.get(tid)
        if not bom_items:
            continue

        dims = {
            "length": float(def_len or 7.5),
            "width":  float(def_wid or 2.6),
            "height": float(def_hgt or 2.6),
            "floor_thickness":      _FALLBACK_DIMS["floor_thickness"],
            "panel_thickness":      _FALLBACK_DIMS["panel_thickness"],
            "insulation_thickness": float(def_ins or _FALLBACK_DIMS["insulation_thickness"]),
            "num_axles": int(def_axles or _FALLBACK_DIMS["num_axles"]),
            "num_doors": int(def_doors or _FALLBACK_DIMS["num_doors"]),
        }

        calc = calculate_bom(bom_items, dims)

        by_sect: dict[str, dict[str, float]] = {}
        for it in calc["items"]:
            sect = it["category"]   # already normed by bom_item build
            name = it["material"]
            by_sect.setdefault(sect, {})
            if name not in by_sect[sect]:   # first BOM entry wins when duplicated
                by_sect[sect][name] = round(it["line_cost"], 2)

        result[tt_name] = by_sect

    return result, [t[1] for t in trailers]


# ── Matching ───────────────────────────────────────────────────────────────────

def match_trailers(excel_sheets: dict, app_names: list[str]) -> list[tuple[str, str, float]]:
    """Bijective greedy Jaccard-similarity matching."""
    candidates = [
        (_match_score(sh, ap), sh, ap)
        for sh in excel_sheets for ap in app_names
    ]
    candidates.sort(key=lambda x: -x[0])
    used_e: set[str] = set()
    used_a: set[str] = set()
    pairs = []
    for score, sh, ap in candidates:
        if sh in used_e or ap in used_a or score == 0:
            continue
        used_e.add(sh); used_a.add(ap)
        pairs.append((sh, ap, round(score, 3)))
    return sorted(pairs, key=lambda x: x[0])


# ── Comparison ─────────────────────────────────────────────────────────────────

def compare_sections(
    excel_sects:   dict[str, list[dict]],
    app_by_sect:   dict[str, dict[str, float]],
) -> dict[str, list[dict]]:
    """
    Compare Excel totals against actual app-computed totals, grouped by
    section.  Only Excel sections are compared; only items that have both a
    unit price AND a total in Excel are included; only items whose name also
    appears in the app BOM are included.

    Returns { section_name: [row_dict, ...] } sorted by |diff| desc.
    Each row_dict: { name, xls_unit, xls_total, app_total, diff, diff_pct }
    """
    flat: dict[str, list[tuple[str, float]]] = {}
    for sect, names in app_by_sect.items():
        for name, cost in names.items():
            flat.setdefault(name, []).append((sect, cost))

    out: dict[str, list[dict]] = {}
    for xls_sect, items in excel_sects.items():
        app_in_sect = app_by_sect.get(xls_sect, {})
        sect_rows: list[dict] = []

        for item in items:
            xls_unit  = item["unit_price"]
            xls_total = item["total"]
            if xls_total is None:   # no Excel total → skip
                continue

            norm_name = item["name"]
            if norm_name not in flat:
                continue   # name not in app BOM — skip

            # Prefer same-section app cost; fall back to first occurrence
            if norm_name in app_in_sect:
                app_total = app_in_sect[norm_name]
            else:
                app_total = flat[norm_name][0][1]

            diff     = app_total - xls_total
            diff_pct = (diff / xls_total * 100) if xls_total else None

            sect_rows.append({
                "name":      norm_name,
                "xls_unit":  xls_unit,
                "xls_total": xls_total,
                "app_total": app_total,
                "diff":      diff,
                "diff_pct":  diff_pct,
            })

        if sect_rows:
            sect_rows.sort(key=lambda r: -abs(r["diff"]))
            out[xls_sect] = sect_rows

    return out


# ── HTML report ────────────────────────────────────────────────────────────────

_CSS = """
* { box-sizing: border-box; }
body { font-family: 'Segoe UI', sans-serif; font-size: 13px;
       background:#0d1117; color:#c9d1d9; margin:0; padding:0; }
h1   { background:#161b22; padding:16px 24px; margin:0; font-size:18px;
       border-bottom:1px solid #30363d; color:#58a6ff; }
.summary { padding:12px 24px; background:#161b22;
           border-bottom:1px solid #30363d; font-size:12px; color:#8b949e; }
.legend  { margin:6px 0 0; font-size:11px; }
.body-block { margin:16px 24px; border:1px solid #30363d; border-radius:8px;
              overflow:hidden; }
.body-hdr { padding:10px 16px; background:#1c2128; font-weight:600;
            font-size:14px; display:flex; justify-content:space-between;
            align-items:center; }
.body-hdr .match { font-size:11px; color:#8b949e; font-weight:400; }
.ok   { color:#3fb950; }
.warn { color:#f0a500; }
.err  { color:#f85149; }
.all-ok   { padding:12px 16px; color:#3fb950; font-size:12px; }
.no-match { padding:12px 16px; color:#8b949e; font-size:12px; font-style:italic; }

.cmp-table { width:100%; border-collapse:collapse; }
.cmp-table th {
    background:#21262d; padding:6px 12px; font-size:10px;
    text-align:left; color:#8b949e; letter-spacing:.5px; white-space:nowrap;
}
.cmp-table th.num { text-align:right; }

/* Section header row */
.cmp-table tr.sect-row td {
    padding:8px 12px 6px; background:#1c2128;
    border-top:2px solid #30363d;
    font-weight:600; font-size:12px; vertical-align:middle;
}
.cmp-table tr.sect-row td.num { text-align:right; font-family:monospace; }
.cmp-table tr.sect-row.sect-ok   td { border-left:3px solid #2ea043; }
.cmp-table tr.sect-row.sect-warn td { border-left:3px solid #d29922; }
.cmp-table tr.sect-row.sect-err  td { border-left:3px solid #f85149; }

/* Item rows (indented) */
.cmp-table tr.item-row td {
    padding:5px 12px 5px 32px;
    border-top:1px solid #21262d; vertical-align:middle;
}
.cmp-table tr.item-row td.num { text-align:right; font-family:monospace; padding-left:12px; }
.cmp-table tr.item-row.big-diff td { background:#1c0505; }
.cmp-table tr.item-row.med-diff td { background:#1f1500; }
.cmp-table tr.item-row.no-diff  td { background:#0a1f0d; }
"""


def _fmt(v: float | None, decimals: int = 2) -> str:
    if v is None:
        return '<span style="color:#444">—</span>'
    return f"{v:,.{decimals}f}"


def _diff_span(diff: float | None, diff_pct: float | None, bold: bool = False) -> str:
    if diff is None:
        return '<span style="color:#444">—</span>'
    if abs(diff) < 0.05:
        return '<span class="ok">&#x2713;</span>'
    sign = "+" if diff > 0 else ""
    cls  = "err" if diff > 0 else "ok"
    pct  = f"&nbsp;<small>({sign}{diff_pct:.1f}%)</small>" if diff_pct is not None else ""
    wt   = "font-weight:600;" if bold else ""
    return f'<span class="{cls}" style="{wt}">{sign}{diff:,.2f}{pct}</span>'


def _item_cls(diff_pct: float | None) -> str:
    if diff_pct is None or abs(diff_pct) < 1:
        return "no-diff"
    return "big-diff" if abs(diff_pct) >= 10 else "med-diff"


def _sect_cls(rows: list[dict]) -> str:
    if all(abs(r["diff_pct"] or 0) < 1 for r in rows):
        return "sect-ok"
    if any(abs(r["diff_pct"] or 0) >= 10 for r in rows):
        return "sect-err"
    return "sect-warn"


def build_html(
    pairs:     list[tuple],
    excel_bom: dict,
    app_bom:   AppBOM,
    show_all:  bool,
) -> str:
    total_body    = len(pairs)
    bodies_w_diff = 0
    items_total   = 0
    items_diff    = 0
    blocks = []

    for sheet_name, app_name, score in pairs:
        exc       = excel_bom[sheet_name]
        app_sects = app_bom.get(app_name, {})
        by_sect   = compare_sections(exc, app_sects)

        body_diff_count = sum(
            1 for rows in by_sect.values()
            for r in rows if abs(r["diff"]) >= 0.05
        )
        body_item_count = sum(len(rows) for rows in by_sect.values())
        items_total += body_item_count
        items_diff  += body_diff_count
        has_diff = body_diff_count > 0
        if has_diff:
            bodies_w_diff += 1

        if not has_diff and not show_all:
            continue

        table_body = ""
        for sect_name, rows in by_sect.items():
            diff_in_sect  = [r for r in rows if abs(r["diff"]) >= 0.05]
            match_in_sect = [r for r in rows if abs(r["diff"]) < 0.05]

            if not diff_in_sect and not show_all:
                continue

            xls_sect_total = sum(r["xls_total"] for r in rows)
            app_sect_total = sum(r["app_total"] for r in rows)
            sect_diff      = app_sect_total - xls_sect_total
            sect_diff_pct  = (sect_diff / xls_sect_total * 100) if xls_sect_total else None
            srow_cls       = _sect_cls(rows)

            table_body += f"""
          <tr class="sect-row {srow_cls}">
            <td><b>{sect_name}</b></td>
            <td style="color:#8b949e;font-size:11px;text-align:right">{len(rows)} items</td>
            <td class="num">{_fmt(xls_sect_total)}</td>
            <td class="num">{_fmt(app_sect_total)}</td>
            <td class="num">{_diff_span(sect_diff, sect_diff_pct, bold=True)}</td>
          </tr>"""

            for r in (diff_in_sect + (match_in_sect if show_all else [])):
                cls = _item_cls(r["diff_pct"])
                xls_u = _fmt(r["xls_unit"]) if r["xls_unit"] else \
                    '<span style="color:#555">—</span>'
                table_body += f"""
          <tr class="item-row {cls}">
            <td>{r["name"]}</td>
            <td class="num" style="color:#8b949e;font-size:11px">{xls_u}</td>
            <td class="num">{_fmt(r["xls_total"])}</td>
            <td class="num">{_fmt(r["app_total"])}</td>
            <td class="num">{_diff_span(r["diff"], r["diff_pct"])}</td>
          </tr>"""

        if not table_body:
            if not by_sect:
                inner = '<div class="no-match">No comparable items found.</div>'
            else:
                inner = f'<div class="all-ok">&#x2713; All {body_item_count} items match.</div>'
        else:
            dims_note = ""
            tt = app_bom.get(app_name)
            status_bar = (
                f'<div style="padding:6px 16px;background:#21262d;font-size:11px;color:#8b949e">'
                f'{body_item_count} comparable items &nbsp;&middot;&nbsp; '
                f'<span class="{"err" if body_diff_count else "ok"}">'
                f'<b>{body_diff_count}</b> with price difference</span>'
                f'</div>'
            )
            inner = status_bar + f"""
          <table class="cmp-table">
            <thead><tr>
              <th style="width:280px">Material name</th>
              <th class="num" style="width:120px">Excel unit&nbsp;price</th>
              <th class="num" style="width:130px">Excel total</th>
              <th class="num" style="width:130px">App total</th>
              <th class="num" style="width:170px">Diff (App &minus; Excel)</th>
            </tr></thead>
            <tbody>{table_body}</tbody>
          </table>"""

        status_cls  = "err" if has_diff else "ok"
        status_icon = "&#x2717;" if has_diff else "&#x2713;"
        blocks.append(f"""
        <div class="body-block">
          <div class="body-hdr">
            <span><span class="{status_cls}">{status_icon}</span>&nbsp;{app_name}</span>
            <span class="match">Excel: &ldquo;{sheet_name}&rdquo;&nbsp;&middot;&nbsp;score:&nbsp;{score}</span>
          </div>
          {inner}
        </div>""")

    pair_rows_html = "".join(
        f'<tr><td style="padding:3px 10px;color:#8b949e">{sh}</td>'
        f'<td style="padding:3px 10px">{ap}</td>'
        f'<td style="padding:3px 10px;color:#8b949e;font-size:10px">{sc:.2f}</td></tr>'
        for sh, ap, sc in pairs
    )
    pair_table = f"""
    <details style="margin:8px 24px 0;font-size:11px">
      <summary style="cursor:pointer;color:#58a6ff;padding:4px 0">
        Show/hide Excel-to-App pairings ({total_body} matched)
      </summary>
      <table style="margin-top:6px;border-collapse:collapse">
        <tr>
          <th style="padding:3px 10px;text-align:left;color:#8b949e;font-size:10px">Excel sheet</th>
          <th style="padding:3px 10px;text-align:left;color:#8b949e;font-size:10px">App body type</th>
          <th style="padding:3px 10px;text-align:left;color:#8b949e;font-size:10px">Score</th>
        </tr>{pair_rows_html}
      </table>
    </details>"""

    summary = (
        f'<div class="summary">'
        f'<b>{total_body}</b> body types compared'
        f'&nbsp;&middot;&nbsp;'
        f'<span class="err"><b>{bodies_w_diff}</b> have differences</span>'
        f'&nbsp;&middot;&nbsp;'
        f'<span class="ok"><b>{total_body - bodies_w_diff}</b> match</span><br>'
        f'<b>{items_total}</b> comparable items &nbsp;&middot;&nbsp; '
        f'<span class="err"><b>{items_diff}</b> with price difference</span>'
        f'<div class="legend">'
        f'App&nbsp;total&nbsp;=&nbsp;formula&nbsp;engine&nbsp;&times;&nbsp;resolved&nbsp;price&nbsp;(skin/taping/floor/cleat)&nbsp;at&nbsp;default&nbsp;dimensions'
        f'&nbsp;&nbsp;|&nbsp;&nbsp;'
        f'<span style="background:#1c0505;padding:1px 6px;border-radius:3px">&#9632;</span>'
        f'&nbsp;&ge;10%&nbsp;&nbsp;'
        f'<span style="background:#1f1500;padding:1px 6px;border-radius:3px">&#9632;</span>'
        f'&nbsp;&lt;10%&nbsp;&nbsp;'
        f'<span style="background:#0a1f0d;padding:1px 6px;border-radius:3px">&#9632;</span>'
        f'&nbsp;match'
        f'</div>'
        f'</div>'
        + pair_table
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>BOM Price Audit</title>
<style>{_CSS}</style>
</head>
<body>
<h1>BOM Price Audit &mdash; Excel vs App</h1>
{summary}
{''.join(blocks)}
</body>
</html>"""


# ── Unmatched reporting ────────────────────────────────────────────────────────

def print_unmatched(pairs, excel_sheets, app_names):
    matched_e = {p[0] for p in pairs}
    matched_a = {p[1] for p in pairs}
    for s in sorted(set(excel_sheets) - matched_e):
        print(f"  [no app match]  xls: {s!r}")
    for s in sorted(set(app_names) - matched_a):
        print(f"  [no xls match]  app: {s!r}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="BOM price audit: Excel totals vs App")
    ap.add_argument("--out",  help="Write HTML to this path (default: open in browser)")
    ap.add_argument("--all",  action="store_true",
                    help="Also show items/sections where totals match")
    args = ap.parse_args()

    print(f"Reading Excel:  {GRP_PATH}")
    excel_bom = parse_excel_bom(GRP_PATH)
    print(f"  {len(excel_bom)} body-type sheets found")

    print(f"Reading DB:     {DATABASE_URL}")
    app_bom, app_names = load_app_bom(DATABASE_URL)
    print(f"  {len(app_names)} trailer types found")

    pairs = match_trailers(excel_bom, app_names)
    print(f"  {len(pairs)} pairs matched")
    print_unmatched(pairs, excel_bom, app_names)

    html = build_html(pairs, excel_bom, app_bom, args.all)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"\nReport written to: {args.out}")
    else:
        tmp = tempfile.NamedTemporaryFile(
            suffix=".html", delete=False, mode="w", encoding="utf-8"
        )
        tmp.write(html)
        tmp.close()
        print(f"\nOpening report in browser: {tmp.name}")
        webbrowser.open(f"file:///{tmp.name}")


if __name__ == "__main__":
    main()
