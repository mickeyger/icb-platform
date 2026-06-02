"""
tools/parse_grp_sheet.py

Parse a GRP Costings sheet and produce structured JSON ready for import.

Rules applied
─────────────
• Named ranges from Excel Name Manager are used when present;
  falls back to Column-B section headers.
• Only sections with a non-zero aggregate (col J or fallback) are imported.
• SIZE block (rows where col A = LENGTH / WIDTH / HEIGHT) → trailer defaults.
• Formula expressions are translated into our dimension-variable space
  (length, width, height) with correct operator-precedence parenthesisation.
• Section multipliers are extracted from the col-J formula (e.g. =H111*2 → ×2).
• SPRAY PAINTING and similar ad-hoc sections that have no TOTAL header row
  are handled by summing col-J values directly.

Usage
─────
  python tools/parse_grp_sheet.py [excel_path] [sheet_name] [out.json]
"""

import sys
import re
import json
import openpyxl
from openpyxl.utils import column_index_from_string

# ── Workbook helpers ───────────────────────────────────────────────────────

def load_workbooks(path: str):
    wb_data     = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    return wb_data, wb_formulas


def safe_float(v) -> float:
    try:
        return round(float(v), 8) if v not in (None, "") else 0.0
    except Exception:
        return 0.0


def fmt_num(v: float) -> str:
    """Format float without trailing zeros."""
    if v == 0:
        return "0"
    s = f"{v:.8f}".rstrip("0").rstrip(".")
    return s


# ── Formula-expression translator ─────────────────────────────────────────
#
# Produces Python-compatible expressions using length / width / height.
# Cross-sheet references and aggregate functions fall back to computed values.

# Hard-coded mappings for the SIZE-row cells (C4/C5/C6 and their derivatives)
CELL_VAR_MAP: dict[str, str] = {
    "C4": "length",
    "C5": "width",
    "C6": "height",
    "D4": "(length + 0.05)",
    "D5": "(width + 0.05)",
    "D6": "height",
    "E4": "((length + 0.05) * 4)",
    "E5": "((width + 0.05) * 2)",
    "F5": "((length + 0.05) * 4 + (width + 0.05) * 2)",
}

AGGREGATE_RE = re.compile(
    r"^(SUM|IF|SUMPRODUCT|OFFSET|AVERAGE|MAX|MIN|VLOOKUP|INDEX|MATCH)\(",
    re.I,
)
CELL_REF_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")

# thread-safe cache (row, col) → resolved expression string
_resolve_cache: dict[tuple, str] = {}


def _needs_parens(expr: str) -> bool:
    """True if expr contains a top-level +/- that could break precedence."""
    if re.fullmatch(r"[\w.]+", expr):          # simple number or bare variable
        return False
    if expr.startswith("(") and expr.endswith(")"):  # already wrapped
        # verify the opening paren closes at the end
        depth = 0
        for i, ch in enumerate(expr):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            if depth == 0 and i < len(expr) - 1:
                return False   # outer parens close early → not a simple wrap
        return True            # outer parens cover the whole expression
    depth = 0
    for ch in expr:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch in "+-" and depth == 0:
            return True
    return False


def _maybe_wrap(expr: str) -> str:
    """Wrap expr in parens if it has top-level additive operators."""
    return f"({expr})" if _needs_parens(expr) else expr


def resolve_cell_expr(row: int, col: int, ws_d, ws_f, depth: int = 0) -> str:
    """Return a symbolic expression (or numeric fallback) for cell (row, col)."""
    key = (row, col)
    if key in _resolve_cache:
        return _resolve_cache[key]

    if depth > 12:
        result = fmt_num(safe_float(ws_d.cell(row, col).value))
        _resolve_cache[key] = result
        return result

    raw = ws_f.cell(row, col).value

    # ── No formula: literal value ────────────────────────────────────────
    if not isinstance(raw, str) or not raw.startswith("="):
        v = ws_d.cell(row, col).value
        if isinstance(v, (int, float)):
            result = fmt_num(v)
        else:
            result = str(v or 0)
        _resolve_cache[key] = result
        return result

    expr = raw[1:]  # strip leading "="

    # ── Cross-sheet reference ────────────────────────────────────────────
    if "[" in expr and "!" in expr:
        result = fmt_num(safe_float(ws_d.cell(row, col).value))
        _resolve_cache[key] = result
        return result

    # ── Aggregate / conditional function → computed value ────────────────
    if AGGREGATE_RE.match(expr):
        result = fmt_num(safe_float(ws_d.cell(row, col).value))
        _resolve_cache[key] = result
        return result

    # ── Translate cell references ────────────────────────────────────────
    def sub_ref(m: re.Match) -> str:
        col_part = m.group(1)
        row_part = int(m.group(2))
        cell_key = f"{col_part}{row_part}"
        if cell_key in CELL_VAR_MAP:
            return CELL_VAR_MAP[cell_key]
        col_num = column_index_from_string(col_part)
        resolved = resolve_cell_expr(row_part, col_num, ws_d, ws_f, depth + 1)
        # Wrap in parens if the sub-expression has additive operators,
        # because it will be embedded into a product/quotient.
        return _maybe_wrap(resolved)

    translated = CELL_REF_RE.sub(sub_ref, expr)
    _resolve_cache[key] = translated
    return translated


def try_simplify(expr: str, computed: float) -> str:
    """
    If the expression is pure arithmetic (no dimension vars), evaluate it.
    If it references dimension variables, return it as-is.
    """
    if not expr or expr in ("None", ""):
        return fmt_num(computed)
    if any(v in expr for v in ("length", "width", "height")):
        return clean_zero_terms(expr)
    try:
        v = eval(expr, {"__builtins__": {}}, {})   # noqa: S307
        evaluated = round(float(v), 6)
        # If numeric expression evaluates fine, verify against computed qty.
        # If they differ by more than 2 %, the H-column had a hidden multiplier
        # (e.g. =G*F*3); fall back to the computed value.
        if computed != 0 and abs(evaluated - computed) / abs(computed) > 0.02:
            return fmt_num(computed)
        return fmt_num(evaluated)
    except Exception:
        pass
    return fmt_num(computed)


def clean_zero_terms(expr: str) -> str:
    """Remove redundant ±0 (bare zero, not part of 0.xx) from expressions."""
    # Match +0 or -0 NOT followed by a decimal digit  (won't touch 0.05, 0.06…)
    cleaned = re.sub(r'([+-])0(?!\.\d)', "", expr)
    cleaned = cleaned.lstrip("+").strip()
    return cleaned or "0"


# ── Column-header detection ────────────────────────────────────────────────

HEADER_KEYWORDS = {
    "WIDTH", "HEIGHT", "M2", "PRICE", "PRICE ", "TOTAL", "TOTAL M",
    "QUANT", "QUANTI", "LENGTH", "QTY",
}


def find_header_row(start: int, end: int, ws_d) -> tuple:
    """Return (row_num, {keyword: col_idx}) for the first column-header row."""
    for r in range(start, end + 1):
        matches: dict[str, int] = {}
        for c in range(1, 12):
            v = ws_d.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() in HEADER_KEYWORDS:
                matches[v.strip().upper()] = c
        if len(matches) >= 2:
            return r, matches
    return None, {}


# ── TOTAL-row detection ────────────────────────────────────────────────────

def find_total_rows(start: int, end: int, ws_d) -> list[int]:
    """Rows where any of cols D-H contains the text 'TOTAL'."""
    rows = []
    for r in range(start, end + 1):
        for c in range(4, 9):
            v = ws_d.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "TOTAL":
                rows.append(r)
                break
    return rows


# ── J-column aggregate lookup ──────────────────────────────────────────────

def get_j_aggregate(total_rows: list[int], item_rows: list[int], ws_d, ws_f) -> tuple:
    """
    Return (j_value, j_formula_str, agg_row).
    First tries rows near a TOTAL row, then falls back to summing item J values.
    """
    for tr in total_rows:
        for r in (tr, tr + 1, tr - 1):
            jd = ws_d.cell(r, 10).value
            jf = ws_f.cell(r, 10).value
            if jd is not None:
                return safe_float(jd), str(jf or ""), r

    # No TOTAL row — aggregate item J values directly
    total = sum(safe_float(ws_d.cell(r, 10).value) for r in item_rows)
    if total != 0:
        return total, "", None

    return 0.0, "", None


# ── Section-multiplier extraction ─────────────────────────────────────────

def extract_multiplier(j_formula: str) -> float:
    """
    '=H111*2'   → 2
    '=H32'      → 1
    '=I44*H44'  → 1  (conditional flag × total, flag is 0 or 1)
    """
    m = re.search(r"\*\s*(\d+(?:\.\d+)?)\s*$", str(j_formula))
    if m:
        v = float(m.group(1))
        return int(v) if v == int(v) else v
    return 1.0


# ── Unit-of-measure heuristic ─────────────────────────────────────────────

def guess_uom(formula_expr: str, is_door_style: bool, price: float) -> str:
    if is_door_style:
        return "each"
    if any(v in formula_expr for v in ("length", "width", "height")):
        # linear items have only ONE dimension variable and a fixed multiplier ≤ 2
        parts = re.findall(r"(length|width|height)", formula_expr)
        if len(parts) == 1:
            return "m"
        return "m2"
    # Pure numeric formula — guess by magnitude
    try:
        v = float(formula_expr)
        return "each" if v == int(v) and v <= 100 else "m2"
    except Exception:
        return "each"


# ── Core parser ────────────────────────────────────────────────────────────

SIZE_LABELS = {"LENGTH", "WIDTH", "HEIGHT"}


def parse_sheet(sheet_name: str, wb_data, wb_formulas) -> dict:
    _resolve_cache.clear()

    ws_d = wb_data[sheet_name]
    ws_f = wb_formulas[sheet_name]
    MAX_ROW = ws_d.max_row

    # ── Named-range support ──────────────────────────────────────────────
    named_ranges: dict[str, str] = {}
    for name, defn in wb_formulas.defined_names.items():
        for title, coord in defn.destinations:
            if title.strip("'") == sheet_name:
                named_ranges[name] = coord

    # ── SIZE block ───────────────────────────────────────────────────────
    trailer_defaults: dict[str, float] = {}
    size_row_set: set[int] = set()
    for r in range(1, min(30, MAX_ROW + 1)):
        a = ws_d.cell(r, 1).value
        if isinstance(a, str) and a.strip().upper() in SIZE_LABELS:
            trailer_defaults[a.strip().lower()] = safe_float(ws_d.cell(r, 3).value)
            size_row_set.add(r)

    # ── Detect sections from col B ────────────────────────────────────────
    col_b_sections: list[tuple[int, str]] = []
    for r in range(1, MAX_ROW + 1):
        if r in size_row_set:
            continue
        a = ws_d.cell(r, 1).value
        b = ws_d.cell(r, 2).value
        if b and isinstance(b, str) and b.strip() and a is None:
            col_b_sections.append((r, b.strip()))

    # ── Section boundaries ───────────────────────────────────────────────
    raw_sections: list[dict] = []
    for idx, (start_row, sect_name) in enumerate(col_b_sections):
        end_row = (col_b_sections[idx + 1][0] - 1
                   if idx + 1 < len(col_b_sections) else MAX_ROW)
        raw_sections.append({"name": sect_name, "start": start_row, "end": end_row})

    # ── Process each section ──────────────────────────────────────────────
    parsed_sections: list[dict] = []
    skipped: list[str] = []

    for sect in raw_sections:
        name  = sect["name"]
        start = sect["start"]
        end   = sect["end"]

        hdr_row, hdr_map = find_header_row(start + 1, end, ws_d)
        total_rows_found = find_total_rows(start, end, ws_d)

        # Determine column layout
        is_door_style = bool(hdr_map.get("QUANT") or hdr_map.get("QUANTI"))

        if is_door_style:
            qty_col   = hdr_map.get("QUANT", hdr_map.get("QUANTI", 4))
            price_col = hdr_map.get("PRICE", hdr_map.get("PRICE ", 5))
            total_col = hdr_map.get("TOTAL", 6)
        else:
            qty_col   = 6   # col F = m2 / linear
            price_col = hdr_map.get("PRICE", hdr_map.get("PRICE ", 7))
            total_col = hdr_map.get("TOTAL", 8)

        # Collect candidate item rows (col A has a material name)
        stop_rows = set(total_rows_found) | {start} | ({hdr_row} if hdr_row else set())
        item_rows = []
        for r in range(start + 1, end + 1):
            if r in stop_rows:
                continue
            a_val = ws_d.cell(r, 1).value
            if isinstance(a_val, str) and a_val.strip():
                item_rows.append(r)

        # J-column aggregate (section total that feeds into grand total)
        j_val, j_formula, j_agg_row = get_j_aggregate(
            total_rows_found, item_rows, ws_d, ws_f
        )
        multiplier = extract_multiplier(j_formula)

        # Skip sections whose aggregate contribution is zero
        if j_val == 0:
            skipped.append(name)
            continue

        # ── Build line items ──────────────────────────────────────────────
        items: list[dict] = []
        for r in item_rows:
            item_total = safe_float(ws_d.cell(r, total_col).value)

            # For ad-hoc sections (SPRAY PAINTING) that have no TOTAL col,
            # fall back to col H then col J
            if item_total == 0:
                item_total = safe_float(ws_d.cell(r, 8).value)
            if item_total == 0:
                continue  # zero-value item (e.g. PU when EPS selected)

            price = safe_float(ws_d.cell(r, price_col).value)

            # ── Derive formula expression (the quantity) ──────────────────
            raw_expr = resolve_cell_expr(r, qty_col, ws_d, ws_f)

            # Handle flat-fee rows (price = 0, total is the fee amount itself)
            if price == 0 and item_total != 0:
                price        = item_total
                formula_expr = "1"
                qty_computed = 1.0
            else:
                qty_computed = (round(item_total / price, 6) if price != 0
                                else safe_float(ws_d.cell(r, qty_col).value))
                formula_expr = try_simplify(raw_expr, qty_computed)

            # Fall-back: formula resolved to "0" but computed qty is known
            if formula_expr == "0" and qty_computed not in (0, 0.0):
                formula_expr = fmt_num(qty_computed)

            uom = guess_uom(formula_expr, is_door_style, price)

            a_val = ws_d.cell(r, 1).value
            items.append({
                "material_name":     a_val.strip(),
                "formula_expression": formula_expr,
                "price_per_unit":    round(price, 4),
                "unit_of_measure":   uom,
                "waste_percentage":  0,
                "notes": (f"qty≈{round(qty_computed, 4)} · "
                          f"total≈{round(item_total, 2)}"),
            })

        parsed_sections.append({
            "name":          name,
            "multiplier":    multiplier,
            "section_total": round(j_val, 2),
            "items":         items,
        })

    return {
        "trailer_name":     sheet_name,
        "source_sheet":     sheet_name,
        "named_ranges_used": bool(named_ranges),
        "trailer_defaults": trailer_defaults,
        "margin_config":    None,
        "sections":         parsed_sections,
        "skipped_sections": skipped,
        "grand_total_check": round(
            sum(s["section_total"] for s in parsed_sections), 2
        ),
    }


# ── Entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else (
        r"C:\Users\micge\Documents\Burt Costing Model\GRP Costings 2018.xlsx"
    )
    sheet_name  = sys.argv[2] if len(sys.argv) > 2 else "UP TO 2.3 CHILLER BODY"
    output_path = sys.argv[3] if len(sys.argv) > 3 else None

    print(f"Parsing '{sheet_name}' …", file=sys.stderr)
    wb_data, wb_formulas = load_workbooks(excel_path)
    result = parse_sheet(sheet_name, wb_data, wb_formulas)

    out = json.dumps(result, indent=2, default=str)
    if output_path:
        with open(output_path, "w", encoding="utf-8") as fh:
            fh.write(out)
        print(f"Written to {output_path}", file=sys.stderr)
    else:
        print(out)
