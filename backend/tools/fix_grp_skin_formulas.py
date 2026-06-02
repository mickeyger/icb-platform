"""
Cross-check 'EXT GRP SKIN 2*300' BOM formulas against GRP Costings 2018.xlsx.

Scope: only rows where the Excel quantity is computed as
  D = (width-style expression involving body-variable cells)
  E = (height-style expression involving body-variable cells)
  F = D * E
i.e. M² calculations using width AND height. Length-based and other
formulas are ignored.

Builds the canonical formula as it appears in Excel (with cell refs
resolved to `width`, `height`, and `{NAME}` Body Variable tokens), and
compares against the DB's stored `formula_expression` for the matching
BOM row. Pairs DB rows to Excel rows by (trailer type, material name,
BOM section). Section is taken from the Excel section header (col B,
the most recent non-empty value above the EXT GRP SKIN row).

Default: dry run. Use --apply to write. Skips rows already matching.

Usage:
    python tools/fix_grp_skin_formulas.py
    python tools/fix_grp_skin_formulas.py --apply
    python tools/fix_grp_skin_formulas.py --db-url mysql+pymysql://USER:PASS@HOST/DB --apply
"""
from __future__ import annotations
import argparse, os, re, sys
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass

DEFAULT_XLSX = r"C:\Users\micge\Documents\Burt Costing Model\GRP Costings 2018.xlsx"

SHEET_ALIASES = {
    "4.9 & UP FREEZER BODY 3": "4.9 & UP FREEZER BODY 2",
}
SKIP_SHEETS = {
    'TRAILER UNITS SOLD', 'CHASSIS COSTINGS', 'SHEET PLANNING',
    'VACUUM PLANNING HEIDELBERG', 'SIDE TIPPER COSTINGS', 'TRAILER PRICE LIST', 'EXAMPLE',
}
TARGET_MATERIAL = "EXT GRP SKIN 2*300"


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).upper()


def body_var_map(sh) -> dict[int, str]:
    """row_num -> body variable name (e.g. {18: 'FLOOR EPS'})"""
    m = {}
    header = None
    for r in range(1, 25):
        b = sh.cell(r, 2).value
        if b and 'BODY OPTIONS' in str(b).upper():
            header = r
            break
    if header is None:
        return m
    for r in range(header + 1, header + 25):
        name = sh.cell(r, 1).value
        if not name:
            continue
        if 'BODY OPTIONS' in str(name).upper():
            continue
        if isinstance(sh.cell(r, 3).value, (int, float)):
            m[r] = str(name).strip()
    return m


def resolve_formula(expr: str, sh, var_rows: dict[int, str]) -> str | None:
    """Replace cell refs with width/height/length or {VAR} tokens. Returns None if any ref can't be resolved."""
    if not isinstance(expr, str):
        return None
    raw = expr.lstrip('=')
    unresolved: list[str] = []

    def repl(m):
        col = m.group(1)
        row = int(m.group(2))
        if col == 'C':
            label = sh.cell(row, 1).value
            if label:
                up = str(label).upper()
                if 'WIDTH' in up:  return 'width'
                if 'HEIGHT' in up: return 'height'
                if 'LENGTH' in up: return 'length'
            if row in var_rows:
                return '{' + var_rows[row] + '}'
        unresolved.append(m.group(0))
        return m.group(0)

    out = re.sub(r'([A-Z])(\d+)', repl, raw)
    if unresolved:
        return None
    return out


def section_for_row(sh, row: int) -> str:
    """Walk up looking for the most recent non-empty col B header (FRONT, SIDES, etc.)."""
    for r in range(row - 1, max(0, row - 40), -1):
        b = sh.cell(r, 2).value
        if b and isinstance(b, str) and b.strip() and b.strip().upper() != 'BODY OPTIONS':
            return b.strip().upper()
    return ''


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--db-url", default=None)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"Workbook not found: {args.xlsx}")
    db_url = args.db_url or os.environ.get("DATABASE_URL") or "sqlite:///costing.db"
    print(f"DB:     {db_url}")
    print(f"Source: {args.xlsx}")
    print(f"Mode:   {'APPLY' if args.apply else 'DRY RUN'}")
    print()

    engine = create_engine(db_url)
    with engine.connect() as c:
        types = {normalize(n): tid for tid, n in c.execute(
            text("SELECT id, name FROM trailer_types WHERE name NOT LIKE '%[deleted-%'")).fetchall()}

    wb = openpyxl.load_workbook(args.xlsx, data_only=False)

    proposals: list[dict] = []
    no_match: list[str] = []
    skipped_sheets: list[str] = []

    for sn in wb.sheetnames:
        if sn.strip() in SKIP_SHEETS:
            continue
        alias = SHEET_ALIASES.get(sn.strip())
        norm = normalize(alias or sn)
        type_id = types.get(norm)
        if not type_id:
            skipped_sheets.append(sn)
            continue
        sh = wb[sn]
        var_rows = body_var_map(sh)

        for r in range(1, sh.max_row + 1):
            if sh.cell(r, 1).value != TARGET_MATERIAL:
                continue
            d_f = sh.cell(r, 4).value
            e_f = sh.cell(r, 5).value
            d_res = resolve_formula(d_f, sh, var_rows) if d_f else None
            e_res = resolve_formula(e_f, sh, var_rows) if e_f else None
            if not (d_res and e_res and 'width' in d_res and 'height' in e_res):
                continue
            new_formula = f"({d_res})*({e_res})"
            section = section_for_row(sh, r)

            with engine.connect() as c:
                row = c.execute(text("""
                    SELECT b.id, b.formula_expression, b.bom_section
                    FROM bill_of_materials b
                    JOIN materials m ON m.id = b.material_id
                    WHERE b.trailer_type_id = :tid AND m.name = :mat AND UPPER(b.bom_section) = :sec
                """), {"tid": type_id, "mat": TARGET_MATERIAL, "sec": section}).fetchone()
            if not row:
                no_match.append(f"{sn} row {r} section={section}")
                continue
            old_formula = row[1] or ''
            if old_formula == new_formula:
                status = "OK (matches)"
            else:
                status = "DIFF"
            proposals.append(dict(
                sheet=sn, xrow=r, section=section, bom_id=row[0],
                old=old_formula, new=new_formula, status=status,
            ))

    diffs = [p for p in proposals if p["status"] == "DIFF"]
    matches = [p for p in proposals if p["status"] == "OK (matches)"]

    print("=" * 100)
    print(f"PROPOSALS: {len(proposals)} | DIFFs: {len(diffs)} | matches: {len(matches)} | no DB match: {len(no_match)} | skipped sheets: {len(skipped_sheets)}")
    print("=" * 100)

    by_sheet: dict[str, list[dict]] = {}
    for p in diffs:
        by_sheet.setdefault(p["sheet"], []).append(p)
    for sn, items in sorted(by_sheet.items()):
        print(f"\n--- {sn} ---")
        for p in items:
            print(f"  row {p['xrow']} (section={p['section']}, bom_id={p['bom_id']}):")
            print(f"    OLD: {p['old']}")
            print(f"    NEW: {p['new']}")

    if matches:
        print(f"\n--- Already matching ({len(matches)}) ---")
        for p in matches:
            print(f"  {p['sheet']} row {p['xrow']} section={p['section']} bom_id={p['bom_id']}")

    if no_match:
        print(f"\n--- No DB row match ({len(no_match)}) ---")
        for u in no_match:
            print(f"  {u}")

    if skipped_sheets:
        print(f"\n--- Skipped (no matching trailer_type) ---")
        for s in skipped_sheets:
            print(f"  {s}")

    print(f"\nWould write {len(diffs)} updates.")
    if not args.apply:
        print("(dry run — no changes made; pass --apply to write)")
        return

    with engine.begin() as c:
        for p in diffs:
            c.execute(text("UPDATE bill_of_materials SET formula_expression = :f WHERE id = :id"),
                      {"f": p["new"], "id": p["bom_id"]})
    print(f"Wrote {len(diffs)} updates.")


if __name__ == "__main__":
    main()
