"""
Import skin/taping/cleat/floor formula links from GRP Costings 2018.xlsx.

Scans every body-template-named sheet, finds PRICE-column formulas that
reference FORMULAS 2018.xls, and links the matching BOM row to the
corresponding skin_formula / taping_block / mounting_cleat / floor_plate.

Default: dry run — prints proposed changes only. Use --apply to write.
Skips rows that already have a non-null link (preserves manual edits)
unless --overwrite is given.

Usage:
    python tools/import_formula_links.py                       # local SQLite, dry run
    python tools/import_formula_links.py --apply               # local SQLite, write
    python tools/import_formula_links.py --db-url mysql+pymysql://USER:PASS@HOST/DB --apply
"""
from __future__ import annotations
import argparse, os, re, sys, zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from sqlalchemy import create_engine, text

# Load .env from project root (script lives in tools/)
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass

DEFAULT_XLSX = r"C:\Users\micge\Documents\Burt Costing Model\GRP Costings 2018.xlsx"

# Cell -> option-name lookup tables for FORMULAS 2018.xls.
# Names must match the rows in skin_formulas / taping_blocks / mounting_cleats / floor_plates.
# Cell positions reflect the FORMULAS 2018.xls layout in the "Latest price
# list" folder (verified 2026-05-09). Keep in sync with
# app/excel_formula_matcher.py.
SKIN_TOTALS = {
    "D13": "450CSM-450",
    "D25": "600CSM-450",
    "D37": "900CSM-450-0",
    "D49": "INTERNAL LAMINATION",
    "D59": "FINAL COAT",
}
TAPING_TOTALS = {
    "F11": "TAPING BLOCK 200MM",
    "F24": "TAPING BLOCK 250MM",
    "F37": "CHEAP TAPPING BLOCK 200MM",
    "F47": "CHEAP TAPPING BLOCK 250MM",
    "F61": "TIMBER ONLY TAPING BLOCK 200MM",
    "F74": "TIMBER ONLY TAPING BLOCK 250MM",
}
CLEAT_TOTALS = {
    "F10": "TOP MOUNTING CLEAT",
    "F22": "BOTTOM MOUNTING CLEAT",
    "F39": "SPRING MOUNTING CLEAT",
}
FLOOR_TOTALS = {
    "F13": "2MM 3CR12",
    "F24": "3MM ALU BUFFER PLATE",
    "F35": "D-RUBBER",
    "F43": "CORNER GUSSETS",
}

# Map FORMULAS 2018 sheet name -> (db_table, fk_column, lookup_dict, extra_columns_to_set)
SHEET_MAP = {
    "FORMULA SKINS":   ("skin_formulas",   "skin_formula_id",   SKIN_TOTALS,   {"is_formula_skin": 1, "skin_formula_region": "standard"}),
    "TAPING BLOCKS":   ("taping_blocks",   "taping_block_id",   TAPING_TOTALS, {}),
    "MOUNTING CLEATS": ("mounting_cleats", "mounting_cleat_id", CLEAT_TOTALS,  {}),
    "SRD FLOOR PLATE": ("floor_plates",    "floor_plate_id",    FLOOR_TOTALS,  {}),
}

PRICE_COL = "G"   # PRICE column on costing sheets
ITEM_COL  = "A"   # item-description column

# Sheet name -> trailer_type name (when they don't match exactly).
SHEET_ALIASES = {
    "4.9 & UP FREEZER BODY 3": "4.9 & UP FREEZER BODY 2",
}

FORMULA_RE = re.compile(r"'\[(\d+)\]([^']+)'!\$?([A-Z]+)\$?(\d+)")


def discover_external_links(xlsx_path: str) -> dict[int, str]:
    """Return {externalLink_id: target_filename} from the xlsx package rels."""
    out: dict[int, str] = {}
    with zipfile.ZipFile(xlsx_path) as z:
        for name in z.namelist():
            m = re.match(r"xl/externalLinks/_rels/externalLink(\d+)\.xml\.rels", name)
            if not m:
                continue
            link_id = int(m.group(1))
            root = ET.fromstring(z.read(name))
            for rel in root:
                target = rel.attrib.get("Target", "")
                # URL-decode
                from urllib.parse import unquote
                out[link_id] = unquote(target.split("/")[-1])
    return out


def load_db_lookups(engine) -> dict[str, dict[str, int]]:
    """Return {table: {name_upper: id}} for all formula tables."""
    out: dict[str, dict[str, int]] = {}
    with engine.connect() as c:
        for table in ("skin_formulas", "taping_blocks", "mounting_cleats", "floor_plates"):
            rows = c.execute(text(f"SELECT id, name FROM {table}")).fetchall()
            out[table] = {r[1].strip().upper(): r[0] for r in rows}
    return out


def load_trailer_types(engine) -> dict[str, int]:
    """Return {normalized_name: id} for active (non-deleted) trailer types."""
    out: dict[str, int] = {}
    with engine.connect() as c:
        rows = c.execute(text("SELECT id, name FROM trailer_types WHERE name NOT LIKE '%[deleted-%'")).fetchall()
        for tid, nm in rows:
            out[normalize(nm)] = tid
    return out


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).upper()


def load_bom_for_type(engine, type_id: int) -> list[dict]:
    """Return BOM rows for a type with material name."""
    sql = text("""
        SELECT b.id, b.skin_formula_id, b.taping_block_id, b.mounting_cleat_id, b.floor_plate_id,
               m.name as material_name
        FROM bill_of_materials b
        LEFT JOIN materials m ON m.id = b.material_id
        WHERE b.trailer_type_id = :tid
    """)
    with engine.connect() as c:
        rows = c.execute(sql, {"tid": type_id}).fetchall()
    return [dict(id=r[0], skin_formula_id=r[1], taping_block_id=r[2],
                 mounting_cleat_id=r[3], floor_plate_id=r[4],
                 material_name=r[5]) for r in rows]


def best_bom_match(item_name: str, bom_rows: list[dict]) -> dict | None:
    """Match an Excel item description to a BOM row by material name (normalized contains)."""
    target = normalize(item_name)
    if not target:
        return None
    # Exact match first
    for r in bom_rows:
        if normalize(r["material_name"]) == target:
            return r
    # Contains either way (ignore very short tokens)
    if len(target) >= 6:
        candidates = [r for r in bom_rows if r["material_name"] and (target in normalize(r["material_name"]) or normalize(r["material_name"]) in target)]
        if len(candidates) == 1:
            return candidates[0]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--db-url", default=None, help="SQLAlchemy URL (default: from DATABASE_URL env or sqlite:///costing.db)")
    ap.add_argument("--apply", action="store_true", help="Write changes (default: dry run)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite rows that already have a link set")
    ap.add_argument("--type", help="Restrict to a single trailer type name")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"Source workbook not found: {args.xlsx}")

    db_url = args.db_url or os.environ.get("DATABASE_URL") or "sqlite:///costing.db"
    print(f"DB:     {db_url}")
    print(f"Source: {args.xlsx}")
    print(f"Mode:   {'APPLY (writing)' if args.apply else 'DRY RUN'}{'  +overwrite' if args.overwrite else ''}")
    print()

    engine = create_engine(db_url)
    db_lookups = load_db_lookups(engine)
    types = load_trailer_types(engine)

    ext_links = discover_external_links(args.xlsx)
    formulas_link_ids = {lid for lid, fn in ext_links.items() if "FORMULAS 2018" in fn.upper()}
    if not formulas_link_ids:
        sys.exit("Workbook has no external link to FORMULAS 2018.xls")
    print(f"FORMULAS 2018 external link IDs: {sorted(formulas_link_ids)}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=False)

    proposals: list[dict] = []   # {type_id, type_name, bom_id, table, fk, value, item, sheet, cell, current_id, status}
    unmatched_items: list[str] = []
    unknown_cells: list[str] = []
    skipped_sheets: list[str] = []

    for sheet_name in wb.sheetnames:
        alias = SHEET_ALIASES.get(sheet_name.strip())
        norm = normalize(alias or sheet_name)
        if args.type and normalize(args.type) != norm:
            continue
        type_id = types.get(norm)
        if not type_id:
            skipped_sheets.append(sheet_name)
            continue
        sh = wb[sheet_name]
        bom_rows = load_bom_for_type(engine, type_id)

        for row in sh.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                if cell.column_letter != PRICE_COL:
                    continue
                m = FORMULA_RE.search(v)
                if not m:
                    continue
                link_id, ref_sheet, col_letters, row_num = int(m.group(1)), m.group(2), m.group(3), int(m.group(4))
                if link_id not in formulas_link_ids:
                    continue
                ref_sheet_u = ref_sheet.upper()
                if ref_sheet_u not in SHEET_MAP:
                    continue
                table, fk_col, totals_map, extras = SHEET_MAP[ref_sheet_u]
                cell_key = f"{col_letters}{row_num}"
                opt_name = totals_map.get(cell_key)
                if not opt_name:
                    unknown_cells.append(f"{sheet_name} {cell.coordinate}: {ref_sheet}!{cell_key}")
                    continue
                opt_id = db_lookups[table].get(opt_name.upper())
                if not opt_id:
                    unknown_cells.append(f"{sheet_name} {cell.coordinate}: {ref_sheet}!{cell_key} -> '{opt_name}' (no DB row)")
                    continue
                # find item name in col A on the same row
                item_name = sh.cell(cell.row, 1).value or ""
                bom = best_bom_match(item_name, bom_rows)
                if not bom:
                    unmatched_items.append(f"{sheet_name} row {cell.row}: '{item_name}' -> {opt_name}")
                    continue
                current = bom.get(fk_col)
                if current is not None and not args.overwrite:
                    status = f"SKIP (already={current})" if current != opt_id else "OK (already set)"
                else:
                    status = "SET" if current is None else f"OVERWRITE ({current}->{opt_id})"
                proposals.append(dict(
                    type_name=sheet_name, type_id=type_id, bom_id=bom["id"],
                    table=table, fk=fk_col, opt_id=opt_id, opt_name=opt_name,
                    item=item_name, cell=cell.coordinate,
                    ref=f"{ref_sheet}!{cell_key}", current=current, status=status,
                    extras=extras,
                ))

    # --- Report ---
    print()
    print("=" * 100)
    print(f"PROPOSALS: {len(proposals)} | unmatched items: {len(unmatched_items)} | unknown cells: {len(unknown_cells)} | skipped sheets: {len(skipped_sheets)}")
    print("=" * 100)
    by_type: dict[str, list[dict]] = {}
    for p in proposals:
        by_type.setdefault(p["type_name"], []).append(p)
    for tname, items in sorted(by_type.items()):
        print(f"\n--- {tname} ---")
        for p in items:
            print(f"  {p['cell']:>6} {p['ref']:<28} -> {p['table']:<16} '{p['opt_name']:<30}' | item='{p['item']}' | bom_id={p['bom_id']} | {p['status']}")

    if unmatched_items:
        print("\n--- UNMATCHED ITEMS (no BOM row found) ---")
        for u in unmatched_items[:50]:
            print(f"  {u}")
        if len(unmatched_items) > 50:
            print(f"  ... +{len(unmatched_items)-50} more")

    if unknown_cells:
        print("\n--- UNKNOWN FORMULA CELLS (not in lookup tables) ---")
        for u in unknown_cells[:30]:
            print(f"  {u}")
        if len(unknown_cells) > 30:
            print(f"  ... +{len(unknown_cells)-30} more")

    if skipped_sheets:
        print(f"\n--- Sheets skipped (no matching trailer_type): {len(skipped_sheets)} ---")
        for s in skipped_sheets:
            print(f"  {s}")

    # --- Apply ---
    to_write = [p for p in proposals if p["status"].startswith(("SET", "OVERWRITE"))]
    print(f"\nWould write {len(to_write)} updates.")
    if not args.apply:
        print("(dry run — no changes made; pass --apply to write)")
        return

    with engine.begin() as c:
        for p in to_write:
            sets = [f"{p['fk']} = :v"]
            params = {"v": p["opt_id"], "id": p["bom_id"]}
            for k, val in p["extras"].items():
                sets.append(f"{k} = :{k}")
                params[k] = val
            sql = text(f"UPDATE bill_of_materials SET {', '.join(sets)} WHERE id = :id")
            c.execute(sql, params)
    print(f"Wrote {len(to_write)} updates.")


if __name__ == "__main__":
    main()
