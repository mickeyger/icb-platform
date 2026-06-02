"""
tools/test_cell_resolver.py

Run the new app.excel_cell_resolver against the real RIGID DRY FREIGHT
sheet to verify each pattern (literal / chain / external / IF gate /
aggregate / cycle / depth) classifies correctly.

This is also the "what to look for" reference — every CASE below names
an expected pattern + ResolvedKind, and the script reports PASS/FAIL.

Usage:
    python tools/test_cell_resolver.py
    python tools/test_cell_resolver.py --xlsx "C:/path/to/GRP Costings 2018.xlsx"
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Windows cp1252 console can't print box-drawing / em-dash chars otherwise
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass

# Make `app` importable when running from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from app.excel_cell_resolver import (
    Resolver, ResolvedKind, discover_external_link_ids,
)

DEFAULT_XLSX = (
    r"C:\Users\micge\Documents\Burt Costing Model"
    r"\Latest price list\GRP Costings 2018.xlsx"
)
DEFAULT_SHEET = "RIGID DRY FREIGHT"


# Each case: (cell, expected_kind, optional human-readable note describing
# what should be true about the result besides kind).
CASES: list[tuple[str, ResolvedKind, str]] = [
    # ── Literals ─────────────────────────────────────────────────────────
    ("C3",  ResolvedKind.LITERAL,
     "length=7.2 (raw number)"),
    ("C4",  ResolvedKind.LITERAL,
     "width=2.6"),
    ("G4",  ResolvedKind.LITERAL,
     "default margin = 0.05"),

    # ── Chain that ends at an aggregate ──────────────────────────────────
    ("H3",  ResolvedKind.AGGREGATE,
     "=J317 -> =SUM(J50:J316) — single hop through the chain ends at SUM"),

    # ── Chain through intra-sheet refs ───────────────────────────────────
    ("G54", ResolvedKind.EXTERNAL_OTHER,
     "=G39 -> ='[1]RESINS + ADHESIVES'!... — single hop, ends at PRICE 2017"),
    ("G55", ResolvedKind.EXTERNAL_OTHER,
     "=G40 -> another PRICE 2017 ref"),

    # ── Direct external FORMULAS 2018 reference ──────────────────────────
    ("G80", ResolvedKind.EXTERNAL_FORMULAS_2018,
     "='[2]SRD FLOOR PLATE'!$F$24"),
    ("G82", ResolvedKind.EXTERNAL_FORMULAS_2018,
     "='[2]SRD FLOOR PLATE'!$F$35 — D-RUBBER"),

    # ── Direct external (PRICE 2017 / [1]) ───────────────────────────────
    ("G69", ResolvedKind.EXTERNAL_OTHER,
     "=[1]FITTINGS!$D$6 — CSLB hinges, PRICE 2017"),
    ("G78", ResolvedKind.EXTERNAL_OTHER,
     "=[1]RIVETS!$C$5 — long rivets"),

    # ── IF gate (per-line conditional inclusion) ────────────────────────
    ("I56", ResolvedKind.IF_CONDITION,
     "=IF(C21=\"Y\",1,0) — gates 20MM EPS by BAKERY BODY option (row 21)"),
    ("I57", ResolvedKind.IF_CONDITION,
     "=IF(C26=\"Y\",1,0) — gates 80DV URETHANE by 80 DV COMPOSITE PANELS"),
    ("I65", ResolvedKind.IF_CONDITION,
     "=IF(C8=\"Y\",1,0) — DRD section master toggle (row 8 = DRD option)"),
    ("I84", ResolvedKind.IF_CONDITION,
     "=IF(C8=\"Y\",1,0) — DRD DOOR FITTINGS section, also gated by DRD"),
    ("I99", ResolvedKind.IF_CONDITION,
     "=IF(C9=\"Y\",1,0) — SRD section, gated by SRD option (row 9)"),
    ("I266", ResolvedKind.IF_CONDITION,
     "=IF(C8 =\"Y\",1,0) — note the leading space typo, must still parse"),

    # ── Aggregate function ──────────────────────────────────────────────
    ("H50",  ResolvedKind.AGGREGATE,
     "=SUM(H39:H49) — FRONT section total"),
    ("H65",  ResolvedKind.AGGREGATE,
     "=SUM(H54:H64) — DRD section total"),
    ("J317", ResolvedKind.AGGREGATE,
     "=SUM(J50:J316) — grand total"),

    # ── Multi-term arithmetic expression (no single ref, no aggregate) ──
    ("H4",  ResolvedKind.EXPRESSION,
     "=H3*G4 — length × default margin, two-term product"),
    ("H56", ResolvedKind.EXPRESSION,
     "=G56*F56*I56 — line total = price × area × gate (3 terms)"),

    # ── Empty cell ──────────────────────────────────────────────────────
    ("A2",  ResolvedKind.EMPTY, "no value at all"),
]


# ── Runner ─────────────────────────────────────────────────────────────────

def green(s: str) -> str: return f"\033[92m{s}\033[0m"
def red(s: str)   -> str: return f"\033[91m{s}\033[0m"
def dim(s: str)   -> str: return f"\033[90m{s}\033[0m"
def bold(s: str)  -> str: return f"\033[1m{s}\033[0m"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--verbose", "-v", action="store_true",
                    help="print all cases, not just failures")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        print(red(f"Workbook not found: {args.xlsx}"))
        return 2

    print(f"Workbook: {args.xlsx}")
    print(f"Sheet:    {args.sheet}")
    print()

    wb       = openpyxl.load_workbook(args.xlsx, data_only=False)
    wb_data  = openpyxl.load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(red(f"Sheet not found in workbook: {args.sheet!r}"))
        return 2

    ext_links = discover_external_link_ids(args.xlsx)
    print(f"External links discovered: {ext_links}")
    print()

    r = Resolver(wb[args.sheet], wb_data[args.sheet], ext_links)

    passed, failed = 0, 0
    for cell, expected_kind, note in CASES:
        res = r.resolve(cell)
        ok = res.kind == expected_kind
        if ok: passed += 1
        else:  failed += 1
        if not ok or args.verbose:
            tag = green("PASS") if ok else red("FAIL")
            line = f"  {tag}  {cell:>5}  expected={expected_kind.value:<22} got={res.kind.value}"
            print(line)
            print(dim(f"         note: {note}"))
            extras = []
            if res.value is not None:           extras.append(f"value={res.value}")
            if res.ref_sheet:                   extras.append(f"ref={res.ref_sheet}!{res.ref_cell} (link {res.ref_link_id})")
            if res.if_test_cell:                extras.append(f"if {res.if_test_cell}={res.if_expected!r}")
            if res.fallback_value is not None:  extras.append(f"fallback={res.fallback_value}")
            if res.reason:                      extras.append(f"reason={res.reason}")
            if extras:
                print(dim(f"         {' · '.join(extras)}"))
            if args.verbose and res.chain:
                print(dim(f"         chain: {' -> '.join(res.chain)}"))

    # ── Cycle + depth simulation (synthetic, no real workbook cell) ─────
    # Use a small in-memory workbook to verify the guards trigger.
    print()
    print(bold("Cycle + depth guards"))
    cycle_ok, depth_ok = _test_guards()
    print(f"  cycle detection: {'PASS' if cycle_ok else 'FAIL'}")
    print(f"  depth limit:     {'PASS' if depth_ok else 'FAIL'}")
    passed += int(cycle_ok) + int(depth_ok)
    failed += int(not cycle_ok) + int(not depth_ok)

    print()
    print(bold(f"  {passed} passed · {failed} failed · {len(CASES)+2} total"))
    return 0 if failed == 0 else 1


def _test_guards() -> tuple[bool, bool]:
    """Build a tiny in-memory workbook and verify the resolver guards
    trigger UNRESOLVED for cycles and overly-deep chains."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    # Cycle: A1 -> A2 -> A1
    ws["A1"] = "=A2"
    ws["A2"] = "=A1"
    # Depth chain: B1 -> B2 -> ... -> B40 -> 0
    for i in range(1, 41):
        ws.cell(i, 2).value = f"=B{i+1}" if i < 40 else 0

    # The resolver expects a data_only twin — fake one with the same cells
    # (data_only would read None for these formulas in a fresh in-memory
    # workbook; not relevant to the guards being tested).
    r = Resolver(ws, ws, ext_links={})
    cyc = r.resolve("A1")
    deep = r.resolve("B1")
    cycle_ok = (cyc.kind == ResolvedKind.UNRESOLVED and "cycle" in (cyc.reason or ""))
    depth_ok = (deep.kind == ResolvedKind.UNRESOLVED and "depth" in (deep.reason or ""))
    return cycle_ok, depth_ok


if __name__ == "__main__":
    sys.exit(main())
