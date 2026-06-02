"""
Import Body Variables (numeric metric values) from GRP Costings 2018.xlsx.

For each body-template-named sheet, finds the "BODY OPTIONS" header in
column B, then reads each row's name (col A) and numeric value (col C).
Updates bill_of_materials.variable_value for matching is_body_option rows.

Default: dry run. Use --apply to write. By default, rows that already have
a non-null variable_value are preserved (--overwrite forces).

Usage:
    python tools/import_body_variables.py
    python tools/import_body_variables.py --apply
    python tools/import_body_variables.py --db-url mysql+pymysql://USER:PASS@HOST/DB --apply
"""
from __future__ import annotations
import argparse, os, re, sys
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass

DEFAULT_XLSX = r"C:\Users\micge\Documents\Burt Costing Model\GRP Costings 2018.xlsx"

# Sheet name -> trailer_type name (when they don't match exactly).
SHEET_ALIASES = {
    "4.9 & UP FREEZER BODY 3": "4.9 & UP FREEZER BODY 2",
}


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).upper()


def load_trailer_types(engine) -> dict[str, int]:
    out: dict[str, int] = {}
    with engine.connect() as c:
        for tid, nm in c.execute(text(
            "SELECT id, name FROM trailer_types WHERE name NOT LIKE '%[deleted-%'"
        )).fetchall():
            out[normalize(nm)] = tid
    return out


def load_body_option_rows(engine, type_id: int) -> list[dict]:
    sql = text("""
        SELECT b.id, b.variable_value, m.name AS material_name
        FROM bill_of_materials b
        LEFT JOIN materials m ON m.id = b.material_id
        WHERE b.trailer_type_id = :tid AND b.is_body_option = 1
    """)
    with engine.connect() as c:
        rows = c.execute(sql, {"tid": type_id}).fetchall()
    return [dict(id=r[0], variable_value=r[1], material_name=r[2]) for r in rows]


def find_body_options_block(sh) -> tuple[int, int] | None:
    """Return (start_row, end_row) of the BODY OPTIONS block, or None."""
    header_row = None
    for r in range(1, sh.max_row + 1):
        b = sh.cell(r, 2).value
        a = sh.cell(r, 1).value
        if (b and "BODY OPTIONS" in str(b).upper()) or (a and "BODY OPTIONS" in str(a).upper() and not (b or sh.cell(r, 3).value)):
            header_row = r
            break
    if header_row is None:
        return None
    # Walk down from header until 3 consecutive blank rows in col A
    blanks = 0
    last = header_row
    for r in range(header_row + 1, sh.max_row + 1):
        a = sh.cell(r, 1).value
        if not a:
            blanks += 1
            if blanks >= 3:
                break
        else:
            blanks = 0
            last = r
    return (header_row + 1, last)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--db-url", default=None)
    ap.add_argument("--apply", action="store_true", help="Write changes")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite rows with existing variable_value")
    ap.add_argument("--type", help="Restrict to a single trailer type")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"Source workbook not found: {args.xlsx}")

    db_url = args.db_url or os.environ.get("DATABASE_URL") or "sqlite:///costing.db"
    print(f"DB:     {db_url}")
    print(f"Source: {args.xlsx}")
    print(f"Mode:   {'APPLY (writing)' if args.apply else 'DRY RUN'}{'  +overwrite' if args.overwrite else ''}")
    print()

    engine = create_engine(db_url)
    types = load_trailer_types(engine)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    proposals: list[dict] = []
    unmatched: list[str] = []
    skipped_sheets: list[str] = []
    no_block: list[str] = []

    for sheet_name in wb.sheetnames:
        alias = SHEET_ALIASES.get(sheet_name.strip())
        norm = normalize(alias or sheet_name)
        if args.type and normalize(args.type) != norm:
            continue
        type_id = types.get(norm)
        if not type_id:
            skipped_sheets.append(sheet_name)
            continue
        sh = wb[sheet_name]
        block = find_body_options_block(sh)
        if not block:
            no_block.append(sheet_name)
            continue
        start, end = block
        bom_rows = load_body_option_rows(engine, type_id)

        for r in range(start, end + 1):
            name = sh.cell(r, 1).value
            value = sh.cell(r, 3).value
            if not name:
                continue
            # Only treat as a variable if col C is a number
            try:
                val = float(value)
            except (TypeError, ValueError):
                continue

            target_norm = normalize(name)
            match = next((b for b in bom_rows if normalize(b["material_name"]) == target_norm), None)
            if not match:
                unmatched.append(f"{sheet_name} '{name}' = {val}")
                continue

            current = match["variable_value"]
            if current is not None and not args.overwrite:
                if abs(float(current) - val) < 1e-9:
                    status = "OK (matches)"
                else:
                    status = f"SKIP (current={current}, would set={val})"
            else:
                status = f"SET ({current} -> {val})" if current is not None else f"SET ({val})"

            proposals.append(dict(
                type_name=sheet_name, bom_id=match["id"], name=name,
                old=current, new=val, status=status,
            ))

    print("=" * 100)
    print(f"PROPOSALS: {len(proposals)} | unmatched: {len(unmatched)} | sheets w/o BODY OPTIONS block: {len(no_block)} | skipped sheets: {len(skipped_sheets)}")
    print("=" * 100)

    by_type: dict[str, list[dict]] = {}
    for p in proposals:
        by_type.setdefault(p["type_name"], []).append(p)
    for tname, items in sorted(by_type.items()):
        print(f"\n--- {tname} ---")
        for p in items:
            print(f"  {p['name']:<22} = {p['new']:<8} | bom_id={p['bom_id']:<6} | {p['status']}")

    if unmatched:
        print("\n--- UNMATCHED (no body-option BOM row found) ---")
        for u in unmatched[:50]:
            print(f"  {u}")
        if len(unmatched) > 50:
            print(f"  ... +{len(unmatched)-50} more")

    if no_block:
        print(f"\n--- Sheets with no BODY OPTIONS block: {len(no_block)} ---")
        for s in no_block:
            print(f"  {s}")

    if skipped_sheets:
        print(f"\n--- Sheets skipped (no matching trailer_type): {len(skipped_sheets)} ---")
        for s in skipped_sheets:
            print(f"  {s}")

    to_write = [p for p in proposals if p["status"].startswith("SET")]
    print(f"\nWould write {len(to_write)} updates.")
    if not args.apply:
        print("(dry run — no changes made; pass --apply to write)")
        return

    with engine.begin() as c:
        for p in to_write:
            c.execute(text("UPDATE bill_of_materials SET variable_value = :v WHERE id = :id"),
                      {"v": p["new"], "id": p["bom_id"]})
    print(f"Wrote {len(to_write)} updates.")


if __name__ == "__main__":
    main()
