"""
Update materials.price_per_unit from a new PRICE 2017 MARCH.xlsx file.

Strategy
--------
1. Scan GRP Costings 2018.xlsx (formula mode) to find every cross-workbook
   reference like ='[PRICE 2017 MARCH.xlsx]EPS'!$C$12.
   Collects ALL unique (sheet, col, row) cells referenced — not just one per
   material name. This is critical for sheets like EPS where different rows
   represent different thicknesses with different prices.

2. For every referenced cell, read the old and new PRICE file values.
   Build a lookup: old_value -> new_value (cells where the price changed).

3. Match DB materials purely by price_per_unit value — if a material's current
   price matches an old PRICE-file cell value (within R0.01), it gets updated
   to the new value. No name matching required.

4. Write price_history records and stamp last_bulk_update_at so the existing
   admin "Undo Last Bulk Update" button works out of the box.

Usage
-----
    # Dry run (no changes written):
    python tools/update_prices_from_price_list.py

    # Apply locally (SQLite):
    python tools/update_prices_from_price_list.py --apply

    # Apply to production MySQL:
    DATABASE_URL="mysql+pymysql://user:pass@host/dbname" ^
        python tools/update_prices_from_price_list.py --apply

    # Undo the last batch:
    python tools/update_prices_from_price_list.py --rollback

File paths
----------
Override with env vars if files live elsewhere:
    GRP_PATH       path to GRP Costings 2018.xlsx  (formulas read from here)
    OLD_PRICE_PATH path to the OLD PRICE 2017 MARCH.xlsx  (baseline)
    NEW_PRICE_PATH path to the NEW PRICE 2017 MARCH.xlsx  (updated prices)
"""

import argparse
import os
import re
from datetime import datetime, timezone
from urllib.parse import unquote

import openpyxl
import sqlalchemy as sa

# ---------------------------------------------------------------------------
# File paths (env-overridable)
# ---------------------------------------------------------------------------
GRP_PATH = os.environ.get(
    "GRP_PATH",
    r"C:\Users\micge\Documents\Burt Costing Model\GRP Costings 2018.xlsx",
)
OLD_PRICE_PATH = os.environ.get(
    "OLD_PRICE_PATH",
    r"C:\Users\micge\Documents\Burt Costing Model\PRICE 2017 MARCH.xlsx",
)
NEW_PRICE_PATH = os.environ.get(
    "NEW_PRICE_PATH",
    r"C:\Users\micge\Documents\Burt Costing Model\Latest price list\PRICE 2017 MARCH.xlsx",
)
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./costing.db")

BATCH_NOTE = "PRICE 2017 MARCH update"
PRICE_MATCH_TOLERANCE = 0.01  # rand — DB price must match old PRICE cell within R0.01

# ---------------------------------------------------------------------------
# Regex helpers
# ---------------------------------------------------------------------------
RE_NAMED   = re.compile(
    r"\[PRICE 2017 MARCH\.xlsx\]'?([^'\]!]+?)'?\!\$?([A-Z]+)\$?(\d+)",
    re.IGNORECASE,
)
RE_INDEXED = re.compile(r"\[(\d+)\]'?([^'\]!]+?)'?\!\$?([A-Z]+)\$?(\d+)")


def _col_to_idx(col: str) -> int:
    """Column letter(s) to 1-based openpyxl column index."""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _parse_ref(formula, ext_files):
    """Return (sheet, col_letter, row_int) if formula references PRICE 2017 MARCH, else None."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    m = RE_NAMED.search(formula)
    if m:
        return m.group(1).strip(), m.group(2), int(m.group(3))
    m = RE_INDEXED.search(formula)
    if m:
        idx = int(m.group(1)) - 1
        if 0 <= idx < len(ext_files) and "PRICE 2017 MARCH" in ext_files[idx].upper():
            return m.group(2).strip(), m.group(3), int(m.group(4))
    return None


def _read_cell(wb, sheet, col_letter, row):
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    ci = _col_to_idx(col_letter)
    for r in ws.iter_rows(min_row=row, max_row=row, min_col=ci, max_col=ci, values_only=True):
        return r[0] if r else None
    return None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Step 1 — collect every unique PRICE cell referenced in GRP Costings
# ---------------------------------------------------------------------------
def collect_price_cell_refs(grp_path):
    print(f"Reading GRP formulas: {grp_path}")
    wb = openpyxl.load_workbook(grp_path, data_only=False)

    ext_files = []
    for el in getattr(wb, "_external_links", []) or []:
        try:
            ext_files.append(unquote(el.file_link.Target))
        except Exception:
            ext_files.append("")

    # Set of unique (sheet, col_letter, row) tuples
    refs = set()
    for ws in wb.worksheets:
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                ref = _parse_ref(cell.value, ext_files)
                if ref:
                    refs.add(ref)

    wb.close()
    print(f"  Found {len(refs)} unique PRICE file cell references")
    return refs


# ---------------------------------------------------------------------------
# Step 2 — build old_value -> new_value map from the two PRICE files
# ---------------------------------------------------------------------------
def build_cell_price_map(refs, old_path, new_path):
    print(f"Reading old prices: {old_path}")
    print(f"Reading new prices: {new_path}")

    wb_old = openpyxl.load_workbook(old_path, data_only=True)
    wb_new = openpyxl.load_workbook(new_path, data_only=True)

    # old_value -> new_value  (only cells where price changed)
    price_map   = {}
    # Track collisions: two different cells with the same old value but different new values
    collisions  = {}
    unchanged   = 0
    na_skipped  = 0

    for sh, col, row in sorted(refs):
        old_v = _to_float(_read_cell(wb_old, sh, col, row))
        new_v = _to_float(_read_cell(wb_new, sh, col, row))

        if old_v is None or new_v is None:
            na_skipped += 1
            continue
        if abs(old_v - new_v) < 0.001:
            unchanged += 1
            continue

        key = round(old_v, 4)
        if key in price_map:
            if abs(price_map[key] - new_v) > 0.001:
                # Same old price, different new price — flag it
                collisions[key] = (price_map[key], new_v, sh, col, row)
        else:
            price_map[key] = round(new_v, 4)

    wb_old.close()
    wb_new.close()

    print(f"  Changed price cells: {len(price_map)}")
    print(f"  Unchanged: {unchanged}  |  N/A skipped: {na_skipped}")

    if collisions:
        print(f"\n  WARNING: {len(collisions)} old-price value(s) map to different new prices")
        print(f"  (two referenced cells had the same old value but different new values)")
        for old_k, (new1, new2, sh, col, row) in collisions.items():
            print(f"    old={old_k}  -> new1={new1}  vs new2={new2}  (conflict at [{sh}]{col}{row})")

    return price_map  # {old_price_rounded: new_price_rounded}


# ---------------------------------------------------------------------------
# Step 3 — match DB materials by price value and build update list
# ---------------------------------------------------------------------------
def build_db_updates(price_map, conn):
    rows = conn.execute(
        sa.text("SELECT id, name, price_per_unit FROM materials")
    ).fetchall()

    updates  = []  # (material_id, name, old_price, new_price)
    skipped  = []  # (material_id, name, price) — no matching price cell found

    for mid, mname, mprice in rows:
        if mprice is None:
            continue
        key = round(mprice, 4)

        # Look for an exact match first, then scan within tolerance
        new_price = price_map.get(key)
        if new_price is None:
            # Tolerance scan
            for old_k, new_k in price_map.items():
                if abs(mprice - old_k) <= PRICE_MATCH_TOLERANCE:
                    new_price = new_k
                    break

        if new_price is not None and abs(new_price - mprice) > 0.0001:
            updates.append((mid, mname, mprice, new_price))
        elif new_price is None and mprice > 0:
            skipped.append((mid, mname, mprice))

    # Deduplicate by material_id (shouldn't happen but be safe)
    seen, deduped = {}, []
    for row in updates:
        if row[0] not in seen:
            seen[row[0]] = True
            deduped.append(row)

    return deduped, skipped


# ---------------------------------------------------------------------------
# Step 4a — print and optionally apply updates
# ---------------------------------------------------------------------------
def apply_updates(updates, skipped, conn, dry_run):
    now  = datetime.now(timezone.utc)
    note = f"{BATCH_NOTE} -- {now.strftime('%Y-%m-%d')}"

    prefix = "DRY RUN -- Would update" if dry_run else "Updating"
    print(f"\n{prefix} {len(updates)} material records")
    print(f"  Batch note: {note}")
    print()
    print(f"  {'id':>5}  {'Material name':<52}  {'Old price':>12}  {'New price':>12}  {'Chg':>6}")
    print(f"  {'-'*5}  {'-'*52}  {'-'*12}  {'-'*12}  {'-'*6}")

    for mid, mname, old_p, new_p in sorted(updates, key=lambda x: x[2]):
        pct = (new_p / old_p - 1) * 100 if old_p else 0
        print(f"  {mid:>5}  {mname:<52}  {old_p:>12.4f}  {new_p:>12.4f}  {pct:>+5.1f}%")

    if skipped:
        # Only print materials that had a non-zero price and weren't matched
        # Filter out zero-price and very-low-price items (likely placeholders)
        notable = [(mid, mname, p) for mid, mname, p in skipped if p >= 1.0]
        if notable:
            print(f"\n  NOTE: {len(notable)} materials with price >= R1 had no matching PRICE cell")
            print(f"  (their prices were not set from a PRICE 2017 MARCH formula — manual if needed)")
            print(f"  {'id':>5}  {'Material name':<52}  {'Price':>12}")
            print(f"  {'-'*5}  {'-'*52}  {'-'*12}")
            for mid, mname, p in sorted(notable, key=lambda x: x[2])[:30]:
                print(f"  {mid:>5}  {mname:<52}  {p:>12.4f}")
            if len(notable) > 30:
                print(f"  ... ({len(notable) - 30} more not shown)")

    if dry_run:
        print("\n  (Dry run -- nothing written. Re-run with --apply to commit.)")
        return

    for mid, mname, old_p, new_p in updates:
        conn.execute(
            sa.text("""
                INSERT INTO price_history
                    (material_id, old_price, new_price, changed_date, changed_by)
                VALUES (:mid, :old, :new, :now, :who)
            """),
            {"mid": mid, "old": old_p, "new": new_p, "now": now,
             "who": "bulk_price_update script"},
        )
        conn.execute(
            sa.text("""
                UPDATE materials
                SET price_per_unit        = :new,
                    last_updated          = :now,
                    last_bulk_update_at   = :now,
                    last_bulk_update_note = :note
                WHERE id = :mid
            """),
            {"new": new_p, "now": now, "note": note, "mid": mid},
        )

    print(f"\n  Done. {len(updates)} records updated.")
    print(f"  To undo: run with --rollback, or use Admin > Materials > Undo Last Bulk Update.")


# ---------------------------------------------------------------------------
# Step 4b — rollback last batch
# ---------------------------------------------------------------------------
def rollback(conn):
    row = conn.execute(
        sa.text("""
            SELECT last_bulk_update_at FROM materials
            WHERE last_bulk_update_at IS NOT NULL
            ORDER BY last_bulk_update_at DESC LIMIT 1
        """)
    ).fetchone()
    if not row:
        print("No bulk update batch found -- nothing to roll back.")
        return

    batch_at = row[0]
    print(f"Rolling back batch: {batch_at}")

    mats = conn.execute(
        sa.text("SELECT id, name, price_per_unit FROM materials WHERE last_bulk_update_at = :ts"),
        {"ts": batch_at},
    ).fetchall()

    reverted = 0
    for mid, mname, cur_price in mats:
        ph = conn.execute(
            sa.text("""
                SELECT id, old_price FROM price_history
                WHERE material_id = :mid
                ORDER BY changed_date DESC LIMIT 1
            """),
            {"mid": mid},
        ).fetchone()
        if not ph:
            print(f"  WARN: no price_history for id={mid} ({mname}), skipping")
            continue
        ph_id, old_price = ph
        conn.execute(
            sa.text("""
                UPDATE materials
                SET price_per_unit        = :old,
                    last_bulk_update_at   = NULL,
                    last_bulk_update_note = NULL
                WHERE id = :mid
            """),
            {"old": old_price, "mid": mid},
        )
        conn.execute(sa.text("DELETE FROM price_history WHERE id = :id"), {"id": ph_id})
        print(f"  Restored id={mid:4d}  {mname:<50}  {cur_price:.4f} -> {old_price:.4f}")
        reverted += 1

    print(f"\n  Rolled back {reverted} records.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Update material prices from new PRICE 2017 MARCH.xlsx"
    )
    parser.add_argument("--apply",    action="store_true",
                        help="Write changes to DB (default is dry run)")
    parser.add_argument("--rollback", action="store_true",
                        help="Undo the last bulk update batch")
    args = parser.parse_args()

    connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
    engine = sa.create_engine(DATABASE_URL, connect_args=connect_args)

    with engine.begin() as conn:
        if args.rollback:
            rollback(conn)
            return

        refs      = collect_price_cell_refs(GRP_PATH)
        price_map = build_cell_price_map(refs, OLD_PRICE_PATH, NEW_PRICE_PATH)
        updates, skipped = build_db_updates(price_map, conn)

        apply_updates(updates, skipped, conn, dry_run=not args.apply)


if __name__ == "__main__":
    main()
